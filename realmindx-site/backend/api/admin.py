import csv
import io
import json
import mimetypes
import re
import secrets
import zipfile
from datetime import date, datetime, timedelta, timezone
from html import escape
from pathlib import Path
from threading import Thread
from types import SimpleNamespace
from uuid import uuid4

from flask import Blueprint, Response, current_app, g, jsonify, request, send_file
from flask_login import current_user, login_required
from sqlalchemy import and_, func, literal, or_
from sqlalchemy.orm import joinedload, selectinload
from werkzeug.utils import secure_filename

from ..analytics import build_analytics_dashboard, build_product_detail, parse_analytics_range
from ..audit_labels import AREA_LABELS, readable_audit_action, readable_audit_summary
from ..communications import (
    CommunicationResult,
    generate_batch_id,
    mask_destination,
    record_attempt,
    resolve_communication_mode,
)
from ..profile_completion import account_status as canonical_account_status
from ..book_requests import (
    BookRequestError,
    mark_available,
    request_json,
    send_available_notification,
)
from ..contacts import (
    MARKETING_ACTIVE,
    TRANSACTIONAL_ONLY,
    UNSUBSCRIBED,
    contact_json,
    newsletter_subscriber_json,
    normalize_contact_email,
    remove_contact_source,
    upsert_contact,
    upsert_contact_safely,
)
from ..default_content import (
    DEFAULT_DONATION_SLIDES,
    DEFAULT_HOME_HERO_SLIDES,
    DEFAULT_PARTNERS,
    DEFAULT_PEOPLE,
    DEFAULT_SERVICES,
    DEFAULT_SITE_COPY,
    DEFAULT_TESTIMONIALS,
)
from ..email_service import (
    EmailAttachment,
    OutboundEmail,
    absolute_app_url,
    app_email_shell,
    bookshop_email_shell,
    bookshop_order_summary_table,
    send_email,
)
from ..extensions import db
from ..image_variants import ensure_product_image_variants, reset_product_image_variants
from ..rich_text import contains_rich_html, sanitize_rich_html
from ..delivery_locations import format_location_aliases
from ..delivery_service import (
    DeliveryError,
    OTP_OVERRIDE_REASONS,
    actor_from_user,
    assign_order_to_company,
    cancel_delivery,
    create_company,
    create_company_user,
    reset_portal_password,
    resend_delivery_otp,
    send_portal_access_notification,
    staff_delivery_contact_warning,
    staff_override_otp,
)
from ..location_data import parse_location_ids
from ..order_status import normalize_order_status
from ..profile_completion import teacher_profile_completion
from ..sms_service import normalise_phone
from ..teacher_ids import ensure_application_id, generate_teacher_id, is_valid_teacher_id
from ..bookshop_search import canonical_taxonomy_value


def _normalize_taxonomy(value, taxonomy):
    if not value:
        return value
    try:
        return canonical_taxonomy_value(taxonomy, value)
    except Exception:
        return value


from ..models import (
    AnalyticsEvent,
    AuditLog,
    BookRequest,
    CommunicationAttempt,
    Contact,
    ContactSource,
    ContactMessage,
    CartInvoice,
    BookshopPaymentIntent,
    DeliveryCompany,
    DeliveryCompanyUser,
    DeliveryEvent,
    DeliveryOtp,
    DeliveryRider,
    DeliverySettlementBatch,
    DeliverySettlementLine,
    DeliveryZone,
    Flyer,
    Job,
    JobAlertPreference,
    JobApplication,
    NewsletterSubscriber,
    NewsletterCampaign,
    NewsletterCampaignRecipient,
    News,
    Order,
    OrderDelivery,
    OrderItem,
    OrderReview,
    Permission,
    PromoCode,
    Product,
    ProductCategory,
    ProductReview,
    Resource,
    Role,
    GalleryItem,
    TeacherPlacement,
    User,
    UserProfile,
    SiteSetting,
    UploadedFile,
)
from ..promo_affiliates import (
    record_completed_order_promo_usage,
    send_promo_usage_notification,
    usage_snapshot,
)
from ..security import generate_temporary_password, admin_or_staff_required, admin_required, permission_required
from ..settlement_service import (
    SettlementError, apply_adjustment, batch_json, line_json, mark_settled,
    raise_dispute, resolve_dispute,
)
from ..serializers import (
    delivery_company_json,
    delivery_company_user_json,
    delivery_event_json,
    delivery_json,
    delivery_rider_json,
    delivery_zone_json,
    job_json,
    order_json,
    order_review_json,
    product_json,
    user_json,
)
from ..upload_utils import save_upload

admin_bp = Blueprint("admin", __name__)


@admin_bp.after_request
def capture_unlogged_management_change(response):
    """Guarantee that successful admin/staff changes leave an audit record."""
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"} or response.status_code >= 400:
        return response
    if getattr(g, "audit_logged", False) or not current_user.is_authenticated:
        return response
    endpoint = str(request.endpoint or "management_action").split(".")[-1]
    try:
        from ..audit import audit as _audit
        _audit(endpoint, "management_portal", details={"page": request.path, "method": request.method})
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Could not record the audit entry for %s", request.path)
    return response


def _book_request_error(exc):
    db.session.rollback()
    return jsonify(error=exc.message, code=exc.code), exc.status_code


@admin_bp.get("/book-requests")
@login_required
@permission_required("bookRequests.view")
def list_book_requests():
    page = max(1, request.args.get("page", 1, type=int))
    page_size = min(100, max(5, request.args.get("page_size", 10, type=int)))
    status = (request.args.get("status") or "").strip().lower()
    search = (request.args.get("q") or "").strip()
    query = BookRequest.query
    if status in {"pending", "available"}:
        query = query.filter(BookRequest.status == status)
    if search:
        needle = f"%{search}%"
        query = query.filter(or_(
            BookRequest.reference.ilike(needle),
            BookRequest.requested_title.ilike(needle),
            BookRequest.customer_name.ilike(needle),
            BookRequest.email.ilike(needle),
            BookRequest.phone.ilike(needle),
        ))
    pagination = query.order_by(BookRequest.created_at.desc()).paginate(page=page, per_page=page_size, error_out=False)
    pending_count = BookRequest.query.filter_by(status="pending").count()
    return jsonify(
        items=[request_json(row, include_private=True) for row in pagination.items],
        page=pagination.page,
        pages=pagination.pages,
        total=pagination.total,
        pending_count=pending_count,
    )


@admin_bp.get("/book-requests/<int:request_id>")
@login_required
@permission_required("bookRequests.view")
def get_book_request(request_id):
    row = db.get_or_404(BookRequest, request_id)
    payload = request_json(row, include_private=True)
    event_labels = {
        "book_request_created": "Request received",
        "book_request_duplicate_reused": "Existing request reused",
        "book_request_acknowledgement": "Acknowledgement sent",
        "book_request_marked_available": "Book marked available",
        "book_request_availability_notification": "Availability notice sent",
        "book_request_notification_retried": "Availability notice retried",
    }
    def request_event_label(event):
        details = event.details or {}
        if event.action == "book_request_acknowledgement" and "sent" not in {details.get("email"), details.get("sms")}:
            return "Acknowledgement could not be sent"
        if event.action in {"book_request_availability_notification", "book_request_notification_retried"} and "sent" not in {details.get("email"), details.get("sms")}:
            return "Availability notice could not be sent"
        return event_labels.get(event.action, event.action.replace("_", " ").capitalize())
    events = AuditLog.query.filter_by(entity_type="book_request", entity_id=str(row.id)).order_by(AuditLog.created_at.desc()).all()
    payload["history"] = [{
        "id": event.id,
        "action": request_event_label(event),
        "details": event.details or {},
        "created_at": event.created_at.isoformat(),
    } for event in events]
    return jsonify(request=payload)


@admin_bp.post("/book-requests/<int:request_id>/available")
@login_required
@permission_required("bookRequests.manage")
def make_book_request_available(request_id):
    payload = request.get_json(silent=True) or {}
    row = db.get_or_404(BookRequest, request_id)
    try:
        notification = mark_available(row, payload.get("product_url"), current_user.id)
        db.session.commit()
    except BookRequestError as exc:
        return _book_request_error(exc)
    return jsonify(request=request_json(row, include_private=True), notification=notification)


@admin_bp.post("/book-requests/<int:request_id>/retry-notification")
@login_required
@permission_required("bookRequests.manage")
def retry_book_request_notification(request_id):
    row = db.get_or_404(BookRequest, request_id)
    if row.status != "available":
        return jsonify(error="Only available requests can be notified."), 409
    try:
        notification = send_available_notification(row, retry=True)
        db.session.commit()
    except BookRequestError as exc:
        return _book_request_error(exc)
    return jsonify(request=request_json(row, include_private=True), notification=notification)


def slugify(value):
    slug = re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")
    return slug or "item"


def _send_internal_account_access_email(user, role_name, temporary_password):
    login_path = "/staff/login" if role_name == "staff" else "/admin/login"
    login_url = f"{current_app.config['BASE_URL'].rstrip('/')}{login_path}"
    role_label = "Staff" if role_name == "staff" else "Admin"
    body = (
        f"<p>Hello {escape(user.full_name or role_label)},</p>"
        f"<p>Your RealMindX {role_label.lower()} account has been created or reset.</p>"
        '<div style="background:#f5f8fc;border:1px solid #d9e3f0;border-radius:10px;padding:18px 20px;margin:20px 0;">'
        f"<p style=\"margin:0 0 8px;\"><strong>Email:</strong> {escape(user.email)}</p>"
        f"<p style=\"margin:0;\"><strong>Temporary password:</strong> {escape(temporary_password)}</p>"
        "</div>"
        "<p>You must change this password immediately after your first sign-in.</p>"
    )
    try:
        result = send_email(
            OutboundEmail(
                to=user.email,
                subject=f"Your RealMindX {role_label} account",
                html=app_email_shell(
                    f"{role_label} account ready",
                    body,
                    cta_label=f"Open {role_label} sign in",
                    cta_url=login_url,
                    eyebrow="RealMindX Secure Access",
                    preheader=f"Your RealMindX {role_label.lower()} account is ready.",
                ),
                text=(
                    f"Your RealMindX {role_label} account is ready. Email: {user.email}. "
                    f"Temporary password: {temporary_password}. Login: {login_url}. "
                    "Change the password on first sign-in."
                ),
            ),
            purpose="security",
            recipient_user_id=user.id,
            template_name="internal_account_access",
        )
        return result.status
    except Exception as exc:
        current_app.logger.warning(
            "Access email failed for %s (role=%s, error=%s)",
            mask_destination("email", user.email),
            role_name,
            type(exc).__name__,
        )
        return "failed"


def placed_order_query():
    return Order.query.filter(
        or_(
            Order.payment_method.is_(None),
            Order.payment_method != "online",
            Order.payment_status == "paid",
        )
    )


def log_action(action, entity_type=None, entity_id=None, metadata=None):
    """Backward-compat wrapper — delegates to the shared audit() helper."""
    from ..audit import audit as _audit
    _audit(action, entity_type, entity_id, metadata)


def _run_in_background(task_name, func, *args):
    app = current_app._get_current_object()

    def runner():
        with app.app_context():
            try:
                func(*args)
            except Exception:
                app.logger.exception("%s failed.", task_name)

    Thread(target=runner, daemon=True).start()


def _order_contact_snapshot(order):
    return SimpleNamespace(
        invoice_id=order.invoice_id,
        order_reference=order.order_reference,
        customer_name=order.customer_name,
        email=order.email,
        phone=order.phone,
        delivery_method=order.delivery_method,
        delivery_zone_name=order.delivery_zone_name,
        location=order.location,
        payment_method=order.payment_method,
        payment_status=order.payment_status,
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        subtotal_amount=order.subtotal_amount,
        bulk_discount_amount=order.bulk_discount_amount,
        promo_code=order.promo_code,
        promo_applies_to=order.promo_applies_to,
        promo_discount_amount=order.promo_discount_amount,
        delivery_fee=order.delivery_fee,
        total_amount=order.total_amount,
        notes=order.notes,
        items=[
            SimpleNamespace(
                product_id=item.product_id,
                product_name=item.product_name,
                unit_price=item.unit_price,
                quantity=item.quantity,
            )
            for item in order.items
        ],
    )


def _collection_item_id(item):
    if not isinstance(item, dict):
        return ""
    return str(item.get("id") or item.get("key") or "")


def _deleted_collection_setting_key(key):
    return f"{key}__deleted_ids"


def _collection_deleted_ids(key):
    row = SiteSetting.query.filter_by(key=_deleted_collection_setting_key(key)).first()
    if row and isinstance(row.value, list):
        return {str(item) for item in row.value if str(item)}
    return set()


def _save_collection_deleted_ids(key, deleted_ids):
    row = SiteSetting.query.filter_by(key=_deleted_collection_setting_key(key)).first()
    if not row:
        row = SiteSetting(key=_deleted_collection_setting_key(key), public=False)
        db.session.add(row)
    row.value = sorted({str(item) for item in deleted_ids if str(item)})
    row.public = False
    return row


def _mark_default_item_deleted(key, default_items, item_id):
    default_ids = {_collection_item_id(item) for item in default_items}
    item_id = str(item_id)
    if item_id not in default_ids:
        return
    deleted_ids = _collection_deleted_ids(key)
    deleted_ids.add(item_id)
    _save_collection_deleted_ids(key, deleted_ids)


def _restore_deleted_item_id(key, item_id):
    item_id = str(item_id)
    deleted_ids = _collection_deleted_ids(key)
    if item_id not in deleted_ids:
        return
    deleted_ids.remove(item_id)
    _save_collection_deleted_ids(key, deleted_ids)


def _active_default_items(key, default_items):
    deleted_ids = _collection_deleted_ids(key)
    return [
        item for item in default_items
        if _collection_item_id(item) not in deleted_ids
    ]


def _collection_setting(key, default_items):
    active_defaults = _active_default_items(key, default_items)
    row = SiteSetting.query.filter_by(key=key).first()
    if row and isinstance(row.value, list):
        deleted_ids = _collection_deleted_ids(key)
        current_items = [
            item for item in row.value
            if isinstance(item, dict) and _collection_item_id(item) not in deleted_ids
        ]
        existing_ids = {_collection_item_id(item) for item in current_items}
        missing_defaults = [
            item for item in active_defaults
            if _collection_item_id(item) not in existing_ids
        ]
        if len(current_items) != len(row.value):
            row.value = current_items
        if missing_defaults:
            row.value = [*row.value, *missing_defaults]
            row.public = True
            db.session.add(row)
            db.session.flush()
        return row, row.value
    row = row or SiteSetting(key=key, public=True)
    row.value = active_defaults
    row.public = True
    db.session.add(row)
    db.session.flush()
    return row, row.value


def _save_collection_setting(key, items, public=True):
    row = SiteSetting.query.filter_by(key=key).first()
    if not row:
        row = SiteSetting(key=key)
        db.session.add(row)
    row.value = items
    row.public = public
    return row


def _next_sort_order(items):
    return max([int(item.get("sort_order") or 0) for item in items] or [0]) + 1


def _item_id(payload):
    return (
        payload.get("id")
        or payload.get("key")
        or slugify(payload.get("label") or payload.get("title") or payload.get("name"))
    )


def _admin_collection_items(setting_key, default_items):
    _, items = _collection_setting(setting_key, default_items)
    return items


def _upload_public_url(uploaded_file):
    if not uploaded_file:
        return None
    return f"/uploads/{uploaded_file.visibility}/{uploaded_file.category}/{uploaded_file.stored_filename}"


def _uploaded_file_payload(uploaded_file):
    if not uploaded_file:
        return None
    return {
        "id": uploaded_file.id,
        "name": uploaded_file.original_filename,
        "url": _upload_public_url(uploaded_file),
        "category": uploaded_file.category,
        "visibility": uploaded_file.visibility,
    }


def _enrich_service_media(items):
    file_ids = {
        int(item.get("image_file_id"))
        for item in items
        if str(item.get("image_file_id") or "").isdigit()
    }
    file_ids.update(
        int(item.get("detail_image_file_id"))
        for item in items
        if str(item.get("detail_image_file_id") or "").isdigit()
    )
    files = {
        row.id: row
        for row in UploadedFile.query.filter(UploadedFile.id.in_(file_ids)).all()
    } if file_ids else {}
    enriched = []
    for item in items:
        row = dict(item)
        file_id = row.get("image_file_id")
        if str(file_id or "").isdigit():
            row["image_url"] = _upload_public_url(files.get(int(file_id))) or row.get("image_url")
        detail_file_id = row.get("detail_image_file_id")
        if str(detail_file_id or "").isdigit():
            row["detail_image_url"] = _upload_public_url(files.get(int(detail_file_id))) or row.get("detail_image_url")
        enriched.append(row)
    return enriched


def _permissions_payload():
    return [
        {"id": row.id, "key": row.key, "description": row.description}
        for row in Permission.query.order_by(Permission.key.asc()).all()
    ]


def _staff_payload(user):
    row = user_json(user)
    row["status"] = "active" if user.is_active else "inactive"
    return row


def _admin_payload(user):
    row = user_json(user)
    row["status"] = "active" if user.is_active else "inactive"
    return row


def _can_view_dashboard_metric(permission_key):
    role_name = current_user.role.name if current_user.role else ""
    if role_name == "admin":
        return True
    return current_user.has_permission(permission_key)


def _boolish(value):
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "published", "active", "featured"}


def _decimalish(value, default=0):
    raw = str(value if value is not None else "").replace("GH", "").replace("GHS", "").replace(",", "").strip()
    raw = raw.replace("GHC", "").replace("GH", "").strip()
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def _dateish(value):
    if value in {None, ""}:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip()[:10])
    except ValueError as exc:
        raise ValueError("Use a valid YYYY-MM-DD date.") from exc


def _intish(value, default=0):
    try:
        return int(float(str(value if value is not None else "").strip()))
    except (TypeError, ValueError):
        return default


def _split_tags(value):
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value or "").replace(";", ",").split(",") if item.strip()]


def _unique_product_slug(base_slug, product_id=None):
    base = slugify(base_slug)
    slug = base
    counter = 2
    while True:
        query = Product.query.filter_by(slug=slug)
        if product_id:
            query = query.filter(Product.id != product_id)
        if not query.first():
            return slug
        slug = f"{base}-{counter}"
        counter += 1


def _ensure_category(name):
    name = (name or "").strip()
    if not name:
        return None
    category = ProductCategory.query.filter(func.lower(ProductCategory.name) == name.lower()).first()
    if category:
        return category
    category = ProductCategory(name=name, slug=_unique_category_slug(name), is_active=True)
    db.session.add(category)
    db.session.flush()
    return category


def _category_from_payload(payload):
    typed_name = (payload.get("category_name") or payload.get("category") or "").strip()
    if typed_name:
        return _ensure_category(typed_name)
    category_id = payload.get("category_id")
    return ProductCategory.query.filter_by(id=category_id).first() if category_id else None


def _unique_category_slug(name):
    base = slugify(name)
    candidate = base
    counter = 2
    while ProductCategory.query.filter_by(slug=candidate).first():
        candidate = f"{base}-{counter}"
        counter += 1
    return candidate


def _image_entry_key(basename):
    """Normalise an image filename to the stored original_filename matching key.

    Stored original filenames are passed through secure_filename, so a ZIP entry
    such as "My Cover.jpg" or "my%20cover.jpg" must reduce to the same key
    ("my_cover.jpg") before matching products. Only the basename is used, so
    nested ZIP folders never affect matching, and the result is lowercased so
    matching is case-insensitive.
    """
    from urllib.parse import unquote
    return secure_filename(unquote(basename or "")).lower() or ""


def _save_imported_image(filename, data, owner_id):
    safe_name = secure_filename(filename or "")
    if not safe_name or "." not in safe_name:
        return None
    extension = safe_name.rsplit(".", 1)[1].lower()
    allowed = current_app.config["ALLOWED_UPLOAD_EXTENSIONS"].get("images", set())
    if extension not in allowed:
        return None
    stored_name = f"{uuid4().hex}.{extension}"
    root = Path(current_app.config["UPLOAD_FOLDER"]) / "public" / "images"
    root.mkdir(parents=True, exist_ok=True)
    target = root / stored_name
    target.write_bytes(data)
    uploaded = UploadedFile(
        owner_id=owner_id,
        original_filename=safe_name,
        stored_filename=stored_name,
        storage_path=str(target),
        mime_type=mimetypes.guess_type(safe_name)[0] or "application/octet-stream",
        size_bytes=len(data),
        category="images",
        visibility="public",
    )
    db.session.add(uploaded)
    db.session.flush()
    return uploaded.id, target


def _read_catalog_rows(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("Upload a CSV or XLSX catalogue file.")
    name = file_storage.filename.lower()
    if name.endswith(".csv"):
        text = file_storage.stream.read().decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX import requires openpyxl. Install backend requirements first.") from exc
        workbook = load_workbook(file_storage.stream, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(cell not in (None, "") for cell in row)
        ]
    raise ValueError("Only CSV and XLSX catalogue files are supported.")


def _save_imported_images(file_storage, owner_id, only_filenames=None):
    if not file_storage or not file_storage.filename:
        return {}, [], []
    if not file_storage.filename.lower().endswith(".zip"):
        raise ValueError("Batch images must be uploaded as a ZIP file.")
    file_storage.stream.seek(0, 2)
    archive_bytes = file_storage.stream.tell()
    file_storage.stream.seek(0)
    if archive_bytes > 100 * 1024 * 1024:
        raise ValueError("Image ZIP must be 100 MB or smaller.")
    if not zipfile.is_zipfile(file_storage.stream):
        raise ValueError("Choose a valid ZIP archive for the product images.")
    file_storage.stream.seek(0)
    total_uncompressed = 0
    max_entries = 500
    max_image_bytes = 12 * 1024 * 1024
    max_uncompressed_bytes = 512 * 1024 * 1024
    only_keys = {_image_entry_key(name) for name in (only_filenames or [])} if only_filenames else None
    saved_images = {}
    saved_paths = []
    saved_ids = []
    with zipfile.ZipFile(file_storage.stream) as archive:
        entries = [
            entry
            for entry in archive.infolist()
            if not entry.is_dir() and "__MACOSX" not in entry.filename
        ]
        if len(entries) > max_entries:
            raise ValueError(f"Image ZIP contains too many files. Use at most {max_entries}.")
        for entry in entries:
            basename = Path(entry.filename).name
            if not basename:
                continue
            if entry.file_size > max_image_bytes:
                raise ValueError(f"{basename} is larger than the 12 MB per-image limit.")
            total_uncompressed += entry.file_size
            if total_uncompressed > max_uncompressed_bytes:
                raise ValueError("Image ZIP expands beyond the 512 MB safety limit.")
        for entry in entries:
            basename = Path(entry.filename).name
            if not basename:
                continue
            key = _image_entry_key(basename)
            if not key:
                continue
            if only_keys is not None and key not in only_keys:
                continue
            if key in saved_images:
                # Duplicate filename inside the ZIP: the first entry wins.
                continue
            saved = _save_imported_image(basename, archive.read(entry), owner_id)
            if not saved:
                continue
            file_id, target = saved
            saved_images[key] = file_id
            saved_paths.append(target)
            saved_ids.append(file_id)
    return saved_images, saved_paths, saved_ids


PRODUCT_IMPORT_FIELDS = (
    {"key": "name", "label": "Product name", "required": True, "aliases": ("name", "product_name", "title")},
    {"key": "slug", "label": "Slug", "aliases": ("slug",)},
    {"key": "category", "label": "Category", "aliases": ("category", "product_category")},
    {"key": "price", "label": "Price", "aliases": ("price", "unit_price")},
    {"key": "old_price", "label": "Old price", "aliases": ("old_price", "compare_at_price")},
    {
        "key": "short_description",
        "label": "Short description",
        "aliases": ("short_description", "description", "summary"),
    },
    {
        "key": "full_description",
        "label": "Full description",
        "aliases": ("full_description", "details", "body"),
    },
    {"key": "stock_status", "label": "Stock status", "aliases": ("stock_status", "stock")},
    {
        "key": "quantity_available",
        "label": "Quantity",
        "aliases": ("quantity_available", "quantity", "qty"),
    },
    {"key": "subject", "label": "Subject", "aliases": ("subject",)},
    {"key": "level", "label": "Level", "aliases": ("level", "class", "grade")},
    {
        "key": "curriculum",
        "label": "Curriculum",
        "aliases": ("curriculum", "curriculum_name", "syllabus"),
    },
    {"key": "author", "label": "Author", "aliases": ("author", "writer")},
    {"key": "publisher", "label": "Publisher", "aliases": ("publisher", "publishing_house")},
    {"key": "product_type", "label": "Product type", "aliases": ("product_type", "item_type", "type")},
    {"key": "delivery_note", "label": "Delivery note", "aliases": ("delivery_note",)},
    {"key": "tags", "label": "Tags", "aliases": ("tags", "badges")},
    {"key": "featured", "label": "Featured", "aliases": ("featured", "is_featured")},
    {"key": "source", "label": "Source", "aliases": ("source", "supplier", "vendor")},
    {
        "key": "image_filename",
        "label": "Image filename",
        "aliases": ("image_filename", "image", "cover_filename", "cover_image"),
    },
)


def _normalise_import_header(value):
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _import_headers(rows):
    if not rows:
        return []
    return [str(header or "").strip() for header in rows[0].keys() if str(header or "").strip()]


def _suggest_import_mapping(headers):
    normalised = {_normalise_import_header(header): header for header in headers}
    mapping = {}
    for field in PRODUCT_IMPORT_FIELDS:
        for alias in field["aliases"]:
            source = normalised.get(_normalise_import_header(alias))
            if source:
                mapping[field["key"]] = source
                break
    return mapping


def _parse_import_mapping(raw_mapping, headers):
    if not raw_mapping:
        return {}
    try:
        mapping = json.loads(raw_mapping)
    except (TypeError, json.JSONDecodeError) as exc:
        raise ValueError("Column mapping is invalid. Review the matched columns and try again.") from exc
    if not isinstance(mapping, dict):
        raise ValueError("Column mapping must be an object.")
    allowed_fields = {field["key"] for field in PRODUCT_IMPORT_FIELDS}
    allowed_headers = set(headers)
    clean_mapping = {}
    for field, source in mapping.items():
        if field not in allowed_fields or not source:
            continue
        if source not in allowed_headers:
            raise ValueError(f'The mapped column "{source}" is no longer present in the catalogue.')
        clean_mapping[field] = source
    return clean_mapping


def _apply_import_mapping(row, mapping):
    if not mapping:
        return row
    return {field: row.get(source) for field, source in mapping.items()}


def _json_safe_import_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _normalise_import_row(row):
    lookup = {
        str(key or "").strip().lower().replace(" ", "_").replace("-", "_"): value
        for key, value in (row or {}).items()
    }

    def pick(*names):
        for name in names:
            value = lookup.get(name)
            if value not in (None, ""):
                return value
        return None

    name = str(pick("name", "product_name", "title") or "").strip()
    if not name:
        return None
    stock = str(pick("stock_status", "stock") or "in_stock").strip().lower().replace(" ", "_")
    if stock in {"in", "available"}:
        stock = "in_stock"
    if stock in {"low"}:
        stock = "low_stock"
    if stock in {"out", "unavailable"}:
        stock = "out_of_stock"

    return {
        "name": name,
        "slug": slugify(str(pick("slug") or name)),
        "category": str(pick("category", "product_category") or "").strip(),
        "price": _decimalish(pick("price", "unit_price")),
        "old_price": _decimalish(pick("old_price", "compare_at_price"), None),
        "short_description": pick("short_description", "description", "summary"),
        "full_description": pick("full_description", "details", "body"),
        "stock_status": stock or "in_stock",
        "quantity_available": (
            int(_decimalish(pick("quantity_available", "quantity", "qty")))
            if pick("quantity_available", "quantity", "qty") not in (None, "")
            else None
        ),
        "subject": pick("subject"),
        "level": pick("level", "class", "grade"),
        "curriculum": pick("curriculum", "curriculum_name", "syllabus"),
        "author": pick("author", "writer"),
        "publisher": pick("publisher", "publishing_house"),
        "product_type": pick("product_type", "type"),
        "delivery_note": pick("delivery_note"),
        "tags": _split_tags(pick("tags", "badges")),
        "featured": _boolish(pick("featured")),
        "source": pick("source", "supplier", "vendor"),
        "image_filename": str(pick("image_filename", "image", "cover_image") or "").strip(),
    }

def _ticket_reference(message_id):
    return f"RMX-{int(message_id):06d}"


def _create_admin_collection_item(setting_key, default_items, payload, entity_type):
    items = _admin_collection_items(setting_key, default_items)
    item_id = str(_item_id(payload))
    if any(str(item.get("id")) == item_id for item in items):
        return None, (jsonify(error="An item with this ID already exists."), 409)
    row = {
        **payload,
        "id": item_id,
        "status": payload.get("status") or "published",
        "sort_order": payload.get("sort_order") or _next_sort_order(items),
    }
    items.append(row)
    _restore_deleted_item_id(setting_key, item_id)
    _save_collection_setting(setting_key, items)
    log_action(f"create_{entity_type}", entity_type, item_id)
    db.session.commit()
    return row, None


def _update_admin_collection_item(setting_key, default_items, item_id, payload, entity_type):
    items = _admin_collection_items(setting_key, default_items)
    found = None
    next_items = []
    for item in items:
        if str(item.get("id")) == str(item_id):
            found = {**item, **payload, "id": str(item_id)}
            next_items.append(found)
        else:
            next_items.append(item)
    if not found:
        return None, (jsonify(error="Item not found."), 404)
    _save_collection_setting(setting_key, next_items)
    log_action(f"update_{entity_type}", entity_type, item_id)
    db.session.commit()
    return found, None


def _delete_admin_collection_item(setting_key, default_items, item_id, entity_type):
    items = _admin_collection_items(setting_key, default_items)
    next_items = [item for item in items if str(item.get("id")) != str(item_id)]
    if len(next_items) == len(items):
        return jsonify(error="Item not found."), 404
    _mark_default_item_deleted(setting_key, default_items, item_id)
    _save_collection_setting(setting_key, next_items)
    log_action(f"delete_{entity_type}", entity_type, item_id)
    db.session.commit()
    return jsonify(message="Item deleted.")


def _matches_job_alert(job, preference, user=None):
    # preference.subject is a comma-joined multi-select value (a teacher can
    # now save several teaching subjects, e.g. "Mathematics, Physics,
    # Chemistry"). Match if ANY one of those subjects appears in the job's
    # subject — a straight whole-string containment check would require the
    # entire joined list to appear verbatim inside a single-subject job
    # posting, which would never happen and would silently stop all subject
    # -based alerts for every multi-subject teacher.
    def values(raw):
        return {
            re.sub(r"[^a-z0-9]+", " ", item.lower()).strip()
            for item in (raw or "").split(",")
            if item.strip()
        }

    def aliases(value):
        normalised = re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()
        alias_map = {
            "english": "english language",
            "jhs": "junior high lower secondary",
            "shs": "senior high upper secondary",
            "primary": "upper primary",
            "full time": "full time",
            "part time": "part time",
        }
        return alias_map.get(normalised, normalised)

    pref_subjects = {aliases(item) for item in values(preference.subject)}
    job_subject = aliases(job.subject)
    subject_match = not job_subject or job_subject in pref_subjects
    pref_location_ids = set(parse_location_ids(preference.location_ids))
    if pref_location_ids:
        location_match = bool(job.delivery_zone_id and job.delivery_zone_id in pref_location_ids)
    else:
        pref_locations = values(preference.location)
        location_match = bool(pref_locations and aliases(job.location) in {aliases(item) for item in pref_locations})
    pref_levels = {aliases(item) for item in values(preference.preferred_level)}
    job_level = aliases(job.level)
    level_match = not job_level or job_level in pref_levels
    pref_curricula = {aliases(item) for item in values(preference.curriculum)}
    job_curriculum = aliases(job.curriculum)
    curriculum_match = not job_curriculum or job_curriculum in pref_curricula
    pref_types = {aliases(item) for item in values(preference.employment_type)}
    job_type = aliases(job.employment_type)
    type_match = not job_type or job_type in pref_types
    required_sex = (job.preferred_sex or "any").strip().lower()
    required_age = (job.preferred_age_range or "any").strip().lower()
    sex_match = required_sex == "any" or (user and (user.sex or "").lower() == required_sex)
    age_match = required_age == "any" or (user and (user.age_range or "").lower() == required_age)
    return subject_match and location_match and level_match and curriculum_match and type_match and sex_match and age_match


def dispatch_job_alerts(job):
    if job.status != "published":
        return 0
    preferences = (
        JobAlertPreference.query
        .join(User, JobAlertPreference.user_id == User.id)
        .filter(
            JobAlertPreference.alert_by_email.is_(True),
            JobAlertPreference.frequency == "instant",
            User.is_active.is_(True),
            User.is_verified.is_(True),
        )
        .all()
    )
    sent = 0
    processed_user_ids = set()
    for preference in preferences:
        user = db.session.get(User, preference.user_id)
        if (
            not user
            or user.id in processed_user_ids
            or teacher_profile_completion(user)[0] < 100
            or not _matches_job_alert(job, preference, user)
        ):
            continue
        processed_user_ids.add(user.id)
        job_url = f"{current_app.config['BASE_URL'].rstrip('/')}/jobs#{job.id}"
        try:
            result = send_email(
                OutboundEmail(
                    to=user.email,
                    from_email=current_app.config["JOBS_FROM_EMAIL"],
                    subject=f"A teaching opportunity matches your preferences: {job.title}",
                    html=app_email_shell(
                        "Good news — a teaching opportunity matches you",
                        f"<p>Hello {escape(user.first_name or 'Teacher')},</p>"
                        "<p>We found a teaching opportunity that matches all of your saved preferences.</p>"
                        f"<p><strong>{escape(job.title)}</strong><br>{escape(job.location)}</p>"
                        "<p>Take a look at the role and apply if it feels like the right next step for you. We are rooting for you!</p>",
                        "View Job & Apply",
                        job_url,
                        preheader=f"{job.title} matches your saved teaching preferences.",
                    ),
                ),
                purpose="service_reminder",
                recipient_user_id=user.id,
                template_name="job_alert_match",
            )
        except Exception:
            current_app.logger.exception("Job alert delivery failed for user %s and job %s", user.id, job.id)
            continue
        if result.status == "mocked":
            current_app.logger.info(
                "Job alert recorded in mock mode for user %s and job %s",
                user.id,
                job.id,
            )
            continue
        if result.status not in ("queued", "accepted", "sent", "delivered"):
            current_app.logger.warning(
                "Job alert was not sent for user %s and job %s (status=%s, error_code=%s)",
                user.id,
                job.id,
                result.status,
                result.error_code,
            )
            continue
        preference.last_sent_at = datetime.now(timezone.utc)
        log_action("job_alert_email_sent", "job_alert_preference", preference.id, {
            "job_id": job.id,
            "job_title": job.title,
            "user_id": user.id,
            "email": user.email,
        })
        sent += 1
    return sent


@admin_bp.get("/dashboard")
@login_required
@admin_or_staff_required
def dashboard():
    recent_jobs = Job.query.order_by(Job.created_at.desc()).limit(5).all() if _can_view_dashboard_metric("jobs.view") else []
    recent_orders = placed_order_query().order_by(Order.created_at.desc()).limit(5).all() if _can_view_dashboard_metric("orders.view") else []
    teacher_count = None
    if _can_view_dashboard_metric("teachers.view"):
        teacher_count = db.session.scalar(
            db.select(func.count(User.id))
            .join(Role, User.role_id == Role.id)
            .where(
                Role.name == "user",
                User.teacher_service_enabled.is_(True),
                User.is_active.is_(True),
                User.is_verified.is_(True),
            )
        )
    return jsonify(
        summary={
            "total_users": teacher_count,
            "total_job_applications": db.session.scalar(db.select(func.count(JobApplication.id))) if _can_view_dashboard_metric("applications.view") else None,
            "pending_applications": JobApplication.query.filter_by(status="pending").count() if _can_view_dashboard_metric("applications.view") else None,
            "new_orders": placed_order_query().filter(Order.status == "new").count() if _can_view_dashboard_metric("orders.view") else None,
            "new_contact_messages": ContactMessage.query.filter_by(status="new").count() if _can_view_dashboard_metric("messages.view") else None,
            "total_products": Product.query.count() if _can_view_dashboard_metric("products.view") else None,
            "newsletter_subscribers": NewsletterSubscriber.query.filter_by(is_active=True).count() if _can_view_dashboard_metric("newsletters.view") else None,
        },
        recent_jobs=[job_json(job) for job in recent_jobs],
        recent_orders=[order_json(order) for order in recent_orders],
    )


def _job_alert_admin_json(preference, user):
    return {
        "id": preference.id,
        "teacher_name": user.full_name if user else "",
        "email": user.email if user else "",
        "subject": preference.subject or "",
        "location": preference.location or "",
        "location_ids": preference.location_ids or "",
        "preferred_level": preference.preferred_level or "",
        "curriculum": preference.curriculum or "",
        "employment_type": preference.employment_type or "",
        "alert_by_email": preference.alert_by_email,
        "frequency": preference.frequency,
        "is_default": preference.is_default,
        "last_sent_at": preference.last_sent_at.isoformat() if preference.last_sent_at else None,
        "created_at": preference.created_at.isoformat() if preference.created_at else None,
        "updated_at": preference.updated_at.isoformat() if preference.updated_at else None,
        "status": "active" if preference.alert_by_email and user and user.is_active else "paused",
    }


@admin_bp.get("/job-alerts")
@login_required
@permission_required("alerts.view")
def list_job_alert_preferences():
    rows = (
        db.session.query(JobAlertPreference, User)
        .join(User, JobAlertPreference.user_id == User.id)
        .order_by(JobAlertPreference.updated_at.desc(), JobAlertPreference.created_at.desc())
        .limit(500)
        .all()
    )
    return jsonify(items=[_job_alert_admin_json(preference, user) for preference, user in rows])


def _csv_response(filename, rows, fieldnames):
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@admin_bp.get("/analytics/dashboard")
@login_required
@permission_required("analytics.view")
def analytics_dashboard():
    range_info = parse_analytics_range(request.args)
    return jsonify(build_analytics_dashboard(range_info))


@admin_bp.delete("/analytics/location-history")
@login_required
@admin_required
def clear_analytics_location_history():
    rows = AnalyticsEvent.query.filter(
        or_(
            AnalyticsEvent.country.isnot(None),
            AnalyticsEvent.region.isnot(None),
            AnalyticsEvent.city.isnot(None),
            AnalyticsEvent.ip_prefix.isnot(None),
        )
    ).update(
        {
            AnalyticsEvent.country: None,
            AnalyticsEvent.region: None,
            AnalyticsEvent.city: None,
            AnalyticsEvent.ip_prefix: None,
        },
        synchronize_session=False,
    )
    log_action("clear_analytics_location_history", "analytics_event", metadata={"events_cleared": rows})
    db.session.commit()
    return jsonify(message="Analytics location history cleared.", events_cleared=rows)


@admin_bp.get("/analytics/products/<int:product_id>")
@login_required
@permission_required("analytics.view")
def analytics_product_detail(product_id):
    range_info = parse_analytics_range(request.args)
    payload = build_product_detail(product_id, range_info)
    if not payload:
        return jsonify(error="Product not found."), 404
    return jsonify(payload)


@admin_bp.get("/analytics/export")
@login_required
@permission_required("analytics.export")
def analytics_export():
    report = str(request.args.get("report") or "products").strip().lower()
    range_info = parse_analytics_range(request.args)
    dashboard_payload = build_analytics_dashboard(range_info)

    if report == "top-pages":
        rows = dashboard_payload["overview"]["top_pages"]
        return _csv_response(
            "realmindx-top-pages.csv",
            rows,
            ["path", "title", "views", "unique_visitors"],
        )

    if report == "search-terms":
        rows = dashboard_payload["search"]["terms"]
        return _csv_response(
            "realmindx-search-terms.csv",
            rows,
            ["term", "searches", "with_results", "no_results", "product_views", "purchases"],
        )

    if report == "product-detail":
        product_id = request.args.get("product_id")
        payload = build_product_detail(int(product_id), range_info) if str(product_id or "").isdigit() else None
        if not payload:
            return jsonify(error="Product not found."), 404
        rows = []
        chart_maps = {
            "views": {item["date"]: item["value"] for item in payload["charts"]["views"]},
            "add_to_cart": {item["date"]: item["value"] for item in payload["charts"]["add_to_cart"]},
            "sales": {item["date"]: item["value"] for item in payload["charts"]["sales"]},
            "revenue": {item["date"]: item["value"] for item in payload["charts"]["revenue"]},
            "search_interest": {item["date"]: item["value"] for item in payload["charts"]["search_interest"]},
        }
        for date_key in chart_maps["views"]:
            rows.append({
                "product_name": payload["product"]["name"],
                "product_category": payload["product"]["category"],
                "product_status": payload["product"]["status"],
                "price": payload["product"]["price"],
                "stock_quantity": payload["product"]["stock_quantity"],
                "date": date_key,
                "views": chart_maps["views"].get(date_key, 0),
                "add_to_cart": chart_maps["add_to_cart"].get(date_key, 0),
                "sales": chart_maps["sales"].get(date_key, 0),
                "revenue": chart_maps["revenue"].get(date_key, 0),
                "search_interest": chart_maps["search_interest"].get(date_key, 0),
            })
        return _csv_response(
            f"realmindx-product-{payload['product']['id']}-analytics.csv",
            rows,
            ["product_name", "product_category", "product_status", "price", "stock_quantity", "date", "views", "add_to_cart", "sales", "revenue", "search_interest"],
        )

    rows = dashboard_payload["products"]["items"]
    return _csv_response(
        "realmindx-product-analytics.csv",
        rows,
        [
            "id",
            "name",
            "category",
            "status",
            "stock_status",
            "stock_quantity",
            "price",
            "views",
            "unique_visitors",
            "add_to_cart",
            "remove_from_cart",
            "purchases",
            "quantity_sold",
            "revenue",
            "conversion_rate",
            "add_to_cart_rate",
            "cart_abandonment_count",
            "wishlist_count",
            "search_impressions",
            "search_clicks",
            "unavailable_searches",
            "last_sale_at",
            "last_view_at",
            "last_add_to_cart_at",
            "top_traffic_source",
            "top_device",
            "top_location",
            "interest_delta",
            "interest_delta_pct",
            "performance_status",
        ],
    )


@admin_bp.get("/jobs")
@login_required
@permission_required("jobs.view")
def list_jobs():
    rows = Job.query.order_by(Job.created_at.desc()).all()
    return jsonify(items=[job_json(row) for row in rows])


@admin_bp.post("/jobs")
@login_required
@permission_required("jobs.create")
def create_job():
    payload = request.get_json(silent=True) or {}
    delivery_zone = db.session.get(DeliveryZone, payload.get("delivery_zone_id")) if payload.get("delivery_zone_id") else None
    if not delivery_zone or not delivery_zone.is_active or "pickup" in delivery_zone.name.lower():
        return jsonify(error="Choose a valid job location from the delivery-area list."), 400
    job = Job(
        title=payload.get("title"),
        organisation=payload.get("organisation") or payload.get("school"),
        location=delivery_zone.name,
        delivery_zone_id=delivery_zone.id,
        subject=payload.get("subject"),
        level=payload.get("level"),
        curriculum=payload.get("curriculum"),
        employment_type=payload.get("employment_type"),
        preferred_sex=payload.get("preferred_sex"),
        preferred_age_range=payload.get("preferred_age_range"),
        description=payload.get("description") or "",
        requirements=payload.get("requirements"),
        responsibilities=payload.get("responsibilities"),
        deadline=date.fromisoformat(payload["deadline"]) if payload.get("deadline") else None,
        salary_min=payload.get("salary_min"),
        salary_max=payload.get("salary_max"),
        status=payload.get("status") or "draft",
        created_by_id=current_user.id,
    )
    if not job.title or not job.location or not job.description:
        return jsonify(error="Title, location, and description are required."), 400
    db.session.add(job)
    db.session.flush()
    alerts_sent = dispatch_job_alerts(job)
    log_action("create_job", "job", job.id)
    if alerts_sent:
        log_action("send_job_alerts", "job", job.id, {"count": alerts_sent})
    db.session.commit()
    return jsonify(job=job_json(job), alerts_sent=alerts_sent), 201


@admin_bp.put("/jobs/<int:job_id>")
@login_required
@permission_required("jobs.edit")
def update_job(job_id):
    job = db.get_or_404(Job, job_id)
    payload = request.get_json(silent=True) or {}
    if "delivery_zone_id" in payload:
        delivery_zone = db.session.get(DeliveryZone, payload.get("delivery_zone_id")) if payload.get("delivery_zone_id") else None
        if not delivery_zone or not delivery_zone.is_active or "pickup" in delivery_zone.name.lower():
            return jsonify(error="Choose a valid job location from the delivery-area list."), 400
        job.delivery_zone_id = delivery_zone.id
        job.location = delivery_zone.name
    for field in ["title", "organisation", "subject", "level", "curriculum", "employment_type", "preferred_sex", "preferred_age_range", "description", "requirements", "responsibilities", "salary_min", "salary_max", "status"]:
        if field in payload:
            setattr(job, field, payload[field])
    if "deadline" in payload:
        job.deadline = date.fromisoformat(payload["deadline"]) if payload["deadline"] else None
    alerts_sent = dispatch_job_alerts(job) if payload.get("send_alerts") else 0
    log_action("update_job", "job", job.id)
    if alerts_sent:
        log_action("send_job_alerts", "job", job.id, {"count": alerts_sent})
    db.session.commit()
    return jsonify(job=job_json(job), alerts_sent=alerts_sent)


@admin_bp.delete("/jobs/<int:job_id>")
@login_required
@permission_required("jobs.delete")
def delete_job(job_id):
    job = db.get_or_404(Job, job_id)
    log_action("delete_job", "job", job.id, {"title": job.title})
    db.session.delete(job)
    db.session.commit()
    return jsonify(message="Job deleted.")


@admin_bp.get("/applications")
@login_required
@permission_required("applications.view")
def applications():
    rows = JobApplication.query.order_by(JobApplication.created_at.desc()).all()
    return jsonify(
        items=[
            {
                "id": row.id,
                "status": row.status,
                "user": user_json(row.user),
                "job": job_json(row.job),
                "created_at": row.created_at.isoformat(),
            }
            for row in rows
        ]
    )


@admin_bp.put("/applications/<int:application_id>/status")
@login_required
@permission_required("applications.edit")
def update_application_status(application_id):
    application = db.get_or_404(JobApplication, application_id)
    status = (request.get_json(silent=True) or {}).get("status")
    if status not in {"pending", "reviewed", "shortlisted", "accepted", "rejected"}:
        return jsonify(error="Invalid status."), 400
    old_status = application.status
    application.status = status
    placement = None
    if status == "accepted":
        placement = _ensure_teacher_placement(application)
    log_action("update_application_status", "job_application", application.id, {"status": status, "prev": old_status})
    db.session.commit()
    return jsonify(
        id=application.id,
        status=application.status,
        placement_id=placement.id if placement else None,
    )


def _ensure_teacher_placement(application):
    existing = TeacherPlacement.query.filter_by(application_id=application.id).first()
    if existing:
        existing.status = "accepted"
        existing.accepted_at = existing.accepted_at or datetime.now(timezone.utc)
        return existing
    job = application.job
    placement = TeacherPlacement(
        user_id=application.user_id,
        application_id=application.id,
        job_id=application.job_id,
        school_name=(job.organisation or job.location or "School placement") if job else "School placement",
        job_title=job.title if job else None,
        status="accepted",
        accepted_at=datetime.now(timezone.utc),
    )
    db.session.add(placement)
    db.session.flush()
    return placement


@admin_bp.post("/change-password")
@login_required
def admin_change_password():
    payload = request.get_json(silent=True) or {}
    current_pw = (payload.get("current_password") or "").strip()
    new_pw = (payload.get("new_password") or "").strip()
    if not current_pw or not new_pw:
        return jsonify(error="Current and new password are required."), 400
    if len(new_pw) < 8:
        return jsonify(error="New password must be at least 8 characters."), 400
    if not current_user.check_password(current_pw):
        return jsonify(error="Current password is incorrect."), 403
    current_user.set_password(new_pw)
    current_user.must_change_password = False
    log_action("change_password", "admin_user", current_user.id)
    db.session.commit()
    return jsonify(message="Password updated successfully.")


@admin_bp.get("/users")
@login_required
@permission_required("teachers.view")
def users():
    # Return verified teacher accounts. The UI defaults to Active but can also
    # filter disabled accounts, while summary counts keep internal roles clear.
    rows = (
        User.query
        .options(joinedload(User.profile))
        .join(User.role)
        .filter(
            Role.name == "user",
            User.teacher_service_enabled.is_(True),
            User.is_verified.is_(True),
        )
        .order_by(User.created_at.desc())
        .all()
    )
    items = []
    for user in rows:
        item = user_json(user)
        profile = user.profile
        item["location"] = profile.location if profile else None
        item["preferred_locations"] = profile.preferred_locations if profile else None
        item["teaching_subject"] = profile.teaching_subject if profile else None
        item["curriculum_experience"] = profile.curriculum_experience if profile else None
        items.append(item)
    active_items = [item for item in items if item.get("is_active") is not False]
    incomplete_items = [
        item for item in active_items
        if int(item.get("profile_completion") or 0) < 100 or not item.get("phone_verified")
    ]
    excluded_internal_accounts = (
        User.query
        .join(User.role)
        .filter(Role.name.in_(["admin", "staff"]))
        .count()
    )
    return jsonify(
        items=items,
        summary={
            "total_teachers": len(items),
            "active_teachers": len(active_items),
            "incomplete_profiles": len(incomplete_items),
            "disabled_accounts": len(items) - len(active_items),
            "excluded_internal_accounts": excluded_internal_accounts,
        },
    )


@admin_bp.get("/bookshop-accounts")
@login_required
@permission_required("orders.view")
def bookshop_accounts():
    rows = (
        db.session.query(User, func.count(Order.id).label("order_count"))
        .join(User.role)
        .outerjoin(Order, Order.user_id == User.id)
        .filter(Role.name == "user", User.bookshop_service_enabled.is_(True))
        .group_by(User.id)
        .order_by(User.created_at.desc())
        .all()
    )
    items = []
    for user, order_count in rows:
        item = user_json(user)
        item["order_count"] = int(order_count or 0)
        items.append(item)
    return jsonify(items=items)


REMINDER_COOLDOWN_HOURS = 24  # configurable; prevents repeated clicks from flooding a teacher
AUTOMATED_PROFILE_REMINDER_STAGES = {
    "completion": (
        ("profile_completion_reminder_24h", timedelta(hours=24)),
        ("profile_completion_reminder_7d", timedelta(days=7)),
        ("profile_completion_reminder_30d", timedelta(days=30)),
    ),
    "submission": (
        ("profile_submission_reminder_24h", timedelta(hours=24)),
        ("profile_submission_reminder_7d", timedelta(days=7)),
        ("profile_submission_reminder_30d", timedelta(days=30)),
    ),
    "revision": (
        ("profile_revision_reminder_24h", timedelta(hours=24)),
        ("profile_revision_reminder_7d", timedelta(days=7)),
        ("profile_revision_reminder_30d", timedelta(days=30)),
    ),
}
SUCCESSFUL_COMMUNICATION_STATUSES = ("queued", "accepted", "sent", "delivered", "mocked")


def _reminder_eligibility(user) -> dict:
    """Check a teacher's eligibility and select the correct reminder kind."""
    if not user:
        return {"eligible": False, "reason": "user_not_found", "status_data": None, "reminder_kind": None}
    if not getattr(user, "is_active", False):
        return {"eligible": False, "reason": "account_disabled", "status_data": None, "reminder_kind": None}
    if not getattr(user, "email", None):
        return {"eligible": False, "reason": "missing_email", "status_data": None, "reminder_kind": None}
    role_name = getattr(user.role, "name", None) if user.role else None
    if role_name not in ("user", None):
        return {"eligible": False, "reason": "not_teacher_role", "status_data": None, "reminder_kind": None}
    if not getattr(user, "teacher_service_enabled", False):
        return {"eligible": False, "reason": "teacher_service_disabled", "status_data": None, "reminder_kind": None}

    status = canonical_account_status(user)
    completion = status.get("completion_percentage", 0)
    profile_status = status.get("profile_status", "incomplete")

    # These states no longer need a completion/submission reminder. Rejected
    # applications must not receive a misleading invitation to resubmit.
    if profile_status in ("submitted", "under_review", "verified", "rejected"):
        return {
            "eligible": False,
            "reason": f"profile_{profile_status}",
            "status_data": status,
            "reminder_kind": None,
        }

    if profile_status == "revision_required":
        reminder_kind = "revision"
    elif completion >= 100:
        # A filled profile is not the same thing as a submitted profile.
        reminder_kind = "submission"
    else:
        reminder_kind = "completion"

    return {
        "eligible": True,
        "reason": None,
        "status_data": status,
        "reminder_kind": reminder_kind,
    }


def _reminder_cooldown_active(user) -> bool:
    """Check if a reminder was sent recently (within REMINDER_COOLDOWN_HOURS)."""
    from ..models import CommunicationAttempt

    hours = current_app.config.get("REMINDER_COOLDOWN_HOURS", REMINDER_COOLDOWN_HOURS)
    cutoff = datetime.now(timezone.utc).timestamp() - (hours * 3600)
    recent = CommunicationAttempt.query.filter(
        CommunicationAttempt.recipient_user_id == user.id,
        CommunicationAttempt.purpose == "service_reminder",
        CommunicationAttempt.channel == "email",
        CommunicationAttempt.status.in_(["queued", "accepted", "sent", "delivered"]),
        CommunicationAttempt.requested_at >= datetime.fromtimestamp(cutoff, tz=timezone.utc),
    ).first()
    return recent is not None


def _automated_profile_reminder_due(user, now=None):
    """Return the next due automated reminder stage, or ``None``.

    The first reminder is due 24 hours after account creation. Each later
    interval starts when the preceding automated reminder was accepted, so
    delayed jobs never compress the 7-day and 30-day waiting periods.
    """
    from ..models import CommunicationAttempt

    eligibility = _reminder_eligibility(user)
    if not eligibility["eligible"]:
        return None

    now = now or datetime.now(timezone.utc)
    reminder_kind = eligibility["reminder_kind"]
    stages = AUTOMATED_PROFILE_REMINDER_STAGES[reminder_kind]
    attempts = (
        CommunicationAttempt.query
        .filter(
            CommunicationAttempt.recipient_user_id == user.id,
            CommunicationAttempt.channel == "email",
            CommunicationAttempt.template_name.in_([stage[0] for stage in stages]),
            CommunicationAttempt.status.in_(SUCCESSFUL_COMMUNICATION_STATUSES),
        )
        .order_by(CommunicationAttempt.requested_at.asc(), CommunicationAttempt.id.asc())
        .all()
    )
    completed_templates = {attempt.template_name for attempt in attempts}
    next_index = next(
        (index for index, (template_name, _) in enumerate(stages)
         if template_name not in completed_templates),
        None,
    )
    if next_index is None:
        return None

    template_name, delay = stages[next_index]
    if next_index == 0:
        profile = getattr(user, "profile", None)
        if reminder_kind == "submission":
            anchor = getattr(profile, "updated_at", None) or user.created_at
        elif reminder_kind == "revision":
            anchor = (
                getattr(profile, "reviewed_at", None)
                or getattr(profile, "updated_at", None)
                or user.created_at
            )
        else:
            anchor = user.created_at
    else:
        preceding_template = stages[next_index - 1][0]
        preceding_attempt = next(
            (attempt for attempt in reversed(attempts) if attempt.template_name == preceding_template),
            None,
        )
        if not preceding_attempt:
            return None
        anchor = preceding_attempt.requested_at

    if not anchor:
        return None
    if anchor.tzinfo is None:
        anchor = anchor.replace(tzinfo=timezone.utc)
    due_at = anchor + delay
    if now < due_at:
        return None
    return {
        "template_name": template_name,
        "stage": next_index + 1,
        "due_at": due_at,
        "reminder_kind": reminder_kind,
    }


def _build_reminder_items(user, missing, status_data):
    """Build the list of missing items shown in the reminder email.

    Only includes items users can actually fix. Does not tell a user without
    a phone number to 'Verify your phone number'.
    """
    items = list(missing or [])
    phone = getattr(user, "phone", None)
    phone_verified = getattr(user, "phone_verified", False)
    if not phone_verified:
        if phone:
            items.append("Verify your phone number")
        else:
            items.append("Add and verify a phone number")
    return items


def _send_teacher_profile_reminder(user, *, template_name=None, enforce_cooldown=True):
    """Send one profile-reminder email. Returns a structured dict for batch aggregation."""
    from ..models import CommunicationAttempt

    eligibility = _reminder_eligibility(user)
    if not eligibility["eligible"]:
        return {
            "status": "skipped",
            "reason": eligibility["reason"],
            "user_id": user.id,
            "masked_email": mask_destination("email", user.email or ""),
        }

    if enforce_cooldown and _reminder_cooldown_active(user):
        return {
            "status": "skipped",
            "reason": "reminder_cooldown",
            "user_id": user.id,
            "masked_email": mask_destination("email", user.email or ""),
        }

    status_data = eligibility["status_data"]
    reminder_kind = eligibility["reminder_kind"]
    completion = status_data.get("completion_percentage", 0)
    missing = status_data.get("missing_requirements", [])
    reminder_items = _build_reminder_items(user, missing, status_data)

    sign_in_url = f"{current_app.config['BASE_URL'].rstrip('/')}/login"
    template_name = template_name or f"profile_{reminder_kind}_reminder_manual"

    if reminder_kind == "submission":
        subject = "Your RealMindX profile is complete — submit it for review"
        title = "Your completed profile is ready for review"
        body_html = (
            f"<p>Hello {escape(user.first_name or 'Teacher')},</p>"
            "<p>Your RealMindX teaching profile is <strong>100% complete</strong>, but it has "
            "<strong>not yet been submitted</strong> to our review team.</p>"
            "<p>Filling every profile field does not automatically start the review. Sign in, open "
            "your profile, and select <strong>Submit Profile for Review</strong>.</p>"
            "<p>Once you submit, your application will enter the teacher review queue and you will "
            "receive a confirmation email.</p>"
        )
        if not getattr(user, "phone_verified", False):
            body_html += "<p>You can also verify your phone number while signed in.</p>"
        cta_label = "Sign In and Submit for Review"
        preheader = "Your profile is complete, but review will not begin until you submit it."
        text_body = (
            "Your RealMindX teaching profile is 100% complete, but it has not yet been submitted "
            "to our review team. Filling the profile does not automatically start the review.\n\n"
            "Sign in, open your profile, and select Submit Profile for Review. Once submitted, "
            "you will receive a confirmation email.\n\n"
            f"Sign in and submit here: {sign_in_url}"
        )
    elif reminder_kind == "revision":
        profile = getattr(user, "profile", None)
        review_note = (
            getattr(profile, "review_notes", None)
            or "Review the requested corrections in your profile."
        ).strip()
        subject = "Update and resubmit your RealMindX teacher profile"
        title = "Your profile needs updates before review can continue"
        body_html = (
            f"<p>Hello {escape(user.first_name or 'Teacher')},</p>"
            "<p>Your teacher profile needs changes requested by the RealMindX review team.</p>"
            f"<p><strong>Review note:</strong> {escape(review_note)}</p>"
            "<p>Sign in, make the requested changes, complete any missing items, and then select "
            "<strong>Submit Profile for Review</strong> again. Saving your changes alone does not "
            "return the profile to the review queue.</p>"
        )
        cta_label = "Sign In to Update and Resubmit"
        preheader = "Update the requested items, then resubmit your teacher profile for review."
        text_body = (
            "Your RealMindX teacher profile needs changes before review can continue.\n\n"
            f"Review note: {review_note}\n\n"
            "Sign in, make the requested changes, and then select Submit Profile for Review again. "
            "Saving changes alone does not return the profile to the review queue.\n\n"
            f"Sign in here: {sign_in_url}"
        )
    else:
        missing_html = "".join(f"<li>{escape(item)}</li>" for item in reminder_items)
        missing_text = "\n".join(f"- {item}" for item in reminder_items)
        subject = "You are almost ready for better-matched teaching opportunities"
        title = "Complete your profile and unlock better job matches"
        body_html = (
            f"<p>Hello {escape(user.first_name or 'Teacher')},</p>"
            f"<p>You are almost there — your RealMindX teaching profile is <strong>{completion}% complete</strong>.</p>"
            "<p>Add the remaining information so we can confidently send opportunities that fit your qualifications and preferences.</p>"
            f"<p><strong>Just a little more to add:</strong></p><ul>{missing_html}</ul>"
            "<p>Finishing these items only takes a moment and gives you a better chance of seeing the right roles.</p>"
            "<p><strong>Important:</strong> Completing the fields does not automatically submit your profile. "
            "After it reaches 100%, select <strong>Submit Profile for Review</strong> so the RealMindX team can review it.</p>"
        )
        cta_label = "Sign In to Finish My Profile"
        preheader = f"Your teaching profile is {completion}% complete — finish it, then submit it for review."
        text_body = (
            "Complete your RealMindX teaching profile to receive tailored jobs.\n\n"
            f"Remaining items:\n{missing_text}\n\n"
            "Important: completing the fields does not automatically submit your profile. After it reaches "
            "100%, select Submit Profile for Review so the RealMindX team can review it.\n\n"
            f"Sign in and finish here: {sign_in_url}"
        )

    result = send_email(
        OutboundEmail(
            to=user.email,
            subject=subject,
            html=app_email_shell(
                title,
                body_html,
                cta_label,
                sign_in_url,
                preheader=preheader,
            ),
            text=text_body,
        ),
        purpose="service_reminder",
        recipient_user_id=user.id,
        template_name=template_name,
    )

    if result.status == "mocked":
        log_action("mock_teacher_profile_reminder", "user", user.id, {
            "email": user.email,
            "profile_completion": completion,
            "missing_fields": reminder_items,
            "reminder_kind": reminder_kind,
            "provider_status": result.status,
        })
        return {
            "status": "mocked",
            "profile_completion": completion,
            "missing_fields": reminder_items,
            "reminder_kind": reminder_kind,
            "user_id": user.id,
            "masked_email": mask_destination("email", user.email or ""),
        }
    if result.status in ("queued", "accepted", "sent", "delivered"):
        log_action("send_teacher_profile_reminder", "user", user.id, {
            "email": user.email,
            "profile_completion": completion,
            "missing_fields": reminder_items,
            "reminder_kind": reminder_kind,
            "provider_status": result.status,
        })
        return {
            "status": "accepted",
            "provider_status": result.status,
            "profile_completion": completion,
            "missing_fields": reminder_items,
            "reminder_kind": reminder_kind,
            "user_id": user.id,
            "masked_email": mask_destination("email", user.email or ""),
        }

    return {
        "status": "failed",
        "reason": result.error_code or "provider_error",
        "profile_completion": completion,
        "missing_fields": reminder_items,
        "reminder_kind": reminder_kind,
        "user_id": user.id,
        "masked_email": mask_destination("email", user.email or ""),
    }


def send_due_teacher_profile_completion_reminders(now=None):
    """Send every due completion, submission, or revision reminder."""
    now = now or datetime.now(timezone.utc)
    rows = (
        User.query
        .join(User.role)
        .filter(
            Role.name == "user",
            User.teacher_service_enabled.is_(True),
            User.is_active.is_(True),
        )
        .order_by(User.created_at.asc())
        .all()
    )
    counts = {"due": 0, "accepted": 0, "mocked": 0, "failed": 0, "skipped": 0}
    stages = {1: 0, 2: 0, 3: 0}
    kinds = {"completion": 0, "submission": 0, "revision": 0}
    for user in rows:
        due = _automated_profile_reminder_due(user, now=now)
        if not due:
            continue
        counts["due"] += 1
        result = _send_teacher_profile_reminder(
            user,
            template_name=due["template_name"],
            enforce_cooldown=False,
        )
        status = result.get("status", "failed")
        if status in counts:
            counts[status] += 1
        else:
            counts["failed"] += 1
        if status in ("accepted", "mocked"):
            stages[due["stage"]] += 1
            kinds[due["reminder_kind"]] += 1
    db.session.commit()
    return {**counts, "stages": stages, "kinds": kinds}


@admin_bp.post("/users/profile-reminders")
@login_required
@permission_required("teachers.edit")
def send_profile_reminders_batch():
    rows = (
        User.query
        .join(User.role)
        .filter(
            Role.name == "user",
            User.teacher_service_enabled.is_(True),
            User.is_active.is_(True),
        )
        .order_by(User.created_at.desc())
        .all()
    )
    batch_id = generate_batch_id()
    accepted = 0
    mocked = 0
    failed = []
    skipped = []
    eligible_count = 0

    for user in rows:
        result = _send_teacher_profile_reminder(user)
        if result["status"] == "accepted":
            accepted += 1
        elif result["status"] == "mocked":
            mocked += 1
        elif result["status"] == "failed":
            failed.append({"user_id": result.get("user_id"), "reason": result.get("reason", "provider_error")})
        else:
            skipped.append({"user_id": result.get("user_id"), "reason": result.get("reason", "skipped")})

    eligible_count = accepted + mocked + len(failed)

    return jsonify(
        eligible=eligible_count,
        attempted=accepted + mocked + len(failed),
        accepted=accepted,
        mocked=mocked,
        failed=len(failed),
        skipped=len(skipped),
        failures=failed[:20],
        skips=skipped[:20],
        batch_id=batch_id,
    )


@admin_bp.post("/users/<int:user_id>/profile-reminder")
@login_required
@permission_required("teachers.edit")
def send_profile_reminder(user_id):
    user = db.get_or_404(User, user_id)
    role_name = getattr(user.role, "name", None) if user.role else None

    eligibility = _reminder_eligibility(user)
    if not eligibility["eligible"]:
        reason_map = {
            "profile_submitted": "This teacher's profile has already been submitted.",
            "profile_under_review": "This teacher's profile is under review.",
            "profile_verified": "This teacher's profile is verified.",
            "profile_rejected": "This teacher's profile was rejected and cannot receive a submission reminder.",
            "account_disabled": "Enable this teacher account before sending a profile reminder.",
            "teacher_service_disabled": "Teacher services are not enabled for this account.",
            "not_teacher_role": "Profile reminders can only be sent to teacher accounts.",
            "missing_email": "This teacher does not have a valid email address.",
        }
        msg = reason_map.get(eligibility["reason"], "This teacher is not eligible for a profile reminder.")
        return jsonify(error=msg), 409

    if _reminder_cooldown_active(user):
        return jsonify(error="A profile reminder was recently sent to this teacher. Please wait before sending another."), 429

    result = _send_teacher_profile_reminder(user)
    mode = resolve_communication_mode()

    if result["status"] == "mocked":
        return jsonify(
            message="Reminder recorded in local mock mode. No real email was sent.",
            profile_completion=result.get("profile_completion"),
            mode=mode,
        )

    if result["status"] == "accepted":
        return jsonify(
            message=f"Profile reminder accepted for {user.email}.",
            profile_completion=result.get("profile_completion"),
            mode=mode,
        )

    if result["status"] == "failed":
        return jsonify(error=result.get("reason", "The reminder could not be delivered.")), 502

    return jsonify(error="Unexpected reminder status."), 500


PAYOUT_FIELDS = [
    "payout_method",
    "payout_momo_network",
    "payout_momo_number",
    "payout_bank_name",
    "payout_bank_account_name",
    "payout_bank_account_number",
    "payout_notes",
]


def _payout_payload(profile):
    if not profile:
        return {field: None for field in PAYOUT_FIELDS}
    return {field: getattr(profile, field, None) for field in PAYOUT_FIELDS}


def _apply_payout_payload(profile, payload):
    source = payload.get("payout") if isinstance(payload.get("payout"), dict) else payload
    changed = False
    for field in PAYOUT_FIELDS:
        if field not in source:
            continue
        value = source.get(field)
        if isinstance(value, str):
            value = value.strip() or None
        setattr(profile, field, value)
        changed = True
    return changed


@admin_bp.patch("/users/<int:user_id>")
@login_required
@permission_required("teachers.edit")
def update_user(user_id):
    """Update a regular-user account and admin-managed teacher payout details."""
    user = db.get_or_404(User, user_id)
    # Safety: only allow toggling regular users, not admins / staff.
    if user.role and user.role.name in ("admin", "staff"):
        return jsonify(error="Cannot modify admin or staff accounts via this endpoint."), 403
    payload = request.get_json(silent=True) or {}
    changed = False
    if "status" in payload:
        user.is_active = payload["status"] == "active"
        changed = True
        log_action("toggle_user_active", "user", user.id, {"status": payload["status"]})
    payout_source = payload.get("payout") if isinstance(payload.get("payout"), dict) else payload
    if any(field in payout_source for field in PAYOUT_FIELDS):
        profile = user.profile
        if not profile:
            profile = UserProfile(user_id=user.id)
            db.session.add(profile)
            db.session.flush()
        changed = _apply_payout_payload(profile, payload) or changed
        log_action("update_teacher_payout", "user", user.id)
    if changed:
        db.session.commit()
    return jsonify(user_json(user))


@admin_bp.delete("/users/<int:user_id>")
@login_required
@permission_required("teachers.delete")
def delete_user(user_id):
    user = db.get_or_404(User, user_id)
    if user.role and user.role.name in ("admin", "staff"):
        return jsonify(error="Cannot delete admin or staff accounts via this endpoint."), 403
    if TeacherPlacement.query.filter_by(user_id=user.id).first():
        return jsonify(error="This teacher has placement history and cannot be permanently deleted. Disable the account instead."), 409

    log_action("delete_user", "user", user.id, {"email": user.email})
    teacher_source = ContactSource.query.filter_by(source="teacher", source_record_id=str(user.id)).first()
    if teacher_source:
        remove_contact_source(teacher_source.contact, "teacher")
    _clear_user_foreign_keys(user)
    db.session.expire(user)
    db.session.delete(user)
    db.session.commit()
    return jsonify(message="Teacher account deleted.")


def _clear_user_foreign_keys(user):
    """Delete account-owned rows and anonymise nullable historical user references."""
    users_table = User.__table__
    for table in db.metadata.tables.values():
        if table is users_table or table.name == "teacher_placements":
            continue
        for foreign_key in table.foreign_keys:
            if foreign_key.column.table is not users_table:
                continue
            column = foreign_key.parent
            predicate = column == user.id
            if column.nullable:
                db.session.execute(table.update().where(predicate).values({column.name: None}))
            else:
                db.session.execute(table.delete().where(predicate))


@admin_bp.get("/users/<int:user_id>")
@login_required
@permission_required("teachers.view")
def get_user(user_id):
    """Return a single teacher's full profile data for the admin detail modal."""
    user = db.get_or_404(User, user_id)
    data = user_json(user)
    profile = getattr(user, "profile", None)
    if profile:
        def _file_payload(file_id):
            if not file_id:
                return {"url": None, "filename": None, "id": None, "mime_type": None, "size_bytes": None}
            f = db.session.get(UploadedFile, file_id)
            if not f:
                return {"url": None, "filename": None, "id": None, "mime_type": None, "size_bytes": None}
            return {
                "url": f"/uploads/{f.visibility}/{f.category}/{f.stored_filename}",
                "filename": f.original_filename,
                "id": f.id,
                "mime_type": f.mime_type,
                "size_bytes": f.size_bytes,
            }

        age = None
        if profile.date_of_birth:
            today = date.today()
            d = profile.date_of_birth
            age = today.year - d.year - ((today.month, today.day) < (d.month, d.day))
        cv_file = _file_payload(profile.cv_file_id)
        certificate_file = _file_payload(profile.certificate_file_id)

        data["profile"] = {
            "location": profile.location,
            "preferred_locations": profile.preferred_locations,
            "preferred_location_ids": profile.preferred_location_ids,
            "teaching_subject": profile.teaching_subject,
            "preferred_level": profile.preferred_level,
            "preferred_employment_type": profile.preferred_employment_type,
            "available_from": profile.available_from,
            "curriculum_experience": profile.curriculum_experience,
            "bio": profile.bio,
            "profile_picture_url": data.get("profile_picture_url"),
            "cv_url": cv_file["url"],
            "cv_filename": cv_file["filename"],
            "cv_file_id": cv_file["id"],
            "cv_mime_type": cv_file["mime_type"],
            "certificate_url": certificate_file["url"],
            "certificate_filename": certificate_file["filename"],
            "certificate_file_id": certificate_file["id"],
            "certificate_mime_type": certificate_file["mime_type"],
            "next_of_kin_name": profile.next_of_kin_name,
            "next_of_kin_phone": profile.next_of_kin_phone,
            "next_of_kin_relationship": profile.next_of_kin_relationship,
            "next_of_kin_email": profile.next_of_kin_email,
            "years_of_experience": profile.years_of_experience,
            "date_of_birth": profile.date_of_birth.isoformat() if profile.date_of_birth else None,
            "age": age,
            "payout": _payout_payload(profile),
        }
    else:
        data["profile"] = None
    preferences = (
        JobAlertPreference.query
        .filter_by(user_id=user.id)
        .order_by(JobAlertPreference.is_default.desc(), JobAlertPreference.updated_at.desc())
        .all()
    )
    data["job_alert_preferences"] = [_job_alert_admin_json(preference, user) for preference in preferences]
    applications = (
        JobApplication.query
        .options(joinedload(JobApplication.job))
        .filter_by(user_id=user.id)
        .order_by(JobApplication.updated_at.desc(), JobApplication.created_at.desc())
        .all()
    )
    data["applications"] = [
        {
            "id": application.id,
            "job_id": application.job_id,
            "status": application.status,
            "cover_note": application.cover_note,
            "created_at": application.created_at.isoformat() if application.created_at else None,
            "updated_at": application.updated_at.isoformat() if application.updated_at else None,
            "job_title": application.job.title if application.job else None,
            "organisation": application.job.organisation if application.job else None,
            "location": application.job.location if application.job else None,
            "subject": application.job.subject if application.job else None,
            "level": application.job.level if application.job else None,
            "employment_type": application.job.employment_type if application.job else None,
        }
        for application in applications
    ]
    placements = (
        TeacherPlacement.query
        .filter_by(user_id=user.id)
        .order_by(TeacherPlacement.accepted_at.desc(), TeacherPlacement.created_at.desc())
        .all()
    )
    data["placements"] = [
        {
            "id": placement.id,
            "application_id": placement.application_id,
            "job_id": placement.job_id,
            "school_name": placement.school_name,
            "job_title": placement.job_title,
            "status": placement.status,
            "accepted_at": placement.accepted_at.isoformat() if placement.accepted_at else None,
            "started_at": placement.started_at.isoformat() if placement.started_at else None,
            "ended_at": placement.ended_at.isoformat() if placement.ended_at else None,
            "notes": placement.notes,
        }
        for placement in placements
    ]
    return jsonify(data)


def _review_queue_item(user):
    profile = getattr(user, "profile", None)
    completion, _ = teacher_profile_completion(user)
    cv_exists = bool(profile and profile.cv_file_id)
    certificate_exists = bool(profile and profile.certificate_file_id)
    return {
        "id": user.id,
        "application_id": user.application_id,
        "teacher_id": user.teacher_id,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "email": user.email,
        "phone": user.phone,
        "profile_status": profile.profile_status if profile else None,
        "profile_completion": completion,
        "submitted_at": profile.submitted_at.isoformat() if profile and profile.submitted_at else None,
        "reviewed_at": profile.reviewed_at.isoformat() if profile and profile.reviewed_at else None,
        "reviewed_by_id": profile.reviewed_by_id if profile else None,
        "teaching_subject": profile.teaching_subject if profile else None,
        "curriculum_experience": profile.curriculum_experience if profile else None,
        "preferred_level": profile.preferred_level if profile else None,
        "location": profile.location if profile else None,
        "cv_present": cv_exists,
        "certificate_present": certificate_exists,
    }


def _review_detail(user):
    profile = getattr(user, "profile", None)
    completion, missing = teacher_profile_completion(user)
    data = user_json(user)
    data["profile_completion"] = completion
    data["profile_missing_fields"] = missing
    if profile:
        def _file_payload(file_id):
            if not file_id:
                return {"url": None, "filename": None, "id": None, "mime_type": None, "size_bytes": None}
            f = db.session.get(UploadedFile, file_id)
            if not f:
                return {"url": None, "filename": None, "id": None, "mime_type": None, "size_bytes": None}
            return {
                "url": f"/uploads/{f.visibility}/{f.category}/{f.stored_filename}",
                "filename": f.original_filename,
                "id": f.id,
                "mime_type": f.mime_type,
                "size_bytes": f.size_bytes,
            }
        cv_file = _file_payload(profile.cv_file_id)
        certificate_file = _file_payload(profile.certificate_file_id)
        data["review"] = {
            "location": profile.location,
            "teaching_subject": profile.teaching_subject,
            "preferred_level": profile.preferred_level,
            "preferred_employment_type": profile.preferred_employment_type,
            "available_from": profile.available_from,
            "curriculum_experience": profile.curriculum_experience,
            "bio": profile.bio,
            "cv_url": cv_file["url"],
            "cv_filename": cv_file["filename"],
            "cv_file_id": cv_file["id"],
            "cv_mime_type": cv_file["mime_type"],
            "certificate_url": certificate_file["url"],
            "certificate_filename": certificate_file["filename"],
            "certificate_file_id": certificate_file["id"],
            "certificate_mime_type": certificate_file["mime_type"],
            "profile_status": profile.profile_status,
            "submitted_at": profile.submitted_at.isoformat() if profile.submitted_at else None,
            "reviewed_at": profile.reviewed_at.isoformat() if profile.reviewed_at else None,
            "reviewed_by_id": profile.reviewed_by_id,
            "review_notes": profile.review_notes,
        }
    else:
        data["review"] = None
    applications = (
        JobApplication.query
        .options(joinedload(JobApplication.job))
        .filter_by(user_id=user.id)
        .order_by(JobApplication.updated_at.desc(), JobApplication.created_at.desc())
        .all()
    )
    data["applications"] = [
        {
            "id": app.id,
            "job_id": app.job_id,
            "status": app.status,
            "cover_note": app.cover_note,
            "created_at": app.created_at.isoformat() if app.created_at else None,
            "updated_at": app.updated_at.isoformat() if app.updated_at else None,
            "job_title": app.job.title if app.job else None,
            "organisation": app.job.organisation if app.job else None,
        }
        for app in applications
    ]
    placements = (
        TeacherPlacement.query
        .filter_by(user_id=user.id)
        .order_by(TeacherPlacement.accepted_at.desc(), TeacherPlacement.created_at.desc())
        .all()
    )
    data["placements"] = [
        {
            "id": p.id,
            "application_id": p.application_id,
            "job_id": p.job_id,
            "school_name": p.school_name,
            "job_title": p.job_title,
            "status": p.status,
            "accepted_at": p.accepted_at.isoformat() if p.accepted_at else None,
            "started_at": p.started_at.isoformat() if p.started_at else None,
            "ended_at": p.ended_at.isoformat() if p.ended_at else None,
            "notes": p.notes,
        }
        for p in placements
    ]
    return data


@admin_bp.get("/teachers/review")
@login_required
@permission_required("teachers.view")
def teacher_review_queue():
    """List teacher profiles in the review lifecycle."""
    page = max(1, request.args.get("page", 1, type=int))
    per_page = max(1, min(100, request.args.get("per_page", 50, type=int)))
    status_filter = request.args.get("status")
    search = request.args.get("search", "").strip()
    location = request.args.get("location", "").strip()
    subjects = [value.strip() for value in request.args.getlist("subject") if value.strip()]
    curricula = [value.strip() for value in request.args.getlist("curriculum") if value.strip()]

    valid_statuses = {"submitted", "under_review", "revision_required", "verified", "rejected"}
    if status_filter and status_filter not in valid_statuses:
        return jsonify(error=f"Invalid status. Choose from: {', '.join(sorted(valid_statuses))}"), 400
    if (
        len(location) > 160
        or len(subjects) > 25
        or len(curricula) > 25
        or any(len(value) > 160 for value in subjects + curricula)
    ):
        return jsonify(error="Too many or invalid teacher filter values."), 400

    query = (
        User.query
        .join(User.role)
        .join(UserProfile, UserProfile.user_id == User.id, isouter=True)
        .filter(Role.name == "user")
    )

    if status_filter:
        query = query.filter(UserProfile.profile_status == status_filter)
    else:
        query = query.filter(UserProfile.profile_status.in_(valid_statuses))

    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            db.or_(
                User.first_name.ilike(search_pattern),
                User.last_name.ilike(search_pattern),
                User.email.ilike(search_pattern),
                User.phone.ilike(search_pattern),
                User.application_id.ilike(search_pattern),
                User.teacher_id.ilike(search_pattern),
                db.func.concat(User.first_name, " ", User.last_name).ilike(search_pattern),
            )
        )

    if location:
        escaped_location = location.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        location_pattern = f"%{escaped_location}%"
        query = query.filter(
            or_(
                UserProfile.location.ilike(location_pattern, escape="\\"),
                UserProfile.preferred_locations.ilike(location_pattern, escape="\\"),
            )
        )

    def exact_comma_list_match(column, values):
        normalized = func.lower(
            func.replace(func.replace(func.coalesce(column, ""), ", ", ","), " ,", ",")
        )
        padded = literal(",") + normalized + literal(",")
        clauses = []
        for value in values:
            escaped = value.lower().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            clauses.append(padded.like(f"%,{escaped},%", escape="\\"))
        return or_(*clauses)

    # OR within each taxonomy and AND between location, subjects, and curricula.
    if subjects:
        query = query.filter(exact_comma_list_match(UserProfile.teaching_subject, subjects))
    if curricula:
        query = query.filter(exact_comma_list_match(UserProfile.curriculum_experience, curricula))

    total = query.count()
    rows = query.order_by(UserProfile.submitted_at.desc().nullslast(), User.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    return jsonify(
        items=[_review_queue_item(u) for u in rows.items],
        total=total,
        page=page,
        per_page=per_page,
        pages=rows.pages,
    )


@admin_bp.get("/teachers/<int:user_id>/review")
@login_required
@permission_required("teachers.view")
def teacher_review_detail(user_id):
    """Return a single teacher's complete review record."""
    user = db.get_or_404(User, user_id)
    if user.role and user.role.name != "user":
        return jsonify(error="Only teacher accounts can be reviewed."), 403
    return jsonify(_review_detail(user))


@admin_bp.patch("/teachers/<int:user_id>/account")
@login_required
@permission_required("teachers.account.manage")
def admin_update_teacher_account(user_id):
    """Apply an authorised, auditable correction on a teacher's behalf."""
    user = db.get_or_404(User, user_id)
    if user.role and user.role.name != "user":
        return jsonify(error="Only teacher accounts can be changed here."), 403
    payload = request.get_json(silent=True) or {}
    reason = str(payload.get("reason") or "").strip()
    if len(reason) < 8:
        return jsonify(error="Enter a reason of at least 8 characters for company records."), 400

    profile = user.profile
    if not profile:
        profile = UserProfile(user_id=user.id)
        db.session.add(profile)
    changed = {}

    for field in ("first_name", "last_name"):
        if field in payload:
            value = str(payload.get(field) or "").strip() or None
            if field == "first_name" and not value:
                return jsonify(error="First name is required."), 400
            if value != getattr(user, field):
                changed[field] = {"from": getattr(user, field), "to": value}
                setattr(user, field, value)

    if "email" in payload:
        email = str(payload.get("email") or "").strip().lower()
        if not email or "@" not in email:
            return jsonify(error="Enter a valid email address."), 400
        if User.query.filter(User.email == email, User.id != user.id).first():
            return jsonify(error="That email is already connected to another account."), 409
        if email != user.email:
            changed["email"] = {"from": user.email, "to": email}
            user.email = email

    if "phone" in payload:
        try:
            phone = normalise_phone(payload.get("phone")) if payload.get("phone") else None
        except ValueError as exc:
            return jsonify(error=str(exc)), 400
        if phone != user.phone:
            changed["phone"] = {"from": user.phone, "to": phone}
            user.phone = phone

    profile_fields = (
        "location", "teaching_subject", "preferred_level", "preferred_employment_type",
        "available_from", "curriculum_experience", "bio", "next_of_kin_name",
        "next_of_kin_phone", "next_of_kin_relationship", "next_of_kin_email",
    )
    for field in profile_fields:
        if field in payload:
            value = payload.get(field)
            if isinstance(value, str):
                value = value.strip() or None
            if value != getattr(profile, field):
                changed[field] = {"from": getattr(profile, field), "to": value}
                setattr(profile, field, value)

    if not changed:
        return jsonify(error="No account changes were supplied."), 400
    if user.teacher_service_enabled:
        upsert_contact_safely(
            user.email,
            full_name=user.full_name,
            phone=user.phone,
            source="teacher",
            source_record_id=user.id,
            metadata={"application_id": user.application_id},
            logger=current_app.logger,
        )
    log_action("teacher_account_admin_updated", "user", user.id, {
        "reason": reason,
        "changed_fields": changed,
        "profile_status_preserved": profile.profile_status,
    })
    db.session.commit()
    return jsonify(_review_detail(user))


@admin_bp.patch("/teachers/<int:user_id>/verification")
@login_required
@permission_required("teachers.verification.manage")
def admin_update_teacher_verification(user_id):
    user = db.get_or_404(User, user_id)
    if user.role and user.role.name != "user":
        return jsonify(error="Only teacher accounts can be changed here."), 403
    payload = request.get_json(silent=True) or {}
    reason = str(payload.get("reason") or "").strip()
    if len(reason) < 8:
        return jsonify(error="Enter a reason of at least 8 characters for company records."), 400
    changed = {}
    if "email_verified" in payload:
        verified = bool(payload.get("email_verified"))
        if verified != bool(user.is_verified):
            changed["email_verified"] = {"from": bool(user.is_verified), "to": verified}
            user.is_verified = verified
    if "phone_verified" in payload:
        verified = bool(payload.get("phone_verified"))
        if verified != bool(user.phone_verified):
            changed["phone_verified"] = {"from": bool(user.phone_verified), "to": verified}
            user.phone_verified = verified
            user.phone_verified_at = datetime.now(timezone.utc) if verified else None
    if not changed:
        return jsonify(error="No verification changes were supplied."), 400
    log_action("teacher_verification_admin_updated", "user", user.id, {"reason": reason, "changed_fields": changed})
    db.session.commit()
    return jsonify(_review_detail(user))


@admin_bp.post("/teachers/<int:user_id>/documents")
@login_required
@permission_required("teachers.documents.manage")
def admin_upload_teacher_document(user_id):
    """Upload a system-valid CV/certificate replacement on a teacher's behalf."""
    user = db.get_or_404(User, user_id)
    if user.role and user.role.name != "user":
        return jsonify(error="Only teacher accounts can be changed here."), 403
    kind = str(request.form.get("kind") or "").strip().lower()
    reason = str(request.form.get("reason") or "").strip()
    if kind not in {"cv", "certificate"}:
        return jsonify(error="Choose CV or certificate."), 400
    if len(reason) < 8:
        return jsonify(error="Enter a reason of at least 8 characters for company records."), 400
    profile = user.profile
    if not profile:
        profile = UserProfile(user_id=user.id)
        db.session.add(profile)
        db.session.flush()
    try:
        uploaded = save_upload(request.files.get("file"), category="documents", owner_id=user.id, visibility="protected")
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    db.session.flush()
    field = "cv_file_id" if kind == "cv" else "certificate_file_id"
    previous_file_id = getattr(profile, field)
    setattr(profile, field, uploaded.id)
    log_action("teacher_document_admin_uploaded", "uploaded_file", uploaded.id, {
        "teacher_user_id": user.id,
        "kind": kind,
        "reason": reason,
        "previous_file_id": previous_file_id,
        "profile_status_preserved": profile.profile_status,
    })
    db.session.commit()
    return jsonify(file={
        "id": uploaded.id,
        "original_filename": uploaded.original_filename,
        "preview_url": f"/api/files/{uploaded.id}/preview",
        "download_url": f"/api/files/{uploaded.id}/download",
    }, review=_review_detail(user)), 201


def _send_revision_required_email(user, note):
    first_name = user.first_name or "Teacher"
    dashboard_url = absolute_app_url("/portal?view=profile")
    escaped_note = escape(note)
    body = (
        f"<p>Dear {escape(first_name)},</p>"
        f"<p><strong>Application ID:</strong> {escape(user.application_id or 'N/A')}</p>"
        "<p>An administrator has reviewed your teacher application and requested some changes before it can proceed.</p>"
        "<p><strong>What the administrator said:</strong></p>"
        f"<blockquote style=\"border-left:3px solid #d1d5db;margin:12px 0;padding:8px 16px;color:#374151;white-space:pre-wrap\">{escaped_note}</blockquote>"
        "<p>Your application has <strong>not</strong> been rejected. You can edit your profile and replace documents to address the items above.</p>"
        "<p>After making the required changes, submit your updated profile so the administrator can continue the review.</p>"
    )
    try:
        send_email(
            OutboundEmail(
                to=user.email,
                subject="Action Required: Update Your Teacher Application",
                html=app_email_shell(
                    "Changes Requested",
                    body,
                    cta_label="Go to Dashboard",
                    cta_url=dashboard_url,
                    eyebrow="RealMindX Teacher Review",
                    preheader="An administrator has requested changes to your teacher application.",
                ),
            ),
            purpose="transactional",
            recipient_user_id=user.id,
            template_name="teacher_profile_revision_required",
        )
    except Exception as exc:
        current_app.logger.warning(
            "Revision-required email failed for user %s (error=%s)",
            user.id,
            type(exc).__name__,
        )


def _send_rejection_email(user, reason):
    first_name = user.first_name or "Teacher"
    from urllib.parse import urlencode
    contact_url = absolute_app_url(f"/contact?{urlencode({
        'subject': 'Request for reconsideration of teacher application',
        'application_id': user.application_id or '',
        'name': first_name,
    })}")
    body = (
        f"<p>Dear {escape(first_name)},</p>"
        f"<p><strong>Application ID:</strong> {escape(user.application_id or 'N/A')}</p>"
        "<p>Thank you for your interest in joining RealMindX as a teacher. After careful review of your application and documents, we are unable to proceed with your application at this time.</p>"
        f"<p><strong>Reason:</strong></p>"
        f"<blockquote style=\"border-left:3px solid #d1d5db;margin:12px 0;padding:8px 16px;color:#374151;white-space:pre-wrap\">{escape(reason)}</blockquote>"
        "<p>Please note that ordinary resubmission of this application is not available.</p>"
        "<p>If you believe the decision was made in error, or you have important new information that was not considered, you may contact RealMindX and quote your Application ID. RealMindX may reopen the application for another review where appropriate.</p>"
    )
    try:
        send_email(
            OutboundEmail(
                to=user.email,
                subject="Update on Your RealMindX Teacher Application",
                html=app_email_shell(
                    "Application Update",
                    body,
                    cta_label="Contact RealMindX",
                    cta_url=contact_url,
                    eyebrow="RealMindX Teacher Review",
                    preheader="An update is available on your teacher application.",
                ),
            ),
            purpose="transactional",
            recipient_user_id=user.id,
            template_name="teacher_profile_rejected",
        )
    except Exception as exc:
        current_app.logger.warning(
            "Rejection email failed for user %s (error=%s)",
            user.id,
            type(exc).__name__,
        )


def _send_reopened_email(user):
    first_name = user.first_name or "Teacher"
    dashboard_url = absolute_app_url("/portal")
    body = (
        f"<p>Dear {escape(first_name)},</p>"
        f"<p><strong>Application ID:</strong> {escape(user.application_id or 'N/A')}</p>"
        "<p>Your teacher application has been reopened and is now back under review by a RealMindX administrator.</p>"
        "<p>Please note that reopening your application does not mean it has been approved. It means your case will be assessed again.</p>"
        "<p>You do not need to make any changes at this time unless RealMindX contacts you to request corrections.</p>"
    )
    try:
        send_email(
            OutboundEmail(
                to=user.email,
                subject="Your Teacher Application Has Been Reopened",
                html=app_email_shell(
                    "Application Reopened",
                    body,
                    cta_label="View Dashboard",
                    cta_url=dashboard_url,
                    eyebrow="RealMindX Teacher Review",
                    preheader="Your teacher application has been reopened for further review.",
                ),
            ),
            purpose="transactional",
            recipient_user_id=user.id,
            template_name="teacher_profile_reopened",
        )
    except Exception as exc:
        current_app.logger.warning(
            "Reopen email failed for user %s (error=%s)",
            user.id,
            type(exc).__name__,
        )


def _send_verification_email(user):
    first_name = user.first_name or "Teacher"
    dashboard_url = absolute_app_url("/portal")
    body = (
        f"<p>Dear {escape(first_name)},</p>"
        f"<p><strong>Application ID:</strong> {escape(user.application_id or 'N/A')}</p>"
        f"<p><strong>Teacher ID:</strong> {escape(user.teacher_id or 'N/A')}</p>"
        "<p>Congratulations! Your RealMindX teacher profile has been visually verified by our team. Your documents and teaching details have been reviewed and confirmed.</p>"
        f"<p>Your permanent Teacher ID is <strong>{escape(user.teacher_id or 'N/A')}</strong>. Please keep this ID safe — it is your official RealMindX teacher reference and should be used in all future communications and placement records.</p>"
        "<p>You can now access features available to verified teachers on the RealMindX platform.</p>"
    )
    try:
        send_email(
            OutboundEmail(
                to=user.email,
                subject="Your RealMindX Teacher Profile Has Been Verified",
                html=app_email_shell(
                    "Profile Verified",
                    body,
                    cta_label="View Dashboard",
                    cta_url=dashboard_url,
                    eyebrow="RealMindX Teacher Verification",
                    preheader=f"Your RealMindX Teacher ID is {user.teacher_id or 'N/A'}.",
                ),
            ),
            purpose="transactional",
            recipient_user_id=user.id,
            template_name="teacher_profile_verified",
        )
    except Exception as exc:
        current_app.logger.warning(
            "Verification email failed for user %s (error=%s)",
            user.id,
            type(exc).__name__,
        )


@admin_bp.post("/teachers/<int:user_id>/start-review")
@login_required
@permission_required("teachers.edit")
def start_teacher_review(user_id):
    """Move a submitted profile to under_review."""
    user = db.get_or_404(User, user_id)
    if not user.teacher_service_enabled:
        return jsonify(error="Teacher service is not enabled for this account."), 403
    profile = user.profile
    if not profile:
        return jsonify(error="Teacher profile not found."), 404

    locked = db.session.query(UserProfile).with_for_update().filter_by(id=profile.id).first()
    profile = locked or profile

    if profile.profile_status == "under_review":
        return jsonify(
            profile_status="under_review",
            reviewed_by_id=profile.reviewed_by_id,
            reviewed_at=profile.reviewed_at.isoformat() if profile.reviewed_at else None,
            message="Review has already been started for this profile."
        ), 200

    if profile.profile_status != "submitted":
        return jsonify(error=f"Cannot start review: current status is '{profile.profile_status}'."), 400

    profile.profile_status = "under_review"
    profile.reviewed_at = datetime.now(timezone.utc)
    profile.reviewed_by_id = current_user.id

    log_action("teacher_profile_review_started", "user", user.id, {
        "application_id": user.application_id,
        "previous_status": "submitted",
        "new_status": "under_review",
    })
    db.session.commit()
    return jsonify(
        profile_status="under_review",
        reviewed_by_id=profile.reviewed_by_id,
        reviewed_at=profile.reviewed_at.isoformat(),
    ), 200


@admin_bp.post("/teachers/<int:user_id>/request-revision")
@login_required
@permission_required("teachers.edit")
def request_teacher_revision(user_id):
    """Move an under_review profile to revision_required with notes."""
    payload = request.get_json(silent=True) or {}
    note = (payload.get("note") or "").strip()
    if not note:
        return jsonify(error="A review note explaining the required corrections is required."), 400

    user = db.get_or_404(User, user_id)
    if not user.teacher_service_enabled:
        return jsonify(error="Teacher service is not enabled for this account."), 403
    profile = user.profile
    if not profile:
        return jsonify(error="Teacher profile not found."), 404

    locked = db.session.query(UserProfile).with_for_update().filter_by(id=profile.id).first()
    profile = locked or profile

    if profile.profile_status == "revision_required":
        return jsonify(
            profile_status="revision_required",
            message="Revision has already been requested for this profile."
        ), 200

    if profile.profile_status != "under_review":
        return jsonify(error=f"Cannot request revision: current status is '{profile.profile_status}'."), 400

    profile.profile_status = "revision_required"
    profile.review_notes = note
    profile.reviewed_at = datetime.now(timezone.utc)
    profile.reviewed_by_id = current_user.id

    log_action("teacher_profile_revision_requested", "user", user.id, {
        "application_id": user.application_id,
        "previous_status": "under_review",
        "new_status": "revision_required",
        "note": note,
    })
    db.session.commit()
    _send_revision_required_email(user, note)
    return jsonify(
        profile_status="revision_required",
        review_notes=note,
        reviewed_by_id=profile.reviewed_by_id,
        reviewed_at=profile.reviewed_at.isoformat(),
    ), 200


@admin_bp.post("/teachers/<int:user_id>/reject")
@login_required
@permission_required("teachers.edit")
def reject_teacher_profile(user_id):
    """Move an under_review profile to rejected with a reason."""
    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "").strip()
    if not reason:
        return jsonify(error="A rejection reason is required."), 400

    user = db.get_or_404(User, user_id)
    if not user.teacher_service_enabled:
        return jsonify(error="Teacher service is not enabled for this account."), 403
    profile = user.profile
    if not profile:
        return jsonify(error="Teacher profile not found."), 404

    locked = db.session.query(UserProfile).with_for_update().filter_by(id=profile.id).first()
    profile = locked or profile

    if profile.profile_status == "rejected":
        return jsonify(
            profile_status="rejected",
            message="This profile has already been rejected."
        ), 200

    if profile.profile_status != "under_review":
        return jsonify(error=f"Cannot reject: current status is '{profile.profile_status}'."), 400

    profile.profile_status = "rejected"
    profile.review_notes = reason
    profile.reviewed_at = datetime.now(timezone.utc)
    profile.reviewed_by_id = current_user.id

    log_action("teacher_profile_rejected", "user", user.id, {
        "application_id": user.application_id,
        "previous_status": "under_review",
        "new_status": "rejected",
        "reason": reason,
    })
    db.session.commit()
    _send_rejection_email(user, reason)
    return jsonify(
        profile_status="rejected",
        review_notes=reason,
        reviewed_by_id=profile.reviewed_by_id,
        reviewed_at=profile.reviewed_at.isoformat(),
    ), 200


@admin_bp.post("/teachers/<int:user_id>/reopen-review")
@login_required
@permission_required("teachers.edit")
def reopen_teacher_review(user_id):
    """Reopen a rejected teacher application for another review."""
    payload = request.get_json(silent=True) or {}
    note = (payload.get("note") or "").strip()
    if not note:
        return jsonify(error="A note explaining why this case is being reopened is required."), 400

    user = db.get_or_404(User, user_id)
    if not user.teacher_service_enabled:
        return jsonify(error="Teacher service is not enabled for this account."), 403
    profile = user.profile
    if not profile:
        return jsonify(error="Teacher profile not found."), 404

    locked = db.session.query(UserProfile).with_for_update().filter_by(id=profile.id).first()
    profile = locked or profile

    if profile.profile_status == "under_review":
        return jsonify(
            profile_status="under_review",
            reviewed_by_id=profile.reviewed_by_id,
            reviewed_at=profile.reviewed_at.isoformat() if profile.reviewed_at else None,
            message="This application has already been reopened."
        ), 200

    if profile.profile_status != "rejected":
        return jsonify(error=f"Cannot reopen: current status is '{profile.profile_status}'."), 400

    profile.profile_status = "under_review"
    profile.review_notes = note
    profile.reviewed_at = datetime.now(timezone.utc)
    profile.reviewed_by_id = current_user.id

    log_action("teacher_profile_review_reopened", "user", user.id, {
        "application_id": user.application_id,
        "previous_status": "rejected",
        "new_status": "under_review",
        "reopening_note": note,
    })
    db.session.commit()
    _send_reopened_email(user)
    return jsonify(
        profile_status="under_review",
        application_id=user.application_id,
        teacher_id=user.teacher_id,
        review_notes=note,
        reviewed_by_id=profile.reviewed_by_id,
        reviewed_at=profile.reviewed_at.isoformat(),
    ), 200


@admin_bp.post("/teachers/<int:user_id>/verify")
@login_required
@permission_required("teachers.edit")
def verify_teacher_profile(user_id):
    """Complete visual verification and issue a permanent Teacher ID."""
    payload = request.get_json(silent=True) or {}

    required_checks = [
        "required_documents_present",
        "documents_readable",
        "identity_details_consistent",
        "qualifications_consistent",
        "teaching_details_consistent",
        "no_obvious_alteration_detected",
    ]
    checklist = {}
    for key in required_checks:
        val = payload.get(key)
        if val is not True:
            return jsonify(error=f"Confirmation '{key}' must be true to proceed with verification."), 400
        checklist[key] = val

    user = db.get_or_404(User, user_id)
    if not user.teacher_service_enabled:
        return jsonify(error="Teacher service is not enabled for this account."), 403

    # Lock both the profile and the user row for the full transaction.
    profile = db.session.query(UserProfile).with_for_update().filter_by(user_id=user.id).first()
    if not profile:
        return jsonify(error="Teacher profile not found."), 404

    user_lock = db.session.query(User).with_for_update().filter_by(id=user.id).first()
    user = user_lock or user

    if profile.profile_status == "verified" and is_valid_teacher_id(user.teacher_id):
        return jsonify(
            application_id=user.application_id,
            teacher_id=user.teacher_id,
            profile_status="verified",
            reviewed_by_id=profile.reviewed_by_id,
            reviewed_at=profile.reviewed_at.isoformat() if profile.reviewed_at else None,
            message="This profile has already been verified."
        ), 200

    if profile.profile_status not in ("under_review", "verified"):
        return jsonify(error=f"Cannot verify: current status is '{profile.profile_status}'."), 400

    completion, _ = teacher_profile_completion(user)
    if completion < 100:
        return jsonify(error="Profile must be 100% complete before verification."), 400

    if not profile.cv_file_id:
        return jsonify(error="CV is missing."), 400
    cv = db.session.get(UploadedFile, profile.cv_file_id)
    if not cv or cv.owner_id != user.id:
        return jsonify(error="CV file record is invalid."), 400

    if not profile.certificate_file_id:
        return jsonify(error="Certificate is missing."), 400
    cert = db.session.get(UploadedFile, profile.certificate_file_id)
    if not cert or cert.owner_id != user.id:
        return jsonify(error="Certificate file record is invalid."), 400

    ensure_application_id(user)
    previous_teacher_id = user.teacher_id
    if is_valid_teacher_id(user.teacher_id):
        existing_id = user.teacher_id
        issued_now = False
    else:
        existing_id = generate_teacher_id()
        issued_now = True

    profile.profile_status = "verified"
    profile.reviewed_at = datetime.now(timezone.utc)
    profile.reviewed_by_id = current_user.id
    user.teacher_id = existing_id
    if issued_now or not user.teacher_id_issued_at:
        user.teacher_id_issued_at = datetime.now(timezone.utc)

    log_action("teacher_profile_visually_verified", "user", user.id, {
        "application_id": user.application_id,
        "teacher_id": user.teacher_id,
        "previous_status": "under_review",
        "new_status": "verified",
        "checklist": checklist,
    })
    if issued_now:
        log_action("teacher_id_issued", "user", user.id, {
            "application_id": user.application_id,
            "teacher_id": user.teacher_id,
            "replaced_malformed_id": previous_teacher_id if previous_teacher_id else None,
        })

    db.session.commit()
    _send_verification_email(user)
    return jsonify(
        application_id=user.application_id,
        teacher_id=user.teacher_id,
        profile_status="verified",
        reviewed_by_id=profile.reviewed_by_id,
        reviewed_at=profile.reviewed_at.isoformat(),
    ), 200


@admin_bp.post("/uploads")
@login_required
@permission_required("uploads.create")
def upload_admin_file():
    file = request.files.get("file")
    category = request.form.get("category") or "images"
    visibility = request.form.get("visibility") or "public"
    try:
        uploaded = save_upload(file, category=category, owner_id=current_user.id, visibility=visibility)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    log_action("upload_admin_file", "uploaded_file", uploaded.id, {"category": uploaded.category})
    db.session.commit()
    url = f"/uploads/{uploaded.visibility}/{uploaded.category}/{uploaded.stored_filename}"
    return jsonify(
        file={
            "id": uploaded.id,
            "original_filename": uploaded.original_filename,
            "category": uploaded.category,
            "visibility": uploaded.visibility,
            "url": url,
        }
    ), 201


@admin_bp.post("/staff")
@login_required
@admin_required
def create_staff():
    payload = request.get_json(silent=True) or {}
    email = (payload.get("email") or "").strip().lower()
    if not email:
        return jsonify(error="Staff email is required."), 400
    if User.query.filter_by(email=email).first():
        return jsonify(error="User already exists."), 409
    role = Role.query.filter_by(name="staff").first() or Role(name="staff", description="Staff")
    permissions = Permission.query.filter(Permission.key.in_(payload.get("permissions") or [])).all()
    user = User(
        email=email,
        first_name=payload.get("first_name") or "Staff",
        last_name=payload.get("last_name"),
        role=role,
        is_verified=True,
        is_active=payload.get("status", "active") != "inactive",
        must_change_password=True,
    )
    temporary_password = generate_temporary_password()
    user.set_password(temporary_password)
    user.direct_permissions = permissions
    db.session.add_all([role, user])
    db.session.flush()
    db.session.add(UserProfile(user_id=user.id))
    log_action("create_staff", "user", user.id, {"permissions": [p.key for p in permissions]})
    notification_status = _send_internal_account_access_email(user, "staff", temporary_password)
    db.session.commit()
    return jsonify(
        user=_staff_payload(user),
        temporary_password=temporary_password,
        notification={"email": notification_status, "email_to": user.email},
    ), 201


@admin_bp.get("/permissions")
@login_required
@admin_required
def list_permissions():
    return jsonify(items=_permissions_payload())


@admin_bp.get("/staff")
@login_required
@admin_required
def list_staff():
    staff_role = Role.query.filter_by(name="staff").first()
    rows = User.query.filter(User.role_id == staff_role.id).order_by(User.created_at.desc()).all() if staff_role else []
    return jsonify(items=[_staff_payload(user) for user in rows])


@admin_bp.put("/staff/<int:user_id>")
@login_required
@admin_required
def update_staff(user_id):
    user = db.get_or_404(User, user_id)
    if not user.role or user.role.name != "staff":
        return jsonify(error="Only staff accounts can be edited here."), 400
    payload = request.get_json(silent=True) or {}
    if "first_name" in payload:
        user.first_name = payload.get("first_name") or user.first_name
    if "last_name" in payload:
        user.last_name = payload.get("last_name") or None
    if "email" in payload and payload.get("email"):
        next_email = payload["email"].strip().lower()
        existing = User.query.filter(User.email == next_email, User.id != user.id).first()
        if existing:
            return jsonify(error="Another account already uses this email."), 409
        user.email = next_email
    if "permissions" in payload:
        user.direct_permissions = Permission.query.filter(Permission.key.in_(payload.get("permissions") or [])).all()
    if "status" in payload:
        user.is_active = payload["status"] == "active"
    log_action("update_staff", "user", user.id, {"permissions": [p.key for p in user.direct_permissions]})
    db.session.commit()
    return jsonify(user=_staff_payload(user))


@admin_bp.delete("/staff/<int:user_id>")
@login_required
@admin_required
def delete_staff(user_id):
    user = db.get_or_404(User, user_id)
    if not user.role or user.role.name != "staff":
        return jsonify(error="Only staff accounts can be deleted here."), 400
    log_action("delete_staff", "user", user.id, {"email": user.email})
    _clear_user_foreign_keys(user)
    db.session.expire(user)
    db.session.delete(user)
    db.session.commit()
    return jsonify(message="Staff account deleted.")


@admin_bp.post("/staff/<int:user_id>/reset-password")
@login_required
@admin_required
def reset_staff_password(user_id):
    user = db.get_or_404(User, user_id)
    if not user.role or user.role.name != "staff":
        return jsonify(error="Only staff accounts can be reset here."), 400
    temporary_password = generate_temporary_password()
    user.set_password(temporary_password)
    user.must_change_password = True
    user.failed_login_count = 0
    user.locked_until = None
    notification_status = _send_internal_account_access_email(user, "staff", temporary_password)
    log_action("reset_staff_password", "user", user.id, {"email": user.email})
    db.session.commit()
    return jsonify(
        message="Staff password reset. They must change it on next login.",
        temporary_password=temporary_password,
        notification={"email": notification_status, "email_to": user.email},
    )


@admin_bp.get("/admins")
@login_required
@admin_required
def list_admins():
    admin_role = Role.query.filter_by(name="admin").first()
    rows = User.query.filter(User.role_id == admin_role.id).order_by(User.created_at.desc()).all() if admin_role else []
    return jsonify(items=[_admin_payload(user) for user in rows])


@admin_bp.post("/admins")
@login_required
@admin_required
def create_admin():
    payload = request.get_json(silent=True) or {}
    email = (payload.get("email") or "").strip().lower()
    if not email:
        return jsonify(error="Admin email is required."), 400
    if User.query.filter_by(email=email).first():
        return jsonify(error="User already exists."), 409
    role = Role.query.filter_by(name="admin").first() or Role(name="admin", description="Full administrator")
    user = User(
        email=email,
        first_name=payload.get("first_name") or "Admin",
        last_name=payload.get("last_name"),
        role=role,
        is_verified=True,
        is_active=payload.get("status", "active") != "inactive",
        must_change_password=True,
    )
    temporary_password = generate_temporary_password()
    user.set_password(temporary_password)
    db.session.add_all([role, user])
    db.session.flush()
    db.session.add(UserProfile(user_id=user.id))
    log_action("create_admin", "user", user.id, {"email": user.email})
    notification_status = _send_internal_account_access_email(user, "admin", temporary_password)
    db.session.commit()
    return jsonify(
        user=_admin_payload(user),
        temporary_password=temporary_password,
        notification={"email": notification_status, "email_to": user.email},
    ), 201


@admin_bp.put("/admins/<int:user_id>")
@login_required
@admin_required
def update_admin(user_id):
    user = db.get_or_404(User, user_id)
    if not user.role or user.role.name != "admin":
        return jsonify(error="Only admin accounts can be edited here."), 400
    payload = request.get_json(silent=True) or {}
    if "first_name" in payload:
        user.first_name = payload.get("first_name") or user.first_name
    if "last_name" in payload:
        user.last_name = payload.get("last_name") or None
    if "email" in payload and payload.get("email"):
        next_email = payload["email"].strip().lower()
        existing = User.query.filter(User.email == next_email, User.id != user.id).first()
        if existing:
            return jsonify(error="Another account already uses this email."), 409
        user.email = next_email
    if "status" in payload:
        if user.id == current_user.id and payload["status"] != "active":
            return jsonify(error="You cannot deactivate the admin account you are currently using."), 400
        user.is_active = payload["status"] == "active"
    user.is_verified = True
    log_action("update_admin", "user", user.id, {"email": user.email})
    db.session.commit()
    return jsonify(user=_admin_payload(user))


@admin_bp.delete("/admins/<int:user_id>")
@login_required
@admin_required
def delete_admin(user_id):
    user = db.get_or_404(User, user_id)
    if not user.role or user.role.name != "admin":
        return jsonify(error="Only admin accounts can be deleted here."), 400
    if user.id == current_user.id:
        return jsonify(error="You cannot delete the admin account you are currently using."), 400
    log_action("delete_admin", "user", user.id, {"email": user.email})
    _clear_user_foreign_keys(user)
    db.session.expire(user)
    db.session.delete(user)
    db.session.commit()
    return jsonify(message="Admin account deleted.")


@admin_bp.post("/admins/<int:user_id>/reset-password")
@login_required
@admin_required
def reset_admin_password(user_id):
    user = db.get_or_404(User, user_id)
    if not user.role or user.role.name != "admin":
        return jsonify(error="Only admin accounts can be reset here."), 400
    if user.id == current_user.id:
        return jsonify(error="Use My Account to change your own password."), 400
    temporary_password = generate_temporary_password()
    user.set_password(temporary_password)
    user.must_change_password = True
    user.failed_login_count = 0
    user.locked_until = None
    notification_status = _send_internal_account_access_email(user, "admin", temporary_password)
    log_action("reset_admin_password", "user", user.id, {"email": user.email})
    db.session.commit()
    return jsonify(
        message="Admin password reset. They must change it on next login.",
        temporary_password=temporary_password,
        notification={"email": notification_status, "email_to": user.email},
    )


@admin_bp.get("/audit-logs")
@login_required
@admin_required
def list_audit_logs():
    rows = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(500).all()
    actor_ids = {row.actor_id for row in rows if row.actor_id}
    actors = {user.id: user for user in User.query.filter(User.id.in_(actor_ids)).all()} if actor_ids else {}
    return jsonify(items=[
        {
            "id": row.id,
            "actor": actors.get(row.actor_id).full_name if actors.get(row.actor_id) else (row.details or {}).get("actor_email") or "System",
            "actor_email": actors.get(row.actor_id).email if actors.get(row.actor_id) else (row.details or {}).get("actor_email"),
            "actor_role": actors.get(row.actor_id).role.name.replace("_", " ").title() if actors.get(row.actor_id) and actors.get(row.actor_id).role else "System",
            "action": readable_audit_action(row.action, row.details),
            "raw_action": row.action,
            "summary": readable_audit_summary(row.action, row.entity_type, row.details),
            "entity_type": (lambda label: label[:1].upper() + label[1:])(
                AREA_LABELS.get(row.entity_type) or str(row.entity_type or "System").replace("_", " ")
            ),
            "entity_id": row.entity_id,
            "details": row.details or {},
            "ip_address": row.ip_address,
            "created_at": row.created_at.isoformat(),
            "updated_at": row.created_at.isoformat(),
        }
        for row in rows
    ])


@admin_bp.get("/products")
@login_required
@permission_required("products.view")
def list_products():
    rows = (
        Product.query
        .options(
            joinedload(Product.category),
            joinedload(Product.image_file),
            joinedload(Product.image_original_file),
            joinedload(Product.image_medium_file),
            joinedload(Product.image_thumb_file),
            selectinload(Product.reviews),
        )
        .order_by(Product.featured.desc(), Product.created_at.desc())
        .all()
    )
    return jsonify(items=[product_json(row, include_private=True) for row in rows])


def _uploaded_file_exists(uploaded_file):
    if not uploaded_file:
        return False

    upload_root = Path(current_app.config["UPLOAD_FOLDER"])
    candidates = []
    if uploaded_file.storage_path:
        stored_path = Path(uploaded_file.storage_path)
        candidates.append(stored_path if stored_path.is_absolute() else upload_root / stored_path)
    if uploaded_file.visibility and uploaded_file.category and uploaded_file.stored_filename:
        candidates.append(upload_root / uploaded_file.visibility / uploaded_file.category / uploaded_file.stored_filename)

    seen = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            resolved = candidate
        if str(resolved) in seen:
            continue
        seen.add(str(resolved))
        try:
            if resolved.is_file() and resolved.stat().st_size > 0:
                return True
        except OSError:
            continue
    return False


def _products_without_display_images(active_only=None):
    query = Product.query.options(joinedload(Product.image_file))
    if active_only is True:
        query = query.filter(Product.is_active.is_(True))
    elif active_only is False:
        query = query.filter(Product.is_active.is_(False))
    return [
        product for product in query.order_by(Product.created_at.desc()).all()
        if not _uploaded_file_exists(product.image_file)
    ]


@admin_bp.get("/products/missing-images")
@login_required
@permission_required("products.view")
def products_missing_images():
    products = _products_without_display_images()
    return jsonify(
        total=len(products),
        published=sum(1 for product in products if product.is_active),
        draft=sum(1 for product in products if not product.is_active),
    )


@admin_bp.post("/products/missing-images/unpublish")
@login_required
@permission_required("products.edit")
def unpublish_products_missing_images():
    products = _products_without_display_images(active_only=True)
    product_ids = [product.id for product in products]
    for product in products:
        product.is_active = False
    log_action(
        "bulk_unpublish_products_missing_images",
        "product",
        None,
        {"count": len(product_ids), "product_ids": product_ids},
    )
    db.session.commit()
    return jsonify(
        message=f"Unpublished {len(product_ids)} product(s) without images.",
        unpublished=len(product_ids),
        product_ids=product_ids,
    )


@admin_bp.post("/products")
@login_required
@permission_required("products.create")
def create_product():
    payload = request.get_json(silent=True) or {}
    name = payload.get("name")
    category = _category_from_payload(payload)
    slug = payload.get("slug") or name
    product = Product(
        name=name,
        slug=_unique_product_slug(slug),
        category=category,
        price=payload.get("price") or 0,
        old_price=payload.get("old_price"),
        short_description=payload.get("short_description"),
        full_description=payload.get("full_description"),
        image_file_id=payload.get("image_file_id") or None,
        stock_status=payload.get("stock_status") or "in_stock",
        quantity_available=payload.get("quantity_available"),
        subject=_normalize_taxonomy(payload.get("subject"), "subject"),
        level=_normalize_taxonomy(payload.get("level"), "level"),
        curriculum=_normalize_taxonomy(payload.get("curriculum"), "curriculum"),
        author=payload.get("author"),
        publisher=_normalize_taxonomy(payload.get("publisher"), "publisher"),
        product_type=payload.get("product_type"),
        source=payload.get("source"),
        featured=bool(payload.get("featured")),
        delivery_note=payload.get("delivery_note"),
        tags=payload.get("tags") or [],
    )
    if not product.name:
        return jsonify(error="Product name is required."), 400
    db.session.add(product)
    db.session.flush()
    if product.image_file_id:
        ensure_product_image_variants(product, owner_id=current_user.id)
    log_action("create_product", "product", product.id)
    db.session.commit()
    return jsonify(product=product_json(product, include_private=True)), 201


@admin_bp.put("/products/<int:product_id>")
@login_required
@permission_required("products.edit")
def update_product(product_id):
    product = db.get_or_404(Product, product_id)
    payload = request.get_json(silent=True) or {}
    previous_image_file_id = product.image_file_id
    for field in [
        "name",
        "slug",
        "price",
        "old_price",
        "short_description",
        "full_description",
        "image_file_id",
        "stock_status",
        "quantity_available",
        "subject",
        "level",
        "curriculum",
        "author",
        "publisher",
        "product_type",
        "source",
        "featured",
        "delivery_note",
        "tags",
        "is_active",
    ]:
        if field in payload:
            setattr(product, field, payload[field])
    for _tax_field, _tax_type in [("subject", "subject"), ("level", "level"), ("curriculum", "curriculum"), ("publisher", "publisher")]:
        if _tax_field in payload:
            setattr(product, _tax_field, _normalize_taxonomy(payload[_tax_field], _tax_type))
    if "slug" in payload:
        product.slug = _unique_product_slug(payload["slug"] or product.name, product.id)
    if "category_id" in payload or "category_name" in payload or "category" in payload:
        product.category = _category_from_payload(payload)
    if "status" in payload:
        product.is_active = payload["status"] == "published"
    if "image_file_id" in payload:
        product.image_file_id = payload.get("image_file_id") or None
        if product.image_file_id != previous_image_file_id:
            reset_product_image_variants(product)
    if product.image_file_id:
        ensure_product_image_variants(product, owner_id=current_user.id)
    log_action("update_product", "product", product.id)
    db.session.commit()
    return jsonify(product=product_json(product, include_private=True))


@admin_bp.delete("/products/<int:product_id>")
@login_required
@permission_required("products.delete")
def delete_product(product_id):
    product = db.get_or_404(Product, product_id)
    log_action("delete_product", "product", product.id, {"name": product.name})
    OrderItem.query.filter_by(product_id=product.id).update(
        {OrderItem.product_id: None},
        synchronize_session=False,
    )
    AnalyticsEvent.query.filter_by(product_id=product.id).update(
        {AnalyticsEvent.product_id: None},
        synchronize_session=False,
    )
    ProductReview.query.filter_by(product_id=product.id).delete(synchronize_session=False)
    db.session.delete(product)
    db.session.commit()
    return jsonify(message="Product deleted.")


def _review_json(review):
    return {
        "id": review.id,
        "product_id": review.product_id,
        "product_name": review.product.name if review.product else "",
        "order_id": review.order_id,
        "customer_name": review.customer_name,
        "email": review.email,
        "rating": review.rating,
        "title": review.title,
        "comment": review.comment,
        "status": review.status,
        "created_at": review.created_at.isoformat(),
        "updated_at": review.updated_at.isoformat(),
    }


@admin_bp.get("/product-reviews")
@login_required
@permission_required("productReviews.view")
def list_product_reviews():
    rows = ProductReview.query.order_by(ProductReview.created_at.desc()).limit(300).all()
    return jsonify(items=[_review_json(row) for row in rows])


@admin_bp.put("/product-reviews/<int:review_id>")
@login_required
@permission_required("productReviews.edit")
def update_product_review(review_id):
    review = db.get_or_404(ProductReview, review_id)
    payload = request.get_json(silent=True) or {}
    if "status" in payload:
        status = payload["status"]
        if status not in {"pending", "approved", "rejected"}:
            return jsonify(error="Choose pending, approved, or rejected."), 400
        review.status = status
    for field in ["title", "comment"]:
        if field in payload:
            setattr(review, field, payload[field])
    log_action("update_product_review", "product_review", review.id, {"status": review.status})
    db.session.commit()
    return jsonify(_review_json(review))


@admin_bp.delete("/product-reviews/<int:review_id>")
@login_required
@permission_required("productReviews.delete")
def delete_product_review(review_id):
    review = db.get_or_404(ProductReview, review_id)
    log_action("delete_product_review", "product_review", review.id, {"product_id": review.product_id})
    db.session.delete(review)
    db.session.commit()
    return jsonify(message="Review deleted.")


@admin_bp.get("/order-reviews")
@login_required
@permission_required("orderReviews.view")
def list_order_reviews():
    rows = OrderReview.query.order_by(OrderReview.created_at.desc()).limit(300).all()
    return jsonify(items=[order_review_json(row) for row in rows])


@admin_bp.put("/order-reviews/<int:review_id>")
@login_required
@permission_required("orderReviews.edit")
def update_order_review(review_id):
    review = db.get_or_404(OrderReview, review_id)
    payload = request.get_json(silent=True) or {}
    if "status" in payload:
        status = str(payload["status"] or "").strip().lower()
        if status not in {"new", "reviewed", "follow_up", "archived"}:
            return jsonify(error="Choose new, reviewed, follow_up, or archived."), 400
        review.status = status
    if "admin_notes" in payload:
        review.admin_notes = (payload.get("admin_notes") or "").strip() or None
    log_action("update_order_review", "order_review", review.id, {"status": review.status})
    db.session.commit()
    return jsonify(order_review_json(review))


@admin_bp.delete("/order-reviews/<int:review_id>")
@login_required
@permission_required("orderReviews.delete")
def delete_order_review(review_id):
    review = db.get_or_404(OrderReview, review_id)
    log_action("delete_order_review", "order_review", review.id, {"order_id": review.order_id})
    db.session.delete(review)
    db.session.commit()
    return jsonify(message="Order review deleted.")


@admin_bp.post("/products/import")
@login_required
@permission_required("products.create")
def import_products():
    saved_paths = []
    saved_ids = []
    try:
        rows = _read_catalog_rows(request.files.get("catalog_file"))
        headers = _import_headers(rows)
        mapping = _parse_import_mapping(request.form.get("column_mapping"), headers)
        overwrite_slugs = set(json.loads(request.form.get("overwrite_slugs") or "[]"))
        if mapping and not mapping.get("name"):
            raise ValueError("Map a catalogue column to Product name before importing.")
        image_ids, saved_paths, saved_ids = _save_imported_images(request.files.get("images_zip"), current_user.id)
    except ValueError as exc:
        db.session.rollback()
        for path in saved_paths:
            path.unlink(missing_ok=True)
        if saved_ids:
            UploadedFile.query.filter(UploadedFile.id.in_(saved_ids)).delete(synchronize_session=False)
            db.session.commit()
        return jsonify(error=str(exc)), 400

    imported = 0
    updated = 0
    skipped = []
    missing_images = set()
    image_dir = Path(current_app.config["UPLOAD_FOLDER"]) / "public" / "images"
    files_before = set(image_dir.rglob("*")) if image_dir.exists() else set()
    try:
        for index, raw_row in enumerate(rows, start=2):
            row = _normalise_import_row(_apply_import_mapping(raw_row, mapping))
            if not row:
                skipped.append({"row": index, "reason": "Missing product name"})
                continue
            category = _ensure_category(row["category"])
            category_slug = slugify(row["category"]) if row.get("category") else None
            product = Product.query.filter_by(slug=row["slug"]).first()
            if not product and category_slug:
                product = Product.query.filter(
                    Product.name.ilike(row["name"]),
                    Product.category.has(ProductCategory.slug == category_slug)
                ).first()
            
            if product and row["slug"] not in overwrite_slugs:
                skipped.append({"row": index, "reason": f"Conflict with existing product '{product.name}' not marked for overwrite"})
                continue

            if not product:
                product = Product(name=row["name"], slug=_unique_product_slug(row["slug"]), price=row["price"])
                db.session.add(product)
                imported += 1
            else:
                updated += 1
            product.name = row["name"]
            product.category = category
            product.price = row["price"]
            product.old_price = row["old_price"]
            product.short_description = row["short_description"]
            product.full_description = row["full_description"]
            product.stock_status = row["stock_status"]
            product.quantity_available = row["quantity_available"]
            product.subject = _normalize_taxonomy(row["subject"], "subject")
            product.level = _normalize_taxonomy(row["level"], "level")
            product.curriculum = _normalize_taxonomy(row["curriculum"], "curriculum")
            product.author = row["author"]
            product.publisher = _normalize_taxonomy(row["publisher"], "publisher")
            product.product_type = row["product_type"]
            product.delivery_note = row["delivery_note"]
            product.tags = row["tags"]
            product.featured = row["featured"]
            product.source = row["source"]
            product.is_active = True
            image_filename = row["image_filename"].strip()
            image_key = _image_entry_key(image_filename)
            if image_key and image_key in image_ids:
                next_image_id = image_ids[image_key]
                if product.image_file_id != next_image_id:
                    product.image_file_id = next_image_id
                    reset_product_image_variants(product)
            elif image_filename:
                missing_images.add(image_filename)
            if product.image_file_id:
                ensure_product_image_variants(product, owner_id=current_user.id)
        log_action(
            "import_products",
            "product",
            None,
            {
                "imported": imported,
                "updated": updated,
                "skipped": skipped,
                "images_saved": len(image_ids),
                "missing_images": sorted(missing_images),
            },
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        for path in saved_paths:
            path.unlink(missing_ok=True)
        if image_dir.exists():
            for path in set(image_dir.rglob("*")) - files_before:
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    pass
        if saved_ids:
            UploadedFile.query.filter(UploadedFile.id.in_(saved_ids)).delete(synchronize_session=False)
            db.session.commit()
        current_app.logger.exception("Product catalogue import failed.")
        return jsonify(error="The catalogue could not be imported. No product changes were saved."), 500
    return jsonify(
        imported=imported,
        updated=updated,
        skipped=skipped,
        images_saved=len(image_ids),
        missing_images=sorted(missing_images),
    )


@admin_bp.post("/products/import/preview")
@login_required
@permission_required("products.create")
def preview_product_import():
    try:
        rows = _read_catalog_rows(request.files.get("catalog_file"))
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    if not rows:
        return jsonify(error="The catalogue contains headers but no product rows."), 400

    headers = _import_headers(rows)
    mapping = _suggest_import_mapping(headers)
    name_source = mapping.get("name")
    blank_names = sum(
        1
        for row in rows
        if not str(row.get(name_source) or "").strip()
    ) if name_source else len(rows)
    warnings = []
    if not name_source:
        warnings.append("No product-name column was matched. Choose the correct column below.")
    elif blank_names:
        warnings.append(f"{blank_names} row(s) have no product name and will be skipped.")

    conflicts = []
    if name_source:
        for index, raw_row in enumerate(rows, start=2):
            row = _normalise_import_row(_apply_import_mapping(raw_row, mapping))
            if not row:
                continue
            category_slug = slugify(row["category"]) if row.get("category") else None
            product = Product.query.filter_by(slug=row["slug"]).first()
            if not product and category_slug:
                product = Product.query.filter(
                    Product.name.ilike(row["name"]),
                    Product.category.has(ProductCategory.slug == category_slug)
                ).first()
            if product:
                conflicts.append({
                    "row": index,
                    "slug": row["slug"],
                    "existing_id": product.id,
                    "existing_name": product.name,
                    "existing_category": product.category.name if product.category else "",
                    "import_name": row["name"],
                })

    sample_rows = [
        {key: _json_safe_import_value(value) for key, value in row.items()}
        for row in rows[:8]
    ]
    fields = [
        {
            "key": field["key"],
            "label": field["label"],
            "required": bool(field.get("required")),
        }
        for field in PRODUCT_IMPORT_FIELDS
    ]
    return jsonify(
        row_count=len(rows),
        headers=headers,
        mapping=mapping,
        fields=fields,
        sample_rows=sample_rows,
        warnings=warnings,
        conflicts=conflicts,
    )


@admin_bp.post("/products/import/images/preview")
@login_required
@permission_required("products.create")
def preview_product_images_import():
    """
    Preview an images-only import. Matches images in a ZIP to existing products
    by original_filename (case-insensitive). Returns matched products, unmatched
    images, duplicate matches, and invalid files.
    """
    images_zip = request.files.get("images_zip")
    if not images_zip or not images_zip.filename:
        return jsonify(error="Upload an image ZIP file."), 400
    if not images_zip.filename.lower().endswith(".zip"):
        return jsonify(error="Only ZIP files are supported for image batches."), 400

    # Validate ZIP
    images_zip.stream.seek(0, 2)
    archive_bytes = images_zip.stream.tell()
    images_zip.stream.seek(0)
    if archive_bytes > 100 * 1024 * 1024:
        return jsonify(error="Image ZIP must be 100 MB or smaller."), 400
    if not zipfile.is_zipfile(images_zip.stream):
        return jsonify(error="Choose a valid ZIP archive for the product images."), 400
    images_zip.stream.seek(0)

    # Build a lookup of existing products by original_filename (case-insensitive).
    # Only products that have an image_file are matchable.
    product_by_filename = {}
    duplicate_filenames = set()
    products = Product.query.filter(Product.image_file_id.isnot(None)).all()
    for product in products:
        if product.image_file and product.image_file.original_filename:
            key = _image_entry_key(product.image_file.original_filename)
            if key in product_by_filename:
                duplicate_filenames.add(key)
            else:
                product_by_filename[key] = product

    matched = []
    unmatched = []
    invalid_files = []
    duplicate_matches = []

    try:
        images_zip.stream.seek(0)
        with zipfile.ZipFile(images_zip.stream) as archive:
            entries = [
                entry
                for entry in archive.infolist()
                if not entry.is_dir() and "__MACOSX" not in entry.filename
            ]
            if len(entries) > 500:
                return jsonify(error="Image ZIP contains too many files. Use at most 500."), 400

            total_uncompressed = 0
            seen_keys = set()

            # First pass: safety validation
            for entry in entries:
                raw_name = entry.filename
                basename = Path(raw_name).name
                if not basename:
                    continue
                if ".." in raw_name.replace("\\", "/").split("/"):
                    invalid_files.append({"filename": raw_name, "reason": "Unsafe ZIP path rejected."})
                    continue
                if entry.file_size > 12 * 1024 * 1024:
                    invalid_files.append({"filename": basename, "reason": "File exceeds 12 MB limit."})
                    continue
                total_uncompressed += entry.file_size
                if total_uncompressed > 512 * 1024 * 1024:
                    return jsonify(error="Image ZIP expands beyond the 512 MB safety limit."), 400

                # Check file extension
                ext = Path(basename).suffix.lower().lstrip(".")
                allowed = current_app.config["ALLOWED_UPLOAD_EXTENSIONS"].get("images", set())
                if ext not in allowed:
                    invalid_files.append({"filename": basename, "reason": f"Unsupported file type: .{ext}"})
                    continue

                key = _image_entry_key(basename)
                if not key:
                    invalid_files.append({"filename": basename, "reason": "Invalid image filename."})
                    continue

                # Duplicate filename inside the ZIP (e.g. two folders with the same image name)
                if key in seen_keys:
                    duplicate_matches.append({
                        "filename": basename,
                        "reason": "The ZIP contains this filename more than once. Skipped to keep one image per product.",
                    })
                    continue
                seen_keys.add(key)

                if key in duplicate_filenames:
                    # Multiple products share this filename - ambiguous match
                    duplicate_matches.append({
                        "filename": basename,
                        "reason": "Multiple products use this image filename. Cannot determine which to update.",
                    })
                    continue

                product = product_by_filename.get(key)
                if product:
                    matched.append({
                        "filename": basename,
                        "file_size": entry.file_size,
                        "product_id": product.id,
                        "product_name": product.name,
                        "product_slug": product.slug,
                        "product_category": product.category.name if product.category else "",
                        "current_image_filename": product.image_file.original_filename if product.image_file else "",
                        "current_image_id": product.image_file_id,
                        "existing_image_url": _upload_public_url(product.image_file),
                    })
                else:
                    unmatched.append({
                        "filename": basename,
                        "reason": "No product found with this image filename.",
                    })
    except zipfile.BadZipFile:
        return jsonify(error="The uploaded file is not a valid ZIP archive."), 400
    except Exception:
        current_app.logger.exception("Image ZIP preview failed.")
        return jsonify(error="Failed to process the image ZIP."), 500

    warnings = []
    if not matched and not unmatched and not invalid_files:
        warnings.append("No valid images found in the ZIP.")
    elif not matched:
        warnings.append("No images matched any existing products.")

    return jsonify(
        total_images=len(matched) + len(unmatched) + len(invalid_files) + len(duplicate_matches),
        matched_count=len(matched),
        unmatched_count=len(unmatched),
        invalid_count=len(invalid_files),
        duplicate_count=len(duplicate_matches),
        matched=matched,
        unmatched=unmatched,
        invalid_files=invalid_files,
        duplicate_matches=duplicate_matches,
        warnings=warnings,
    )


@admin_bp.post("/products/import/images")
@login_required
@permission_required("products.create")
def import_product_images():
    """
    Import product images from a ZIP file. Updates images for existing products
    matched by original_filename (case-insensitive). Does not create new products
    or modify any other product fields.
    """
    images_zip = request.files.get("images_zip")
    if not images_zip or not images_zip.filename:
        return jsonify(error="Upload an image ZIP file."), 400
    if not images_zip.filename.lower().endswith(".zip"):
        return jsonify(error="Only ZIP files are supported for image batches."), 400

    # Get the list of product IDs to update from the request
    product_ids = set(json.loads(request.form.get("product_ids") or "[]"))
    if not product_ids:
        return jsonify(error="No products selected for image update."), 400

    # Validate ZIP
    images_zip.stream.seek(0, 2)
    archive_bytes = images_zip.stream.tell()
    images_zip.stream.seek(0)
    if archive_bytes > 100 * 1024 * 1024:
        return jsonify(error="Image ZIP must be 100 MB or smaller."), 400
    if not zipfile.is_zipfile(images_zip.stream):
        return jsonify(error="Choose a valid ZIP archive for the product images."), 400
    images_zip.stream.seek(0)

    # Build lookup of products to update
    products = Product.query.filter(Product.id.in_(product_ids)).all()
    product_by_id = {p.id: p for p in products}
    if len(product_by_id) != len(product_ids):
        return jsonify(error="One or more selected products no longer exist."), 400

    # Only save the images that the selected products actually need, so unselected
    # files never linger in storage.
    needed_keys = {
        _image_entry_key(product.image_file.original_filename)
        for product in product_by_id.values()
        if product.image_file and product.image_file.original_filename
    }

    saved_paths = []
    saved_ids = []
    try:
        image_ids, saved_paths, saved_ids = _save_imported_images(
            images_zip, current_user.id, only_filenames=needed_keys
        )
    except ValueError as exc:
        db.session.rollback()
        for path in saved_paths:
            path.unlink(missing_ok=True)
        if saved_ids:
            UploadedFile.query.filter(UploadedFile.id.in_(saved_ids)).delete(synchronize_session=False)
            db.session.commit()
        return jsonify(error=str(exc)), 400

    updated = 0
    skipped = []
    attached_ids = set()
    image_dir = Path(current_app.config["UPLOAD_FOLDER"]) / "public" / "images"
    files_before = set(image_dir.rglob("*")) if image_dir.exists() else set()
    try:
        for product_id in product_ids:
            product = product_by_id.get(product_id)
            if not product:
                skipped.append({"product_id": product_id, "reason": "Product not found"})
                continue

            # Find the image file that matches this product's original filename
            if not product.image_file or not product.image_file.original_filename:
                skipped.append({"product_id": product_id, "reason": "Product has no existing image to replace"})
                continue

            key = _image_entry_key(product.image_file.original_filename)
            if key not in image_ids:
                skipped.append({"product_id": product_id, "reason": f"No image found in ZIP for filename '{product.image_file.original_filename}'"})
                continue

            # Update the product image. No other product field is touched.
            new_image_id = image_ids[key]
            product.image_file_id = new_image_id
            attached_ids.add(new_image_id)
            reset_product_image_variants(product)
            ensure_product_image_variants(product, owner_id=current_user.id)
            updated += 1
    except Exception:
        db.session.rollback()
        for path in saved_paths:
            path.unlink(missing_ok=True)
        if image_dir.exists():
            for path in set(image_dir.rglob("*")) - files_before:
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    pass
        if saved_ids:
            UploadedFile.query.filter(UploadedFile.id.in_(saved_ids)).delete(synchronize_session=False)
            db.session.commit()
        current_app.logger.exception("Product image update failed.")
        return jsonify(error="The product images could not be updated. No changes were saved."), 500

    if updated == 0:
        db.session.rollback()
        for path in saved_paths:
            path.unlink(missing_ok=True)
        if saved_ids:
            UploadedFile.query.filter(UploadedFile.id.in_(saved_ids)).delete(synchronize_session=False)
            db.session.commit()
        return jsonify(
            updated=0,
            skipped=skipped,
            message="No images were updated.",
        ), 400

    # Remove any images saved for products that ended up skipped so nothing is
    # left orphaned in storage.
    orphan_ids = set(saved_ids) - attached_ids
    if orphan_ids:
        for file_id in orphan_ids:
            uploaded = db.session.get(UploadedFile, file_id)
            if uploaded:
                if uploaded.storage_path:
                    try:
                        Path(uploaded.storage_path).unlink(missing_ok=True)
                    except OSError:
                        current_app.logger.warning("Could not remove orphaned image file %s", uploaded.storage_path)
                db.session.delete(uploaded)

    log_action(
        "update_product_images",
        "product",
        None,
        {
            "updated": updated,
            "skipped": skipped,
        },
    )
    db.session.commit()
    return jsonify(
        updated=updated,
        skipped=skipped,
    )


def _build_admin_export_pdf(title, headers, data_rows):
    """Build a readable, repeat-header table for admin exports.

    CSV/XLSX preserve every field for operational work. The PDF intentionally
    uses the same columns in landscape A3 so it remains a complete printable
    report instead of a truncated list of pipe-separated values.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

    stream = io.BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=landscape(A3),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
        author="RealMindX",
    )
    styles = getSampleStyleSheet()
    heading = styles["Title"]
    heading.fontName = "Helvetica-Bold"
    heading.fontSize = 17
    heading.leading = 21
    heading.textColor = colors.HexColor("#143670")
    cell = styles["BodyText"]
    cell.fontName = "Helvetica"
    cell.fontSize = 5.8
    cell.leading = 7.2
    cell.textColor = colors.HexColor("#172b4d")
    header = styles["BodyText"]
    header.fontName = "Helvetica-Bold"
    header.fontSize = 5.8
    header.leading = 7
    header.textColor = colors.white

    def value(item):
        text = str(item if item not in (None, "") else "-")
        return escape(text).replace("\n", "<br/>")

    table_data = [[Paragraph(escape(column.replace("_", " ").title()), header) for column in headers]]
    for row in data_rows:
        table_data.append([Paragraph(value(row.get(column)), cell) for column in headers])
    usable_width = landscape(A3)[0] - document.leftMargin - document.rightMargin
    table = LongTable(table_data, colWidths=[usable_width / max(len(headers), 1)] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#143670")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9e2f0")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f9fd")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    document.build([Paragraph(title, heading), Spacer(1, 5 * mm), table])
    stream.seek(0)
    return stream


@admin_bp.get("/products/export")
@login_required
@permission_required("products.export")
def export_products():
    export_format = request.args.get("format", "zip").lower()
    
    rows = Product.query.order_by(Product.created_at.desc()).all()
    headers = [
        "id", "name", "category", "price", "old_price", "stock_status", "quantity_available",
        "subject", "level", "curriculum", "author", "publisher", "product_type", "source", "featured", "is_active", "tags",
        "image_filename",
    ]
    data_rows = [
        {
            "id": product.id,
            "name": product.name,
            "category": product.category.name if product.category else "",
            "price": float(product.price or 0),
            "old_price": float(product.old_price or 0) if product.old_price is not None else "",
            "stock_status": product.stock_status,
            "quantity_available": product.quantity_available or "",
            "subject": product.subject or "",
            "level": product.level or "",
            "curriculum": getattr(product, "curriculum", None) or "",
            "author": getattr(product, "author", None) or "",
            "publisher": getattr(product, "publisher", None) or "",
            "product_type": product.product_type or "",
            "source": product.source or "",
            "featured": product.featured,
            "is_active": product.is_active,
            "tags": ", ".join(product.tags or []),
            "image_filename": getattr(product.image_file, "original_filename", "") if getattr(product, "image_file", None) else "",
        }
        for product in rows
    ]

    if export_format == "csv":
        csv_out = io.StringIO(); writer = csv.DictWriter(csv_out, fieldnames=headers); writer.writeheader(); writer.writerows(data_rows)
        return Response(csv_out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=products.csv"})
    if export_format == "xlsx":
        try: from openpyxl import Workbook
        except ImportError: return jsonify(error="XLSX export requires openpyxl."), 501
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Products"; sheet.append(headers)
        for row in data_rows: sheet.append([row.get(key) for key in headers])
        stream = io.BytesIO(); workbook.save(stream); stream.seek(0)
        return send_file(stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="products.xlsx")
    if export_format == "pdf":
        try:
            pdf_stream = _build_admin_export_pdf("RealMindX Bookshop Products", headers, data_rows)
        except ImportError: return jsonify(error="PDF export requires reportlab."), 501
        return send_file(pdf_stream, mimetype="application/pdf", as_attachment=True, download_name="products.pdf")
    if export_format == "zip":
        zip_stream = io.BytesIO()
        with zipfile.ZipFile(zip_stream, "w", zipfile.ZIP_DEFLATED) as zf:
            csv_out = io.StringIO(); writer = csv.DictWriter(csv_out, fieldnames=headers); writer.writeheader(); writer.writerows(data_rows)
            zf.writestr("products.csv", csv_out.getvalue())
            for product in rows:
                if getattr(product, "image_file", None) and product.image_file.storage_path:
                    try:
                        with open(product.image_file.storage_path, "rb") as f: zf.writestr(f"images/{product.image_file.original_filename}", f.read())
                    except Exception: pass
        zip_stream.seek(0)
        return send_file(zip_stream, mimetype="application/zip", as_attachment=True, download_name="realmindx-products-export.zip")
    return jsonify(error="Use csv, xlsx, pdf, or zip."), 400


@admin_bp.get("/users/export")
@login_required
@permission_required("teachers.export")
def export_users():
    """Export registered teachers with their complete work preferences."""
    from openpyxl import Workbook as XlsxWorkbook
    export_format = (request.args.get("format") or "xlsx").lower()
    if export_format not in {"csv", "xlsx", "pdf"}:
        return jsonify(error="Unsupported format. Use csv, xlsx, or pdf."), 400
    rows = (
        User.query
        .join(Role, User.role_id == Role.id, isouter=True)
        .filter(Role.name == "user")
        .order_by(User.created_at.desc())
        .all()
    )
    headers = [
        "id", "first_name", "last_name", "email", "phone", "is_verified", "last_login_at", "created_at",
        "location", "teaching_subject", "preferred_level", "preferred_employment_type", "available_from",
        "curriculum_experience", "preferred_locations", "years_of_experience", "date_of_birth", "bio",
        "cv_filename", "certificate_filename", "next_of_kin_name", "next_of_kin_phone",
        "next_of_kin_relationship", "next_of_kin_email",
    ]

    data_rows = []
    for u in rows:
        profile = u.profile
        data_rows.append({
            "id": u.id, "first_name": u.first_name or "", "last_name": u.last_name or "",
            "email": u.email, "phone": u.phone or "", "is_verified": "Yes" if u.is_verified else "No",
            "last_login_at": str(u.last_login_at or ""), "created_at": str(u.created_at or ""),
            "location": getattr(profile, "location", "") or "",
            "teaching_subject": getattr(profile, "teaching_subject", "") or "",
            "preferred_level": getattr(profile, "preferred_level", "") or "",
            "preferred_employment_type": getattr(profile, "preferred_employment_type", "") or "",
            "available_from": getattr(profile, "available_from", "") or "",
            "curriculum_experience": getattr(profile, "curriculum_experience", "") or "",
            "preferred_locations": getattr(profile, "preferred_locations", "") or "",
            "years_of_experience": getattr(profile, "years_of_experience", "") or "",
            "date_of_birth": str(getattr(profile, "date_of_birth", "") or ""),
            "bio": getattr(profile, "bio", "") or "",
            "cv_filename": getattr(db.session.get(UploadedFile, getattr(profile, "cv_file_id", None)), "original_filename", "") if profile and getattr(profile, "cv_file_id", None) else "",
            "certificate_filename": getattr(db.session.get(UploadedFile, getattr(profile, "certificate_file_id", None)), "original_filename", "") if profile and getattr(profile, "certificate_file_id", None) else "",
            "next_of_kin_name": getattr(profile, "next_of_kin_name", "") or "",
            "next_of_kin_phone": getattr(profile, "next_of_kin_phone", "") or "",
            "next_of_kin_relationship": getattr(profile, "next_of_kin_relationship", "") or "",
            "next_of_kin_email": getattr(profile, "next_of_kin_email", "") or "",
        })

    if export_format == "csv":
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data_rows)
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=realmindx-teachers.csv"})

    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Teachers"
    ws.append([column.replace("_", " ").title() for column in headers])
    for row in data_rows:
        ws.append([row[column] for column in headers])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    if export_format == "pdf":
        try:
            return send_file(_build_admin_export_pdf("RealMindX Teachers", headers, data_rows), mimetype="application/pdf", as_attachment=True, download_name="realmindx-teachers.pdf")
        except ImportError:
            return jsonify(error="PDF export requires reportlab."), 501
    return send_file(stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="realmindx-teachers.xlsx")


@admin_bp.get("/jobs/export")
@login_required
@permission_required("jobs.export")
def export_jobs():
    """Export job posts as CSV, XLSX, or PDF."""
    export_format = (request.args.get("format") or "csv").lower()
    rows = Job.query.order_by(Job.created_at.desc()).all()
    headers = [
        "id", "title", "organisation", "location", "delivery_zone", "subject", "level", "curriculum",
        "employment_type", "preferred_sex", "preferred_age_range", "description", "requirements",
        "responsibilities", "salary_min", "salary_max", "salary_currency", "deadline", "status",
        "created_by_user_id", "created_at", "updated_at",
    ]

    data_rows = [
        {
            "id": j.id,
            "title": j.title,
            "organisation": j.organisation or "",
            "location": j.location or "",
            "delivery_zone": j.delivery_zone.name if j.delivery_zone else "",
            "subject": j.subject or "",
            "level": j.level or "",
            "curriculum": j.curriculum or "",
            "employment_type": j.employment_type or "",
            "preferred_sex": j.preferred_sex or "",
            "preferred_age_range": j.preferred_age_range or "",
            "description": j.description or "",
            "requirements": j.requirements or "",
            "responsibilities": j.responsibilities or "",
            "salary_min": float(j.salary_min) if j.salary_min is not None else "",
            "salary_max": float(j.salary_max) if j.salary_max is not None else "",
            "salary_currency": j.salary_currency or "",
            "deadline": str(j.deadline) if j.deadline else "",
            "status": j.status,
            "created_by_user_id": j.created_by_id or "",
            "created_at": j.created_at.isoformat() if j.created_at else "",
            "updated_at": j.updated_at.isoformat() if j.updated_at else "",
        }
        for j in rows
    ]

    if export_format == "csv":
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data_rows)
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=realmindx-jobs.csv"})

    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            return jsonify(error="XLSX export requires openpyxl."), 501
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Posts"
        ws.append(headers)
        for row in data_rows:
            ws.append([row[h] for h in headers])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return send_file(stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="realmindx-jobs.xlsx")

    if export_format == "pdf":
        try:
            return send_file(_build_admin_export_pdf("RealMindX Job Posts", headers, data_rows), mimetype="application/pdf", as_attachment=True, download_name="realmindx-jobs.pdf")
        except ImportError:
            return jsonify(error="PDF export requires reportlab."), 501

    return jsonify(error="Unsupported format. Use csv, xlsx, or pdf."), 400


@admin_bp.get("/applications/export")
@login_required
@permission_required("applications.export")
def export_applications():
    """Export job applications as CSV, XLSX, or PDF."""
    export_format = (request.args.get("format") or "csv").lower()
    rows = (
        JobApplication.query
        .join(JobApplication.user, isouter=True)
        .join(JobApplication.job, isouter=True)
        .order_by(JobApplication.created_at.desc())
        .all()
    )
    headers = [
        "id", "job_id", "job_title", "job_organisation", "job_location", "job_subject", "job_level",
        "applicant_id", "applicant_name", "applicant_email", "applicant_phone", "applicant_sex",
        "applicant_age_range", "applicant_location", "teaching_subject", "preferred_level",
        "preferred_employment_type", "available_from", "years_of_experience", "curriculum_experience",
        "preferred_locations", "status", "cover_note", "cv_filename", "certificate_filename", "applied_at", "updated_at",
    ]

    data_rows = [
        {
            "id": a.id,
            "job_id": a.job_id,
            "job_title": a.job.title if a.job else "",
            "job_organisation": a.job.organisation if a.job else "",
            "job_location": a.job.location if a.job else "",
            "job_subject": a.job.subject if a.job else "",
            "job_level": a.job.level if a.job else "",
            "applicant_id": a.user_id,
            "applicant_name": f"{a.user.first_name or ''} {a.user.last_name or ''}".strip() if a.user else "",
            "applicant_email": a.user.email if a.user else "",
            "applicant_phone": a.user.phone if a.user else "",
            "applicant_sex": a.user.sex if a.user else "",
            "applicant_age_range": a.user.age_range if a.user else "",
            "applicant_location": a.user.profile.location if a.user and a.user.profile else "",
            "teaching_subject": a.user.profile.teaching_subject if a.user and a.user.profile else "",
            "preferred_level": a.user.profile.preferred_level if a.user and a.user.profile else "",
            "preferred_employment_type": a.user.profile.preferred_employment_type if a.user and a.user.profile else "",
            "available_from": a.user.profile.available_from if a.user and a.user.profile else "",
            "years_of_experience": a.user.profile.years_of_experience if a.user and a.user.profile else "",
            "curriculum_experience": a.user.profile.curriculum_experience if a.user and a.user.profile else "",
            "preferred_locations": a.user.profile.preferred_locations if a.user and a.user.profile else "",
            "status": a.status,
            "cover_note": a.cover_note or "",
            "cv_filename": getattr(db.session.get(UploadedFile, a.cv_file_id), "original_filename", "") if a.cv_file_id else "",
            "certificate_filename": getattr(db.session.get(UploadedFile, a.certificate_file_id), "original_filename", "") if a.certificate_file_id else "",
            "applied_at": a.created_at.isoformat() if a.created_at else "",
            "updated_at": a.updated_at.isoformat() if a.updated_at else "",
        }
        for a in rows
    ]

    if export_format == "csv":
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data_rows)
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=realmindx-applications.csv"})

    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            return jsonify(error="XLSX export requires openpyxl."), 501
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(headers)
        for row in data_rows:
            ws.append([row[h] for h in headers])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return send_file(stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="realmindx-applications.xlsx")

    if export_format == "pdf":
        try:
            return send_file(_build_admin_export_pdf("RealMindX Job Applications", headers, data_rows), mimetype="application/pdf", as_attachment=True, download_name="realmindx-applications.pdf")
        except ImportError:
            return jsonify(error="PDF export requires reportlab."), 501

    return jsonify(error="Unsupported format. Use csv, xlsx, or pdf."), 400


@admin_bp.get("/categories")
@login_required
@permission_required("categories.view")
def categories():
    rows = ProductCategory.query.order_by(ProductCategory.sort_order.asc(), ProductCategory.name.asc()).all()
    return jsonify(items=[{"id": row.id, "name": row.name, "slug": row.slug, "description": row.description,
                            "is_active": row.is_active, "bulk_discount_percent": float(row.bulk_discount_percent or 0),
                            "bulk_min_qty": int(row.bulk_min_qty or 10)} for row in rows])


def _category_bulk_payload(payload, category=None):
    current_discount = float(getattr(category, "bulk_discount_percent", 0) or 0) if category else 0
    current_min_qty = int(getattr(category, "bulk_min_qty", 10) or 10) if category else 10
    discount = _decimalish(payload.get("bulk_discount_percent"), current_discount)
    min_qty = _intish(payload.get("bulk_min_qty"), current_min_qty)
    if discount < 0 or discount > 100:
        raise ValueError("Bulk discount percent must be between 0 and 100.")
    if min_qty < 1:
        raise ValueError("Bulk minimum quantity must be at least 1.")
    return discount, min_qty


@admin_bp.post("/categories")
@login_required
@permission_required("categories.create")
def create_category():
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or payload.get("label") or "").strip()
    if not name:
        return jsonify(error="Category name is required."), 400
    try:
        bulk_discount_percent, bulk_min_qty = _category_bulk_payload(payload)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    category = ProductCategory(
        name=name,
        slug=payload.get("slug") or slugify(name),
        description=payload.get("description"),
        sort_order=_intish(payload.get("sort_order"), 0),
        is_active=_boolish(payload.get("is_active", True)),
        bulk_discount_percent=bulk_discount_percent,
        bulk_min_qty=bulk_min_qty,
    )
    db.session.add(category)
    db.session.flush()
    log_action("create_category", "product_category", category.id)
    db.session.commit()
    return jsonify(id=category.id, name=category.name, slug=category.slug), 201


@admin_bp.put("/categories/<int:category_id>")
@login_required
@permission_required("categories.edit")
def update_category(category_id):
    category = db.get_or_404(ProductCategory, category_id)
    payload = request.get_json(silent=True) or {}
    for field in ["name", "slug", "description", "is_active"]:
        if field in payload:
            setattr(category, field, payload[field])
    if "sort_order" in payload:
        category.sort_order = _intish(payload.get("sort_order"), category.sort_order or 0)
    if "bulk_discount_percent" in payload or "bulk_min_qty" in payload:
        try:
            category.bulk_discount_percent, category.bulk_min_qty = _category_bulk_payload(payload, category)
        except ValueError as exc:
            return jsonify(error=str(exc)), 400
    log_action("update_category", "product_category", category.id)
    db.session.commit()
    return jsonify(id=category.id, name=category.name, slug=category.slug)


@admin_bp.delete("/categories/<int:category_id>")
@login_required
@permission_required("categories.delete")
def delete_category(category_id):
    category = db.get_or_404(ProductCategory, category_id)
    log_action("delete_category", "product_category", category.id, {"name": category.name})
    db.session.delete(category)
    db.session.commit()
    return jsonify(message="Category deleted.")


@admin_bp.get("/delivery-zones")
@login_required
@permission_required("deliveryZones.view")
def delivery_zones():
    rows = DeliveryZone.query.order_by(DeliveryZone.sort_order.asc(), DeliveryZone.name.asc()).all()
    return jsonify(items=[delivery_zone_json(row) for row in rows])


def _delivery_zone_payload(zone, payload):
    name = (payload.get("name") or (zone.name if zone else "")).strip()
    if not name:
        raise ValueError("Delivery zone name is required.")
    duplicate = DeliveryZone.query.filter(func.lower(DeliveryZone.name) == name.lower())
    if zone:
        duplicate = duplicate.filter(DeliveryZone.id != zone.id)
    if duplicate.first():
        raise ValueError("A delivery zone with this name already exists.")
    raw_aliases = payload.get("aliases", payload.get("aliases_text", None))

    def text_value(field):
        if field in payload:
            return (str(payload.get(field) or "").strip() or None)
        return getattr(zone, field, None) if zone else None

    def numeric_value(field, default=0):
        if field in payload:
            return _decimalish(payload.get(field), default)
        return float(getattr(zone, field, default) or default) if zone else default

    def integer_value(field, default=0):
        if field in payload:
            return int(payload.get(field) or default)
        return int(getattr(zone, field, default) or default) if zone else default

    def bool_value(field, default):
        if field in payload:
            return _boolish(payload.get(field))
        return bool(getattr(zone, field, default)) if zone else default

    values = {
        "name": name,
        "fee": numeric_value("fee"),
        "description": text_value("description"),
        "aliases": format_location_aliases(raw_aliases, name) if raw_aliases is not None else (zone.aliases if zone else None),
        "region": text_value("region"),
        "district_or_municipality": text_value("district_or_municipality"),
        "nearby_major_town": text_value("nearby_major_town"),
        "delivery_zone_label": text_value("delivery_zone_label"),
        "sort_order": integer_value("sort_order"),
        "is_active": bool_value("is_active", True),
        "is_delivery_area": bool_value("is_delivery_area", True),
        "is_search_alias_only": bool_value("is_search_alias_only", False),
    }
    return values


@admin_bp.post("/delivery-zones")
@login_required
@permission_required("deliveryZones.create")
def create_delivery_zone():
    payload = request.get_json(silent=True) or {}
    try:
        values = _delivery_zone_payload(None, payload)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    zone = DeliveryZone(**values)
    db.session.add(zone)
    db.session.flush()
    log_action("create_delivery_zone", "delivery_zone", zone.id)
    db.session.commit()
    return jsonify(delivery_zone=delivery_zone_json(zone)), 201


@admin_bp.put("/delivery-zones/<int:zone_id>")
@login_required
@permission_required("deliveryZones.edit")
def update_delivery_zone(zone_id):
    zone = db.get_or_404(DeliveryZone, zone_id)
    payload = request.get_json(silent=True) or {}
    try:
        values = _delivery_zone_payload(zone, payload)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    for field, value in values.items():
        setattr(zone, field, value)
    log_action("update_delivery_zone", "delivery_zone", zone.id)
    db.session.commit()
    return jsonify(delivery_zone=delivery_zone_json(zone))


@admin_bp.delete("/delivery-zones/<int:zone_id>")
@login_required
@permission_required("deliveryZones.delete")
def delete_delivery_zone(zone_id):
    zone = db.get_or_404(DeliveryZone, zone_id)
    log_action("delete_delivery_zone", "delivery_zone", zone.id, {"name": zone.name})
    db.session.delete(zone)
    db.session.commit()
    return jsonify(message="Delivery zone deleted.")


def _delivery_error_response(exc):
    return jsonify(error=exc.message, code=exc.code), exc.status_code


@admin_bp.get("/delivery-companies")
@login_required
@permission_required("delivery.view")
def delivery_companies():
    rows = DeliveryCompany.query.order_by(DeliveryCompany.name.asc()).all()
    return jsonify(items=[delivery_company_json(company) for company in rows])


@admin_bp.post("/delivery-companies")
@login_required
@permission_required("delivery.companies.manage")
def create_delivery_company():
    payload = request.get_json(silent=True) or {}
    try:
        company, manager = create_company(payload, actor=actor_from_user(current_user))
    except DeliveryError as exc:
        return _delivery_error_response(exc)
    temporary_password = getattr(manager, "_temporary_password", None) if manager else None
    notification = send_portal_access_notification(manager, "manager", temporary_password) if manager else None
    log_action("create_delivery_company", "delivery_company", company.id, {"name": company.name})
    db.session.commit()
    return jsonify(
        company=delivery_company_json(company),
        manager=delivery_company_user_json(manager) if manager else None,
        temporary_password=temporary_password,
        notification=notification,
    ), 201


@admin_bp.get("/delivery-companies/<int:company_id>")
@login_required
@permission_required("delivery.view")
def delivery_company_detail(company_id):
    company = db.get_or_404(DeliveryCompany, company_id)
    deliveries = (
        OrderDelivery.query
        .filter_by(company_id=company.id)
        .order_by(OrderDelivery.updated_at.desc())
        .limit(100)
        .all()
    )
    include_events = current_user.has_permission("delivery.audit.view")
    serialized_deliveries = []
    for delivery in deliveries:
        try:
            serialized_deliveries.append(delivery_json(delivery, include_events=include_events))
        except Exception as exc:
            current_app.logger.exception("Could not serialize delivery %s for company %s: %s", delivery.id, company.id, exc)
            serialized_deliveries.append({
                "id": delivery.id,
                "order_id": delivery.order_id,
                "order_reference": delivery.order.order_reference if delivery.order else None,
                "company_id": delivery.company_id,
                "company_name": company.name,
                "rider_id": delivery.rider_id,
                "rider_name": delivery.rider.name if delivery.rider else None,
                "status": delivery.status,
                "assigned_at": delivery.assigned_at.isoformat() if delivery.assigned_at else None,
                "updated_at": delivery.updated_at.isoformat() if delivery.updated_at else None,
                "serialization_warning": True,
            })
    return jsonify(
        company=delivery_company_json(company),
        managers=[delivery_company_user_json(user) for user in company.company_users],
        riders=[delivery_rider_json(rider) for rider in company.riders],
        deliveries=serialized_deliveries,
    )


@admin_bp.get("/delivery-riders/<int:rider_id>")
@login_required
@permission_required("delivery.view")
def admin_delivery_rider_detail(rider_id):
    rider = db.get_or_404(DeliveryRider, rider_id)
    deliveries = (
        OrderDelivery.query
        .filter_by(rider_id=rider.id)
        .order_by(OrderDelivery.updated_at.desc())
        .limit(200)
        .all()
    )
    include_events = current_user.has_permission("delivery.audit.view")
    return jsonify(
        rider=delivery_rider_json(rider),
        company=delivery_company_json(rider.company),
        deliveries=[delivery_json(delivery, include_events=include_events) for delivery in deliveries],
    )


@admin_bp.put("/delivery-companies/<int:company_id>")
@login_required
@permission_required("delivery.companies.manage")
def update_delivery_company(company_id):
    company = db.get_or_404(DeliveryCompany, company_id)
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or company.name).strip()
    if not name:
        return jsonify(error="Delivery company name is required."), 400
    duplicate = DeliveryCompany.query.filter(func.lower(DeliveryCompany.name) == name.lower(), DeliveryCompany.id != company.id).first()
    if duplicate:
        return jsonify(error="A delivery company with this name already exists."), 409
    company.name = name
    company.contact_name = (payload.get("contact_name") if "contact_name" in payload else company.contact_name) or None
    company.contact_email = (payload.get("contact_email") if "contact_email" in payload else company.contact_email) or None
    if "contact_phone" in payload:
        raw_phone = payload.get("contact_phone") or ""
        company.contact_phone = normalise_phone(raw_phone) if raw_phone else None
    company.notes = (payload.get("notes") if "notes" in payload else company.notes) or None
    if "default_delivery_payable" in payload:
        company.default_delivery_payable = payload.get("default_delivery_payable") or None
    if "is_active" in payload:
        company.is_active = _boolish(payload.get("is_active"))
        company.status = "active" if company.is_active else "inactive"
    if "status" in payload:
        company.status = str(payload.get("status") or company.status).strip() or company.status
        company.is_active = company.status == "active"
    log_action("update_delivery_company", "delivery_company", company.id, {"name": company.name})
    db.session.commit()
    return jsonify(company=delivery_company_json(company))


@admin_bp.post("/delivery-companies/<int:company_id>/managers")
@login_required
@permission_required("delivery.companies.manage")
def create_delivery_company_manager(company_id):
    company = db.get_or_404(DeliveryCompany, company_id)
    payload = request.get_json(silent=True) or {}
    try:
        manager = create_company_user(company, payload, actor=actor_from_user(current_user))
    except DeliveryError as exc:
        return _delivery_error_response(exc)
    temporary_password = getattr(manager, "_temporary_password", None)
    notification = send_portal_access_notification(manager, "manager", temporary_password)
    log_action("create_delivery_company_manager", "delivery_company_user", manager.id, {"company_id": company.id})
    db.session.commit()
    return jsonify(
        manager=delivery_company_user_json(manager),
        temporary_password=temporary_password,
        notification=notification,
    ), 201


@admin_bp.put("/delivery-company-users/<int:company_user_id>")
@login_required
@permission_required("delivery.companies.manage")
def update_delivery_company_user(company_user_id):
    company_user = db.get_or_404(DeliveryCompanyUser, company_user_id)
    payload = request.get_json(silent=True) or {}
    if "name" in payload:
        company_user.name = (payload.get("name") or company_user.name).strip()
        first, _, last = company_user.name.partition(" ")
        company_user.user.first_name = first or company_user.user.first_name
        company_user.user.last_name = last
    if "title" in payload:
        company_user.title = (payload.get("title") or "").strip() or None
    if "phone" in payload:
        phone = normalise_phone(payload.get("phone") or "")
        if not phone:
            return jsonify(error="Enter a valid Ghana phone number."), 400
        duplicate = DeliveryCompanyUser.query.filter(DeliveryCompanyUser.phone == phone, DeliveryCompanyUser.id != company_user.id).first()
        if duplicate:
            return jsonify(error="A company user already exists for this phone number."), 409
        company_user.phone = phone
        company_user.user.phone = phone
    if "is_active" in payload:
        company_user.is_active = _boolish(payload.get("is_active"))
        company_user.user.is_active = company_user.is_active
    log_action("update_delivery_company_user", "delivery_company_user", company_user.id)
    db.session.commit()
    return jsonify(manager=delivery_company_user_json(company_user))


@admin_bp.post("/delivery-company-users/<int:company_user_id>/reset-password")
@login_required
@permission_required("delivery.companies.manage")
def reset_delivery_company_user_password(company_user_id):
    company_user = db.get_or_404(DeliveryCompanyUser, company_user_id)
    try:
        temporary_password = reset_portal_password(company_user.user)
    except DeliveryError as exc:
        return _delivery_error_response(exc)
    notification = send_portal_access_notification(company_user, "manager", temporary_password)
    log_action("reset_delivery_company_user_password", "delivery_company_user", company_user.id)
    db.session.commit()
    return jsonify(
        message="Company manager password reset. They must change it on next login.",
        temporary_password=temporary_password,
        notification=notification,
    )


@admin_bp.get("/deliveries")
@login_required
@permission_required("delivery.view")
def admin_deliveries():
    status = (request.args.get("status") or "").strip()
    query = OrderDelivery.query.order_by(OrderDelivery.updated_at.desc())
    if status:
        query = query.filter_by(status=status)
    rows = query.limit(200).all()
    return jsonify(items=[delivery_json(delivery) for delivery in rows])


@admin_bp.get("/orders/<int:order_id>/delivery")
@login_required
@permission_required("delivery.view")
def admin_order_delivery(order_id):
    order = db.get_or_404(Order, order_id)
    delivery = getattr(order, "delivery", None)
    return jsonify(
        order=order_json(order),
        delivery=delivery_json(delivery, include_events=True) if delivery else None,
        contact_warning=staff_delivery_contact_warning(order),
        override_reasons=OTP_OVERRIDE_REASONS,
    )


@admin_bp.post("/orders/<int:order_id>/delivery/assign")
@login_required
@permission_required("delivery.assign")
@permission_required("orders.edit")
def admin_assign_order_delivery(order_id):
    order = db.get_or_404(Order, order_id)
    payload = request.get_json(silent=True) or {}
    company = db.session.get(DeliveryCompany, payload.get("company_id"))
    if not company:
        return jsonify(error="Choose a delivery company."), 400
    try:
        delivery = assign_order_to_company(
            order, company, actor_from_user(current_user), note=payload.get("note"),
            company_payable_amount=payload.get("company_payable_amount"),
            promotion_payer=payload.get("promotion_payer"), promotion_amount=payload.get("promotion_amount"),
        )
    except DeliveryError as exc:
        return _delivery_error_response(exc)
    log_action("assign_order_delivery_company", "order_delivery", delivery.id, {
        "order_reference": order.order_reference,
        "company_id": company.id,
    })
    db.session.commit()
    return jsonify(
        delivery=delivery_json(delivery, include_events=True),
        contact_warning=staff_delivery_contact_warning(order),
    )


@admin_bp.post("/deliveries/<int:delivery_id>/otp-resend")
@login_required
@permission_required("delivery.assign")
@permission_required("orders.edit")
def admin_resend_delivery_otp(delivery_id):
    delivery = db.get_or_404(OrderDelivery, delivery_id)
    try:
        resend_delivery_otp(delivery, actor_from_user(current_user))
    except DeliveryError as exc:
        return _delivery_error_response(exc)
    log_action("resend_delivery_otp", "order_delivery", delivery.id)
    db.session.commit()
    return jsonify(delivery=delivery_json(delivery, include_events=True))


@admin_bp.post("/deliveries/<int:delivery_id>/otp-override")
@login_required
@permission_required("delivery.override_otp")
@permission_required("orders.edit")
def admin_override_delivery_otp(delivery_id):
    delivery = db.get_or_404(OrderDelivery, delivery_id)
    payload = request.get_json(silent=True) or {}
    try:
        staff_override_otp(
            delivery,
            actor_from_user(current_user),
            payload.get("reason"),
            note=payload.get("note"),
        )
    except DeliveryError as exc:
        return _delivery_error_response(exc)
    log_action("override_delivery_otp", "order_delivery", delivery.id, {
        "reason": payload.get("reason"),
        "order_reference": delivery.order.order_reference if delivery.order else None,
    })
    db.session.commit()
    return jsonify(delivery=delivery_json(delivery, include_events=True))


@admin_bp.post("/deliveries/<int:delivery_id>/cancel")
@login_required
@permission_required("delivery.assign")
@permission_required("orders.edit")
def admin_cancel_delivery(delivery_id):
    delivery = db.get_or_404(OrderDelivery, delivery_id)
    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "cancelled_by_realmindx").strip()
    cancel_delivery(delivery, actor_from_user(current_user), reason=reason, note=payload.get("note"))
    if payload.get("cancel_order"):
        delivery.order.status = "cancelled"
    log_action("cancel_delivery_assignment", "order_delivery", delivery.id, {"reason": reason, "cancel_order": bool(payload.get("cancel_order"))})
    db.session.commit()
    return jsonify(delivery=delivery_json(delivery, include_events=True), order=order_json(delivery.order))


def _settlement_query():
    query = DeliverySettlementBatch.query.order_by(DeliverySettlementBatch.settlement_date.desc(), DeliverySettlementBatch.id.desc())
    company_id = request.args.get("company_id", type=int)
    status = (request.args.get("status") or "").strip()
    start = (request.args.get("start_date") or "").strip()
    end = (request.args.get("end_date") or "").strip()
    payment_method = (request.args.get("payment_method") or "").strip()
    if company_id: query = query.filter_by(company_id=company_id)
    if status: query = query.filter_by(status=status)
    if start: query = query.filter(DeliverySettlementBatch.settlement_date >= date.fromisoformat(start))
    if end: query = query.filter(DeliverySettlementBatch.settlement_date <= date.fromisoformat(end))
    rows = query.all()
    if payment_method:
        rows = [row for row in rows if any(line.payment_method == payment_method for line in row.lines)]
    return rows


@admin_bp.get("/delivery-settlements")
@login_required
@permission_required("delivery.settlements.view")
def admin_delivery_settlements():
    return jsonify(items=[batch_json(batch) for batch in _settlement_query()])


@admin_bp.get("/delivery-settlements/<int:batch_id>")
@login_required
@permission_required("delivery.settlements.view")
def admin_delivery_settlement_detail(batch_id):
    return jsonify(settlement=batch_json(db.get_or_404(DeliverySettlementBatch, batch_id), include_lines=True, include_events=True))


@admin_bp.post("/delivery-settlements/<int:batch_id>/adjust")
@login_required
@permission_required("delivery.settlements.adjust")
def admin_adjust_delivery_settlement(batch_id):
    payload = request.get_json(silent=True) or {}
    try: batch = apply_adjustment(db.get_or_404(DeliverySettlementBatch, batch_id), payload.get("amount"), payload.get("reason"), actor_from_user(current_user))
    except SettlementError as exc: return jsonify(error=exc.message, code=exc.code), exc.status_code
    db.session.commit()
    return jsonify(settlement=batch_json(batch, include_lines=True, include_events=True))


@admin_bp.post("/delivery-settlements/<int:batch_id>/mark-paid")
@login_required
@permission_required("delivery.settlements.mark_paid")
def admin_mark_delivery_settlement_paid(batch_id):
    payload = request.get_json(silent=True) or {}
    try:
        batch = mark_settled(db.get_or_404(DeliverySettlementBatch, batch_id), payload.get("payment_reference"), payload.get("payment_date"), actor_from_user(current_user), payload.get("payment_proof_url"))
    except (SettlementError, ValueError) as exc:
        return jsonify(error=getattr(exc, "message", "Enter a valid payment date.")), getattr(exc, "status_code", 400)
    db.session.commit()
    company_email = (batch.company.contact_email or "").strip() if batch.company else ""
    if company_email:
        portal_url = f"{current_app.config['DELIVERY_URL'].rstrip('/')}/manager/"
        try:
            send_email(
                OutboundEmail(
                    to=company_email,
                    subject=f"Delivery settlement confirmed: {batch.reference}",
                    html=app_email_shell(
                        "Delivery settlement confirmed",
                        f"<p>Settlement <strong>{escape(batch.reference)}</strong> has been marked settled.</p><p>Payment reference: <strong>{escape(batch.payment_reference)}</strong>.</p>",
                        cta_label="Open delivery company portal", cta_url=portal_url,
                    ),
                    text=f"Settlement {batch.reference} has been marked settled. Payment reference: {batch.payment_reference}. {portal_url}",
                ),
                purpose="transactional",
                recipient_user_id=None,
                template_name="delivery_settlement_paid",
            )
        except Exception:
            current_app.logger.exception("Could not send settlement confirmation for %s", batch.reference)
    return jsonify(settlement=batch_json(batch, include_lines=True, include_events=True))


@admin_bp.post("/delivery-settlements/<int:batch_id>/resolve-dispute")
@login_required
@permission_required("delivery.settlements.dispute_resolve")
def admin_resolve_delivery_settlement_dispute(batch_id):
    payload = request.get_json(silent=True) or {}
    try: batch = resolve_dispute(db.get_or_404(DeliverySettlementBatch, batch_id), payload.get("note"), actor_from_user(current_user))
    except SettlementError as exc: return jsonify(error=exc.message, code=exc.code), exc.status_code
    db.session.commit()
    return jsonify(settlement=batch_json(batch, include_lines=True, include_events=True))


def _settlement_export_response(batch, export_format):
    rows = [line_json(line) for line in batch.lines]
    for row in rows:
        row.update(
            batch_reference=batch.reference,
            batch_status=batch.status,
            batch_payment_reference=batch.payment_reference,
            batch_payment_date=batch.payment_date.isoformat() if batch.payment_date else "",
            batch_adjustment_amount=float(batch.adjustment_amount or 0),
            batch_adjustment_reason=batch.adjustment_reason or "",
            dispute_status=batch.dispute_status,
            dispute_notes=batch.dispute_notes or "",
            resolution_notes=batch.resolution_notes or "",
        )
    headers = ["id", "batch_id", "batch_reference", "settlement_date", "batch_status", "order_id", "order_reference", "delivery_id", "company_id", "company_name", "rider_id", "rider_name", "customer_name", "delivery_location", "payment_method", "book_subtotal", "customer_delivery_fee", "company_payable", "promotion_amount", "promotion_payer", "amount_collected_realmindx", "amount_collected_company", "amount_due_realmindx", "amount_due_company", "net_balance", "adjustment_amount", "adjustment_reason", "batch_adjustment_amount", "batch_adjustment_reason", "status", "delivered_at", "created_at", "batch_payment_reference", "batch_payment_date", "dispute_status", "dispute_notes", "resolution_notes"]
    if export_format == "csv":
        output = io.StringIO(); writer = csv.DictWriter(output, fieldnames=headers); writer.writeheader()
        writer.writerows([{key: row.get(key) for key in headers} for row in rows])
        return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={batch.reference}.csv"})
    if export_format == "xlsx":
        try: from openpyxl import Workbook
        except ImportError: return jsonify(error="XLSX export requires openpyxl."), 501
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Settlement"; sheet.append(headers)
        for row in rows: sheet.append([row.get(key) for key in headers])
        stream = io.BytesIO(); workbook.save(stream); stream.seek(0)
        return send_file(stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"{batch.reference}.xlsx")
    if export_format == "pdf":
        try:
            return send_file(_build_admin_export_pdf(f"RealMindX Delivery Settlement {batch.reference}", headers, rows), mimetype="application/pdf", as_attachment=True, download_name=f"{batch.reference}.pdf")
        except ImportError: return jsonify(error="PDF export requires reportlab."), 501
    if export_format == "zip":
        try: from openpyxl import Workbook
        except ImportError: return jsonify(error="ZIP export requires openpyxl."), 501
        try: pdf_export = _build_admin_export_pdf(f"RealMindX Delivery Settlement {batch.reference}", headers, rows)
        except ImportError: return jsonify(error="ZIP export requires reportlab."), 501
        zip_stream = io.BytesIO()
        with zipfile.ZipFile(zip_stream, "w", zipfile.ZIP_DEFLATED) as zf:
            output = io.StringIO(); writer = csv.DictWriter(output, fieldnames=headers); writer.writeheader(); writer.writerows([{key: row.get(key) for key in headers} for row in rows])
            zf.writestr(f"{batch.reference}.csv", output.getvalue())
            workbook = Workbook(); sheet = workbook.active; sheet.title = "Settlement"; sheet.append(headers)
            for row in rows: sheet.append([row.get(key) for key in headers])
            xlsx_stream = io.BytesIO(); workbook.save(xlsx_stream); xlsx_stream.seek(0); zf.writestr(f"{batch.reference}.xlsx", xlsx_stream.getvalue())
            zf.writestr(f"{batch.reference}.pdf", pdf_export.getvalue())
        zip_stream.seek(0)
        return send_file(zip_stream, mimetype="application/zip", as_attachment=True, download_name=f"{batch.reference}.zip")
    return jsonify(error="Use csv, xlsx, pdf, or zip."), 400


@admin_bp.get("/delivery-settlements/<int:batch_id>/export/<string:export_format>")
@login_required
@permission_required("delivery.settlements.export")
def admin_export_delivery_settlement(batch_id, export_format):
    batch = db.get_or_404(DeliverySettlementBatch, batch_id)
    log_action("settlement_exported", "delivery_settlement", batch.id, {"format": export_format})
    db.session.commit()
    return _settlement_export_response(batch, export_format)


@admin_bp.get("/orders")
@login_required
@permission_required("orders.view")
def orders():
    rows = placed_order_query().order_by(Order.created_at.desc()).all()
    return jsonify(items=[order_json(order) for order in rows])


def _ledger_iso(value):
    return value.isoformat() if value else None


def _ledger_amount(value):
    return float(value or 0)


def _cart_invoice_status(invoice):
    if invoice.converted_at or invoice.converted_order_id:
        return "converted"
    if invoice.emailed_at:
        return "emailed"
    return invoice.status or "generated"


def _receipt_invoice_order_row(order):
    document_id = order.invoice_id or order.order_reference
    return {
        "id": f"receipt-{order.id}",
        "record_id": order.id,
        "document_type": "receipt",
        "document_label": "Receipt",
        "document_id": document_id,
        "lookup_id": document_id,
        "order_reference": order.order_reference,
        "customer_name": order.customer_name,
        "email": order.email,
        "phone": order.phone,
        "recipients": [order.email] if order.email else [],
        "status": normalize_order_status(order.status),
        "payment_status": order.payment_status or "",
        "source": "bookshop_order",
        "subtotal_amount": _ledger_amount(order.subtotal_amount),
        "delivery_fee": _ledger_amount(order.delivery_fee),
        "total_amount": _ledger_amount(order.total_amount),
        "created_at": _ledger_iso(order.created_at),
        "issued_at": _ledger_iso(order.paid_at or order.created_at),
        "emailed_at": None,
        "viewed_at": None,
        "converted_at": _ledger_iso(order.created_at),
        "converted_order_id": order.id,
        "linked_cart_invoice_id": order.cart_invoice.invoice_id if order.cart_invoice else None,
        "item_count": len(order.items or []),
        "pdf_document": "receipt",
    }


def _receipt_invoice_cart_row(invoice):
    recipients = [str(email) for email in (invoice.recipients or []) if email]
    return {
        "id": f"cart-invoice-{invoice.id}",
        "record_id": invoice.id,
        "document_type": "cart_invoice",
        "document_label": "Cart Invoice",
        "document_id": invoice.invoice_id,
        "lookup_id": invoice.invoice_id,
        "order_reference": "",
        "customer_name": "Cart invoice",
        "email": recipients[0] if recipients else "",
        "phone": "",
        "recipients": recipients,
        "status": _cart_invoice_status(invoice),
        "payment_status": "not_applicable",
        "source": "cart_invoice",
        "subtotal_amount": _ledger_amount(invoice.subtotal_amount),
        "delivery_fee": _ledger_amount(invoice.delivery_fee),
        "total_amount": _ledger_amount(invoice.total_amount),
        "created_at": _ledger_iso(invoice.created_at),
        "issued_at": _ledger_iso(invoice.created_at),
        "emailed_at": _ledger_iso(invoice.emailed_at),
        "viewed_at": _ledger_iso(invoice.viewed_at),
        "converted_at": _ledger_iso(invoice.converted_at),
        "converted_order_id": invoice.converted_order_id,
        "linked_cart_invoice_id": invoice.invoice_id,
        "item_count": len(invoice.items or []),
        "pdf_document": "",
    }


@admin_bp.get("/receipts-invoices")
@login_required
@permission_required("orders.view")
def receipts_invoices():
    orders = (
        placed_order_query()
        .options(joinedload(Order.items), joinedload(Order.cart_invoice))
        .order_by(Order.created_at.desc())
        .limit(300)
        .all()
    )
    invoices = (
        CartInvoice.query
        .options(joinedload(CartInvoice.items))
        .order_by(CartInvoice.created_at.desc())
        .limit(300)
        .all()
    )
    rows = [_receipt_invoice_order_row(order) for order in orders]
    rows.extend(_receipt_invoice_cart_row(invoice) for invoice in invoices)

    query = (request.args.get("q") or "").strip().lower()
    document_type = (request.args.get("type") or "all").strip()
    status = (request.args.get("status") or "all").strip().lower()

    if document_type != "all":
        rows = [row for row in rows if row["document_type"] == document_type]
    if status != "all":
        rows = [row for row in rows if str(row["status"]).lower() == status]
    if query:
        rows = [
            row for row in rows
            if query in " ".join([
                str(row.get("document_id") or ""),
                str(row.get("order_reference") or ""),
                str(row.get("customer_name") or ""),
                str(row.get("email") or ""),
                " ".join(row.get("recipients") or []),
                str(row.get("source") or ""),
                str(row.get("status") or ""),
            ]).lower()
        ]

    rows.sort(key=lambda row: row.get("created_at") or "", reverse=True)
    summary = {
        "total": len(rows),
        "receipts": sum(1 for row in rows if row["document_type"] == "receipt"),
        "cart_invoices": sum(1 for row in rows if row["document_type"] == "cart_invoice"),
        "converted": sum(1 for row in rows if row["status"] == "converted"),
        "emailed": sum(1 for row in rows if row["status"] == "emailed"),
    }
    return jsonify(items=rows[:400], summary=summary)


@admin_bp.get("/orders/export")
@login_required
@permission_required("orders.export")
def export_orders():
    export_format = (request.args.get("format") or "csv").lower()
    rows = placed_order_query().order_by(Order.created_at.desc()).all()
    headers = [
        "id",
        "order_reference",
        "customer_name",
        "customer_sex",
        "customer_age_range",
        "email",
        "phone",
        "invoice_id",
        "payment_reference",
        "delivery_method",
        "delivery_zone_name",
        "delivery_region",
        "location",
        "payment_method",
        "payment_provider",
        "payment_status",
        "status",
        "subtotal_amount",
        "bulk_discount_amount",
        "promo_code",
        "promo_applies_to",
        "promo_discount_amount",
        "delivery_fee",
        "total_amount",
        "paid_at",
        "items",
        "notes",
        "created_at",
        "updated_at",
    ]
    data_rows = [
        {
            "id": order.id,
            "order_reference": order.order_reference,
            "customer_name": order.customer_name,
            "customer_sex": order.customer_sex or "",
            "customer_age_range": order.customer_age_range or "",
            "email": order.email,
            "phone": order.phone,
            "invoice_id": order.invoice_id or "",
            "payment_reference": order.payment_reference or "",
            "delivery_method": order.delivery_method,
            "delivery_zone_name": order.delivery_zone_name or "",
            "delivery_region": order.delivery_region or "",
            "location": order.location or "",
            "payment_method": order.payment_method or "",
            "payment_provider": order.payment_provider or "",
            "payment_status": order.payment_status or "",
            "status": order.status,
            "subtotal_amount": float(order.subtotal_amount or 0) if order.subtotal_amount is not None else "",
            "bulk_discount_amount": float(order.bulk_discount_amount or 0),
            "promo_code": order.promo_code or "",
            "promo_applies_to": order.promo_applies_to or "",
            "promo_discount_amount": float(order.promo_discount_amount or 0),
            "delivery_fee": float(order.delivery_fee or 0),
            "total_amount": float(order.total_amount or 0) if order.total_amount is not None else "",
            "paid_at": str(order.paid_at or ""),
            "items": "; ".join(
                f"{item.product_name} x{item.quantity} @ GHS {float(item.unit_price or 0):.2f}"
                for item in order.items
            ),
            "notes": order.notes or "",
            "created_at": str(order.created_at.date()) if order.created_at else "",
            "updated_at": str(order.updated_at.date()) if order.updated_at else "",
        }
        for order in rows
    ]

    if export_format == "csv":
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data_rows)
        return Response(
            out.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=realmindx-orders.csv"},
        )

    if export_format == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            return jsonify(error="XLSX export requires openpyxl."), 501
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Orders"
        sheet.append(headers)
        for row in data_rows:
            sheet.append([row[header] for header in headers])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="realmindx-orders.xlsx",
        )

    if export_format == "pdf":
        try:
            stream = _build_admin_export_pdf("RealMindX Bookshop Orders", headers, data_rows)
        except ImportError:
            return jsonify(error="PDF export requires reportlab."), 501
        return send_file(stream, mimetype="application/pdf", as_attachment=True, download_name="realmindx-orders.pdf")

    return jsonify(error="Unsupported format. Use csv, xlsx, or pdf."), 400


@admin_bp.put("/orders/<int:order_id>/status")
@login_required
@permission_required("orders.edit")
def update_order_status(order_id):
    order = db.get_or_404(Order, order_id)
    payload = request.get_json(silent=True) or {}
    raw_status = payload.get("status")
    if raw_status in {None, ""}:
        return jsonify(error="Select a valid order status."), 400
    status = normalize_order_status(raw_status)
    delivery = getattr(order, "delivery", None)
    if delivery and delivery.company_id:
        return jsonify(
            error="This order is managed by an external delivery company. Use delivery actions instead of the manual order status control.",
            code="external_delivery_status_managed",
        ), 409
    cancel_reason = (payload.get("cancel_reason") or "").strip()
    valid_statuses = {"new", "confirmed", "shipped", "complete", "cancelled", "archived"}
    if status not in valid_statuses:
        return jsonify(error="Invalid order status."), 400
    old_status = normalize_order_status(order.status)
    order.status = status
    if cancel_reason:
        order.notes = cancel_reason
    if status == "cancelled" and getattr(order, "delivery", None):
        try:
            cancel_delivery(order.delivery, actor_from_user(current_user), reason=cancel_reason or "order_cancelled")
        except DeliveryError:
            pass
    promo_snapshot = None
    if status == "complete" and old_status != "complete":
        usage, created = record_completed_order_promo_usage(order)
        if created and usage and usage.affiliate_email and getattr(usage.promo_code, "affiliate_notify_on_use", True):
            promo_snapshot = usage_snapshot(usage)
    log_action("update_order_status", "order", order.id, {"status": status, "prev": old_status})
    db.session.commit()

    if status != old_status:
        snapshot = _order_contact_snapshot(order)
        if snapshot.email:
            _run_in_background("Order status email", _send_order_status_email, snapshot, status, cancel_reason)
        if snapshot.phone:
            _run_in_background("Order status SMS", _send_order_status_sms, snapshot, status, cancel_reason)
    if promo_snapshot:
        _run_in_background("Promo commission email", send_promo_usage_notification, promo_snapshot)

    return jsonify(order=order_json(order))


def _send_order_status_sms(order, status, cancel_reason=""):
    """Send an SMS to the customer when their order status changes."""
    from ..sms_service import send_sms
    if not order.phone:
        return
    first_name = (order.customer_name or "").split()[0] or "there"
    ref = order.order_reference
    status = normalize_order_status(status)
    messages = {
        "new": (
            f"Hi {first_name}, your RealMindX Bookshop order {ref} has been placed. "
            f"Our team will contact you within 1 business day to arrange receipt of your package."
        ),
        "confirmed": (
            f"Hi {first_name}, your RealMindX Bookshop order {ref} is confirmed "
            f"and being packaged. We will contact you shortly with any final receipt details."
        ),
        "shipped": (
            f"Hi {first_name}, great news! Your order {ref} is on its way. "
            f"Expected delivery within 48 working hours. - RealMindX Bookshop"
        ),
        "complete": (
            f"Hi {first_name}, your order {ref} has been delivered. "
            f"Thank you for choosing RealMindX Bookshop. Please check your email to rate your experience."
        ),
        "cancelled": (
            f"Hi {first_name}, your order {ref} has been cancelled."
            + (f" Reason: {cancel_reason}." if cancel_reason else "")
            + " Contact us on WhatsApp for help. - RealMindX Bookshop"
        ),
    }
    msg = messages.get(status)
    if msg:
        send_sms(
            order.phone,
            msg,
            purpose="transactional",
            recipient_user_id=order.user_id,
            template_name=f"bookshop_order_status_{status}",
        )


def _send_order_status_email(order, status, cancel_reason=""):
    """Send a friendly, branded email when order status changes."""
    first_name = (order.customer_name or "").split()[0] or "there"
    ref = order.order_reference
    base_url = current_app.config["BASE_URL"].rstrip("/")
    bookshop_url = current_app.config.get("BOOKSHOP_URL", f"{base_url}/bookshop").rstrip("/")
    status = normalize_order_status(status)
    delivery_info = (
        "Pickup from our Dome Pillar 2 shop"
        if order.delivery_method == "pickup"
        else f"Delivery to {order.location or 'the address on file'}"
    )
    payment_info = (
        "Payment on delivery"
        if order.payment_method == "cash_on_delivery"
        else "Online payment via Paystack"
    )
    order_meta_html = f"""
    <div style="background:#f5f8fc;border:1px solid #dce5f0;border-radius:12px;padding:16px 20px;margin:18px 0;">
      <p style="margin:0 0 6px;"><strong>Reference:</strong> {escape(ref)}</p>
      <p style="margin:0 0 6px;"><strong>Fulfilment:</strong> {escape(delivery_info)}</p>
      <p style="margin:0 0 6px;"><strong>Payment:</strong> {escape(payment_info)}</p>
      <p style="margin:0;"><strong>Contact number:</strong> {escape(order.phone or "not provided")}</p>
    </div>
    """
    order_summary_html = bookshop_order_summary_table(order)
    feedback_url = f"{bookshop_url}/review?ref={escape(ref, quote=True)}"
    feedback_rows = (
        "<tr>"
        + "".join(
            f'<td style="padding:0 2px 8px;"><a href="{feedback_url}&score={score}" '
            f'style="display:inline-block;min-width:28px;padding:9px 0;border-radius:7px;'
            f'background:#143670;color:#ffffff;font-weight:800;font-family:Arial,Helvetica,sans-serif;'
            f'font-size:13px;text-decoration:none;text-align:center;">{score}</a></td>'
            for score in range(1, 11)
        )
        + "</tr>"
    )
    feedback_scale_html = f"""
    <div style="margin:22px 0 6px;">
      <p style="margin:0 0 12px;font-weight:700;color:#143670;">How likely are you to recommend RealMindX Bookshop to others?</p>
      <table role="presentation" cellspacing="0" cellpadding="0" style="margin:0 auto;">
        {feedback_rows}
      </table>
      <p style="margin:8px 0 0;color:#53657d;font-size:13px;">1 = Not likely at all, 10 = Extremely likely.</p>
    </div>
    """

    status_messages = {
        "new": {
            "subject": f"Your RealMindX Bookshop order has been placed: {ref}",
            "title": "Your order has been placed!",
            "body": (
                f"<p>Hello {escape(first_name)},</p>"
                f"<p>Your order <strong>{escape(ref)}</strong> has been placed successfully and our team is reviewing it now.</p>"
                f"{order_meta_html}"
                f"{order_summary_html}"
                f"<p>Our team will contact you within <strong>1 business day</strong> to confirm your order.</p>"
                f"<p>If you need anything sooner, contact us on any of the channels below and we&rsquo;ll help right away.</p>"
            ),
            "cta_label": "Track Your Order",
            "cta_url": "track",
        },
        "confirmed": {
            "subject": f"Your RealMindX Bookshop order is confirmed: {ref}",
            "title": "Your order is confirmed!",
            "body": (
                f"<p>Hello {escape(first_name)},</p>"
                f"<p>Great news! Your order <strong>{escape(ref)}</strong> has been confirmed and our team is getting it ready for you.</p>"
                f"{order_meta_html}"
                f"{order_summary_html}"
                f"<p>Our team will contact you when your order is ready to ensure you receive your order.</p>"
                f"<p>In the meantime, feel free to reach us through any of the channels below if you have any questions or concerns.</p>"
            ),
            "cta_label": "Track Your Order",
            "cta_url": "track",
        },
        "shipped": {
            "subject": f"Your RealMindX order is on its way: {ref}",
            "title": "Your order is on its way!",
            "body": (
                f"<p>Hello {escape(first_name)},</p>"
                f"<p>Good news! Your order <strong>{escape(ref)}</strong> has been dispatched and is heading your way.</p>"
                f"{order_meta_html}"
                f"{order_summary_html}"
                f"<p>Expected delivery is within 48 working hours. Our team will contact you if any final handover details are needed.</p>"
            ),
            "cta_label": "Track Your Order",
            "cta_url": "track",
        },
        "complete": {
            "subject": f"Order delivered. Thank you, {escape(first_name)}!",
            "title": "Order delivered. Thank you!",
            "body": (
                f"<p>Hello {escape(first_name)},</p>"
                f"<p>Your order <strong>{escape(ref)}</strong> has been marked as delivered. We hope you&rsquo;re pleased with our service.</p>"
                f"{order_meta_html}"
                f"{order_summary_html}"
                f"<p>If anything is missing or not as expected, please contact us through our channels below and we&rsquo;ll investigate and rectify.</p>"
                f"{feedback_scale_html}"
                f"<p>Thank you for choosing RealMindX Bookshop. We look forward to serving you again.</p>"
            ),
            "cta_label": "Rate Your Experience",
            "cta_url": feedback_url,
        },
        "cancelled": {
            "subject": f"Your RealMindX order {ref} has been cancelled",
            "title": "Order cancelled",
            "body": (
                f"<p>Hello {escape(first_name)},</p>"
                f"<p>We&rsquo;re sorry to let you know that your order <strong>{escape(ref)}</strong> has been cancelled.</p>"
                + order_summary_html
                + (f"<p><strong>Reason:</strong> {escape(cancel_reason)}</p>" if cancel_reason else "")
                + "<p>If you believe this is an error or would like to place a new order, please reach out to us and we&rsquo;ll be happy to help.</p>"
            ),
            "cta_label": "Contact Us",
            "cta_url": f"{base_url}/contact",
        },
    }

    info = status_messages.get(status)
    if not info:
        return  # no email for other status changes (archived, etc.)

    try:
        attachments = []
        if status == "complete":
            try:
                from ..invoices import build_receipt_pdf

                receipt_stream = build_receipt_pdf(order)
                attachments.append(EmailAttachment(
                    filename=f"{ref}-receipt.pdf",
                    content=receipt_stream.getvalue(),
                    content_type="application/pdf",
                ))
            except Exception:
                current_app.logger.exception("Could not build receipt PDF for order %s.", ref)

        send_email(
            OutboundEmail(
                to=order.email,
                from_email=current_app.config["BOOKSHOP_FROM_EMAIL"],
                subject=info["subject"],
                html=bookshop_email_shell(
                    info["title"],
                    info["body"],
                    cta_label=info.get("cta_label"),
                    cta_url=info.get("cta_url"),
                    eyebrow="RealMindX Bookshop",
                    preheader=info["subject"],
                ),
                attachments=attachments,
            ),
            purpose="transactional",
            recipient_user_id=order.user_id,
            template_name=f"bookshop_order_status_{status}",
        )
    except Exception as exc:
        current_app.logger.warning("Order status email failed (error=%s)", type(exc).__name__)


@admin_bp.delete("/orders/<int:order_id>")
@login_required
@permission_required("orders.delete")
def delete_order(order_id):
    order = db.get_or_404(Order, order_id)
    if DeliverySettlementLine.query.filter_by(order_id=order.id).first():
        return jsonify(error="This order has settlement history and cannot be permanently deleted."), 409
    normalized_email = normalize_contact_email(order.email)
    BookshopPaymentIntent.query.filter_by(order_id=order.id).update({"order_id": None})
    CartInvoice.query.filter_by(converted_order_id=order.id).update({"converted_order_id": None})
    ProductReview.query.filter_by(order_id=order.id).update({"order_id": None})
    log_action("delete_order", "order", order.id, {"order_reference": order.order_reference})
    db.session.delete(order)
    db.session.flush()
    remaining_orders = Order.query.filter(
        func.lower(Order.email) == normalized_email,
        Order.id != order_id,
    ).order_by(Order.created_at.asc()).all()
    contact = Contact.query.filter_by(email=normalized_email).first()
    if contact:
        if remaining_orders:
            source = ContactSource.query.filter_by(contact_id=contact.id, source="bookshop").first()
            if source:
                source.first_seen_at = remaining_orders[0].created_at
                source.last_seen_at = remaining_orders[-1].updated_at or remaining_orders[-1].created_at
                source.details = {
                    **(source.details or {}),
                    "latest_order_id": remaining_orders[-1].id,
                    "latest_order_reference": remaining_orders[-1].order_reference,
                }
        else:
            remove_contact_source(contact, "bookshop")
    db.session.commit()
    return jsonify(message="Order deleted.")


def _content_payload(model, payload, fields):
    row = model()
    for field in fields:
        if field in payload:
            setattr(row, field, payload[field])
    return row


def _clean_news_sections(sections):
    cleaned = []
    if not isinstance(sections, list):
        return cleaned
    for section in sections:
        if not isinstance(section, dict):
            continue
        heading = (section.get("heading") or "").strip()
        body = (section.get("body") or "").strip()
        caption = (section.get("caption") or "").strip()
        image_position = (section.get("image_position") or "auto").strip().lower()
        image_size = (section.get("image_size") or "medium").strip().lower()
        if image_position not in {"auto", "left", "right", "full", "top", "bottom"}:
            image_position = "top"
        if image_size not in {"small", "medium", "large"}:
            image_size = "medium"
        image_file_id = section.get("image_file_id") or None
        if image_file_id in ("", "null", "None"):
            image_file_id = None
        if not heading and not body and not image_file_id and not caption:
            continue
        cleaned.append(
            {
                "heading": heading,
                "body": body,
                "caption": caption,
                "image_position": image_position,
                "image_size": image_size,
                "image_file_id": int(image_file_id) if str(image_file_id or "").isdigit() else None,
            }
        )
    return cleaned


def _enrich_news_sections(sections):
    rows = _clean_news_sections(sections)
    file_ids = {row["image_file_id"] for row in rows if row.get("image_file_id")}
    files = {
        uploaded.id: uploaded
        for uploaded in UploadedFile.query.filter(UploadedFile.id.in_(file_ids)).all()
    } if file_ids else {}
    enriched = []
    for row in rows:
        item = dict(row)
        uploaded = files.get(item.get("image_file_id"))
        item["image_url"] = _upload_public_url(uploaded) if uploaded else None
        enriched.append(item)
    return enriched


def _news_json(row):
    image_url = _upload_public_url(row.image_file) if row.image_file else None
    return {
        "id": row.id,
        "title": row.title,
        "slug": row.slug,
        "category": row.category,
        "summary": row.summary,
        "body": row.body,
        "sections": _enrich_news_sections(row.sections or []),
        "image_url": image_url,
        "image_file_id": row.image_file_id,
        "status": row.status,
        "date": str(row.display_date) if row.display_date else None,
        "created_at": row.created_at.isoformat(),
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        "published_at": row.published_at.isoformat() if row.published_at else None,
    }


@admin_bp.get("/news")
@login_required
@permission_required("news.view")
def list_news():
    rows = News.query.order_by(News.created_at.desc()).all()
    return jsonify(items=[_news_json(r) for r in rows])


@admin_bp.post("/news")
@login_required
@permission_required("news.create")
def create_news():
    payload = request.get_json(silent=True) or {}
    row = News(
        title=payload.get("title"),
        slug=payload.get("slug") or slugify(payload.get("title")),
        category=payload.get("category"),
        summary=payload.get("summary"),
        body=payload.get("body") or "",
        sections=_clean_news_sections(payload.get("sections")),
        image_file_id=payload.get("image_file_id") or None,
        status=payload.get("status") or "draft",
        display_date=date.fromisoformat(payload["date"]) if payload.get("date") else None,
    )
    if not row.title:
        return jsonify(error="Title is required."), 400
    db.session.add(row)
    db.session.flush()
    if row.status == "published" and not row.published_at:
        row.published_at = datetime.now(timezone.utc)
    log_action("create_news", "news", row.id)
    db.session.commit()
    return jsonify(_news_json(row)), 201


@admin_bp.put("/news/<int:news_id>")
@login_required
@permission_required("news.edit")
def update_news(news_id):
    row = db.get_or_404(News, news_id)
    payload = request.get_json(silent=True) or {}
    for field in ["title", "slug", "category", "summary", "body", "image_file_id", "status"]:
        if field in payload:
            setattr(row, field, payload[field])
    if "sections" in payload:
        row.sections = _clean_news_sections(payload.get("sections"))
    if "date" in payload:
        row.display_date = date.fromisoformat(payload["date"]) if payload["date"] else None
    if row.status == "published" and not row.published_at:
        row.published_at = datetime.now(timezone.utc)
    log_action("update_news", "news", row.id)
    db.session.commit()
    return jsonify(_news_json(row))


@admin_bp.delete("/news/<int:news_id>")
@login_required
@permission_required("news.delete")
def delete_news(news_id):
    row = db.get_or_404(News, news_id)
    log_action("delete_news", "news", row.id, {"title": row.title})
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="News post deleted.")


@admin_bp.get("/gallery")
@login_required
@permission_required("gallery.view")
def list_gallery():
    rows = GalleryItem.query.order_by(GalleryItem.sort_order.asc(), GalleryItem.created_at.desc()).limit(100).all()
    def _img(r):
        return f"/uploads/{r.image_file.visibility}/{r.image_file.category}/{r.image_file.stored_filename}" if r.image_file else None
    return jsonify(items=[{"id": r.id, "title": r.title, "description": r.description, "image_url": _img(r), "image_file_id": r.image_file_id, "is_published": r.is_published, "status": "published" if r.is_published else "draft", "created_at": r.created_at.isoformat()} for r in rows])


@admin_bp.post("/gallery")
@login_required
@permission_required("gallery.create")
def create_gallery_item():
    payload = request.get_json(silent=True) or {}
    row = GalleryItem(
        title=payload.get("title"),
        description=payload.get("description"),
        image_file_id=payload.get("image_file_id"),
        sort_order=payload.get("sort_order") or 0,
        is_published=bool(payload.get("is_published", payload.get("status") == "published")),
    )
    if not row.title:
        return jsonify(error="Title is required."), 400
    db.session.add(row)
    db.session.flush()
    log_action("create_gallery_item", "gallery_item", row.id)
    db.session.commit()
    return jsonify(id=row.id, title=row.title), 201


@admin_bp.put("/gallery/<int:item_id>")
@login_required
@permission_required("gallery.edit")
def update_gallery_item(item_id):
    row = db.get_or_404(GalleryItem, item_id)
    payload = request.get_json(silent=True) or {}
    for field in ["title", "description", "image_file_id", "sort_order", "is_published"]:
        if field in payload:
            setattr(row, field, payload[field])
    if "status" in payload:
        row.is_published = payload["status"] == "published"
    log_action("update_gallery_item", "gallery_item", row.id)
    db.session.commit()
    return jsonify(id=row.id, title=row.title)


@admin_bp.delete("/gallery/<int:item_id>")
@login_required
@permission_required("gallery.delete")
def delete_gallery_item(item_id):
    row = db.get_or_404(GalleryItem, item_id)
    log_action("delete_gallery_item", "gallery_item", row.id, {"title": row.title})
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="Gallery item deleted.")


@admin_bp.get("/resources")
@login_required
@permission_required("resources.view")
def list_resources():
    rows = Resource.query.order_by(Resource.created_at.desc()).all()
    items = []
    for r in rows:
        file_payload = _uploaded_file_payload(r.resource_file)
        file_url = file_payload["url"] if file_payload else None
        items.append({
            "id": r.id,
            "title": r.title,
            "description": r.description,
            "source": r.source,
            "category": r.category,
            "level": r.level,
            "subject": r.subject,
            "curriculum": r.curriculum,
            "publication_year": r.publication_year,
            "tags": r.tags,
            "audience": r.audience,
            "official_source_url": r.official_source_url,
            "featured": r.featured,
            "last_verified_at": r.last_verified_at.isoformat() if r.last_verified_at else None,
            "copyright_status": r.copyright_status,
            "document_type": r.document_type,
            "original_filename": r.original_filename or (file_payload["name"] if file_payload else ""),
            "external_url": r.external_url,
            "url": r.external_url or file_url,
            "file_url": file_url,
            "resource_file_id": r.resource_file_id,
            "resource_file_name": file_payload["name"] if file_payload else "",
            "is_published": r.is_published,
            "status": "published" if r.is_published else "draft",
            "created_at": r.created_at.isoformat(),
            "updated_at": r.updated_at.isoformat(),
        })
    return jsonify(items=items)


def _resource_file_id_from_payload(payload):
    if "resource_file_id" not in payload:
        return None, False
    raw = payload.get("resource_file_id")
    if raw in (None, ""):
        return None, True
    try:
        file_id = int(raw)
    except (TypeError, ValueError) as exc:
        raise ValueError("Resource file is invalid.") from exc
    uploaded = db.session.get(UploadedFile, file_id)
    if not uploaded:
        raise ValueError("Uploaded resource file was not found.")
    if uploaded.category != "resources":
        raise ValueError("Choose a file uploaded as a resource.")
    return uploaded.id, True


@admin_bp.post("/resources")
@login_required
@permission_required("resources.create")
def create_resource():
    payload = request.get_json(silent=True) or {}
    try:
        resource_file_id, has_resource_file = _resource_file_id_from_payload(payload)
        last_verified_at = _dateish(payload.get("last_verified_at"))
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    row = Resource(
        title=(payload.get("title") or "").strip(),
        category=(payload.get("category") or "").strip(),
        description=payload.get("description"),
        source=payload.get("source"),
        level=payload.get("level") or None,
        subject=payload.get("subject") or None,
        curriculum=payload.get("curriculum") or None,
        publication_year=_intish(payload.get("publication_year"), None),
        tags=payload.get("tags") or None,
        audience=payload.get("audience") or None,
        official_source_url=payload.get("official_source_url") or None,
        featured=bool(payload.get("featured")),
        last_verified_at=last_verified_at,
        copyright_status=payload.get("copyright_status") or None,
        document_type=payload.get("document_type") or None,
        original_filename=payload.get("original_filename") or None,
        external_url=payload.get("external_url") or payload.get("url"),
        resource_file_id=resource_file_id if has_resource_file else None,
        is_published=bool(payload.get("is_published", payload.get("status") == "published")),
    )
    if not row.title:
        return jsonify(error="Title is required."), 400
    if not row.category:
        return jsonify(error="Category is required."), 400
    if not row.resource_file_id and not row.external_url:
        return jsonify(error="Upload a document or enter an external URL."), 400
    db.session.add(row)
    db.session.flush()
    log_action("create_resource", "resource", row.id)
    db.session.commit()
    return jsonify(id=row.id, title=row.title), 201


@admin_bp.put("/resources/<int:resource_id>")
@login_required
@permission_required("resources.edit")
def update_resource(resource_id):
    row = db.get_or_404(Resource, resource_id)
    payload = request.get_json(silent=True) or {}
    try:
        resource_file_id, has_resource_file = _resource_file_id_from_payload(payload)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    for field in ["title", "category", "description", "source", "level", "subject", "curriculum", "tags", "audience", "official_source_url", "copyright_status", "document_type", "original_filename", "external_url", "is_published", "featured"]:
        if field in payload:
            value = payload[field]
            if field in {"title", "category"}:
                value = (value or "").strip()
            elif field in {"is_published", "featured"}:
                value = bool(value)
            else:
                value = value or None
            setattr(row, field, value)
    if "publication_year" in payload:
        row.publication_year = _intish(payload.get("publication_year"), None)
    if "last_verified_at" in payload:
        try:
            row.last_verified_at = _dateish(payload.get("last_verified_at"))
        except ValueError as exc:
            return jsonify(error=str(exc)), 400
    if "url" in payload:
        row.external_url = payload["url"]
    if has_resource_file:
        row.resource_file_id = resource_file_id
    if "status" in payload:
        row.is_published = payload["status"] == "published"
    if not (row.title or "").strip() or not (row.category or "").strip():
        return jsonify(error="Title and category are required."), 400
    if not row.resource_file_id and not row.external_url:
        return jsonify(error="Upload a document or enter an external URL."), 400
    log_action("update_resource", "resource", row.id)
    db.session.commit()
    return jsonify(id=row.id, title=row.title)


@admin_bp.delete("/resources/<int:resource_id>")
@login_required
@permission_required("resources.delete")
def delete_resource(resource_id):
    row = db.get_or_404(Resource, resource_id)
    log_action("delete_resource", "resource", row.id, {"title": row.title})
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="Resource deleted.")


@admin_bp.get("/messages")
@login_required
@permission_required("messages.view")
def list_messages():
    status_filter = request.args.get("status")
    source_filter = request.args.get("source")
    q = ContactMessage.query
    if status_filter:
        q = q.filter(ContactMessage.status == status_filter)
    if source_filter:
        q = q.filter(ContactMessage.source == source_filter)
    rows = q.order_by(ContactMessage.created_at.desc()).limit(500).all()
    return jsonify(items=[{
        "id": r.id,
        "ticket_reference": r.ticket_reference or _ticket_reference(r.id),
        "name": r.name,
        "email": r.email,
        "phone": r.phone or "",
        "subject": r.subject,
        "service": r.source,
        "status": r.status,
        "message": r.message,
        "notes": r.notes or "",
        "date": r.created_at.strftime("%d %b %Y"),
        "created_at": r.created_at.isoformat(),
    } for r in rows])


@admin_bp.put("/messages/<int:message_id>")
@login_required
@permission_required("messages.edit")
def update_message(message_id):
    row = db.get_or_404(ContactMessage, message_id)
    payload = request.get_json(silent=True) or {}
    if "status" in payload:
        row.status = payload["status"]
    if "notes" in payload:
        row.notes = (payload["notes"] or "").strip() or None
    log_action("update_message", "contact_message", row.id, {"status": row.status})
    db.session.commit()
    return jsonify(id=row.id, status=row.status, notes=row.notes)


@admin_bp.post("/messages/<int:message_id>/reply")
@login_required
@permission_required("messages.edit")
def reply_to_message(message_id):
    row = db.get_or_404(ContactMessage, message_id)
    payload = request.get_json(silent=True) or {}
    reply_body = (payload.get("message") or "").strip()
    if not reply_body:
        return jsonify(error="Reply message is required."), 400
    ticket_reference = row.ticket_reference or _ticket_reference(row.id)
    is_bookshop = (row.source or "").lower() == "bookshop"
    _shell = bookshop_email_shell if is_bookshop else app_email_shell
    _eyebrow = "RealMindX Bookshop" if is_bookshop else "RealMindX Education"
    result = send_email(
        OutboundEmail(
            to=row.email,
            subject=f"Re: [{ticket_reference}] {row.subject}",
            html=_shell(
                f"Reply: {ticket_reference}",
                (
                    f"<p>Hello {escape(row.name.split()[0] if row.name else 'there')},</p>"
                    f"<p>{escape(reply_body).replace(chr(10), '<br>')}</p>"
                    "<hr style=\"border:none;border-top:1px solid #e5e7eb;margin:22px 0;\" />"
                    f"<p style=\"font-size:13px;color:#5b667a;\"><strong>Your original message:</strong><br>"
                    f"{escape(row.message).replace(chr(10), '<br>')}</p>"
                ),
                eyebrow=_eyebrow,
                preheader=f"A reply to your enquiry ref {ticket_reference}.",
            ),
        ),
        purpose="transactional",
        recipient_user_id=None,
        template_name="contact_message_reply",
    )
    if result.status == "mocked":
        log_action(
            "reply_contact_message_mocked",
            "contact_message",
            row.id,
            {"ticket_reference": ticket_reference},
        )
        db.session.commit()
        return jsonify(
            id=row.id,
            status=row.status,
            ticket_reference=ticket_reference,
            delivery_status="mocked",
            message="Reply recorded in mock mode; no email was sent.",
        ), 202
    if result.status not in ("queued", "accepted", "sent", "delivered"):
        log_action(
            "reply_contact_message_failed",
            "contact_message",
            row.id,
            {
                "ticket_reference": ticket_reference,
                "delivery_status": result.status,
                "error_code": result.error_code,
            },
        )
        db.session.commit()
        return jsonify(
            error="The reply could not be delivered. The message was not marked replied.",
            code="reply_delivery_failed",
            delivery_status=result.status,
        ), 503
    row.status = "replied"
    log_action("reply_contact_message", "contact_message", row.id, {"ticket_reference": ticket_reference})
    db.session.commit()
    return jsonify(id=row.id, status=row.status, ticket_reference=ticket_reference)


@admin_bp.delete("/messages/<int:message_id>")
@login_required
@permission_required("messages.delete")
def delete_message(message_id):
    row = db.get_or_404(ContactMessage, message_id)
    log_action("delete_message", "contact_message", row.id, {"subject": row.subject})
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="Message deleted.")


def _contact_summary():
    def source_count(source):
        return (
            db.session.query(func.count(func.distinct(Contact.id)))
            .join(ContactSource, ContactSource.contact_id == Contact.id)
            .filter(ContactSource.source == source)
            .scalar()
            or 0
        )

    return {
        "total_contacts": Contact.query.count(),
        "teachers": source_count("teacher"),
        "bookshop": source_count("bookshop"),
        "newsletter": source_count("newsletter"),
        "schools": source_count("school"),
    }


@admin_bp.get("/contacts")
@login_required
@permission_required("contacts.view")
def list_contacts():
    query = Contact.query
    q = (request.args.get("q") or "").strip()
    source = (request.args.get("source") or "").strip().lower().replace(" ", "_")
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Contact.full_name.ilike(like),
            Contact.email.ilike(like),
            Contact.phone.ilike(like),
        ))
    if source:
        query = query.filter(Contact.sources.any(ContactSource.source == source))
    page = max(request.args.get("page", 1, type=int), 1)
    page_size = min(max(request.args.get("page_size", 25, type=int), 1), 100)
    pagination = query.order_by(Contact.last_activity_at.desc(), Contact.id.desc()).paginate(
        page=page,
        per_page=page_size,
        error_out=False,
    )
    return jsonify(
        items=[contact_json(row) for row in pagination.items],
        summary=_contact_summary(),
        pagination={
            "page": pagination.page,
            "page_size": pagination.per_page,
            "pages": pagination.pages,
            "total": pagination.total,
            "has_next": pagination.has_next,
            "has_prev": pagination.has_prev,
        },
    )


def _communication_attempt_json(row):
    return {
        "id": row.id,
        "subject": row.subject or row.template_name or "Email",
        "purpose": row.purpose,
        "provider": row.provider,
        "provider_message_id": row.provider_message_id,
        "status": row.status,
        "error_code": row.error_code,
        "error_message": row.error_message,
        "initiated_by": row.initiated_by,
        "requested_at": row.requested_at.isoformat() if row.requested_at else None,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


@admin_bp.get("/contacts/<int:contact_id>")
@login_required
@permission_required("contacts.view")
def get_contact(contact_id):
    row = db.get_or_404(Contact, contact_id)
    payload = contact_json(row)
    payload["emails"] = [
        _communication_attempt_json(attempt)
        for attempt in CommunicationAttempt.query.filter_by(contact_id=row.id, channel="email")
        .order_by(CommunicationAttempt.requested_at.desc())
        .limit(50)
        .all()
    ]
    payload["linked_records"] = {
        "teachers": [
            {"id": user.id, "application_id": user.application_id, "created_at": user.created_at.isoformat() if user.created_at else None}
            for user in User.query.filter(func.lower(User.email) == row.email).limit(5).all()
            if user.teacher_service_enabled
        ],
        "orders": [
            {
                "id": order.id,
                "order_reference": order.order_reference,
                "status": order.status,
                "created_at": order.created_at.isoformat() if order.created_at else None,
            }
            for order in Order.query.filter(func.lower(Order.email) == row.email)
            .order_by(Order.created_at.desc())
            .limit(20)
            .all()
        ],
    }
    return jsonify(item=payload)


@admin_bp.post("/contacts")
@login_required
@permission_required("contacts.create")
def create_contact():
    payload = request.get_json(silent=True) or {}
    try:
        row = upsert_contact(
            payload.get("email"),
            full_name=(payload.get("full_name") or "").strip() or None,
            phone=(payload.get("phone") or "").strip() or None,
            source="admin_added",
            metadata={"created_by": current_user.id},
        )
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    log_action("create_contact", "contact", row.id, {"email": row.email})
    db.session.commit()
    return jsonify(item=contact_json(row)), 201


@admin_bp.put("/contacts/<int:contact_id>")
@login_required
@permission_required("contacts.edit")
def update_contact(contact_id):
    row = db.get_or_404(Contact, contact_id)
    payload = request.get_json(silent=True) or {}
    if "full_name" in payload:
        row.full_name = (payload.get("full_name") or "").strip() or None
    if "phone" in payload:
        row.phone = (payload.get("phone") or "").strip() or None
    log_action("update_contact", "contact", row.id)
    db.session.commit()
    return jsonify(item=contact_json(row))


@admin_bp.post("/contacts/<int:contact_id>/email")
@login_required
@permission_required("contacts.email")
def send_contact_email(contact_id):
    row = db.get_or_404(Contact, contact_id)
    payload = request.get_json(silent=True) or {}
    subject = (payload.get("subject") or "").strip()
    message = (payload.get("message") or "").strip()
    idempotency_key = (payload.get("idempotency_key") or "").strip()
    if not subject or not message:
        return jsonify(error="Subject and message are required."), 400
    if len(idempotency_key) < 8 or len(idempotency_key) > 80:
        return jsonify(error="A valid send request identifier is required."), 400

    existing = CommunicationAttempt.query.filter_by(idempotency_key=idempotency_key).first()
    if existing:
        if existing.contact_id != row.id:
            return jsonify(error="That send request identifier is already in use."), 409
        return jsonify(message="This email request was already processed.", attempt=_communication_attempt_json(existing))

    now = datetime.now(timezone.utc)
    attempt = CommunicationAttempt(
        contact_id=row.id,
        channel="email",
        purpose="transactional",
        recipient_user_id=None,
        masked_destination=mask_destination("email", row.email),
        template_name="admin_contact_email",
        provider="pending",
        mode=resolve_communication_mode(),
        status="queued",
        initiated_by=current_user.id,
        idempotency_key=idempotency_key,
        subject=subject,
        requested_at=now,
    )
    db.session.add(attempt)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        existing = CommunicationAttempt.query.filter_by(idempotency_key=idempotency_key).first()
        if existing:
            return jsonify(message="This email request was already processed.", attempt=_communication_attempt_json(existing))
        raise

    result = send_email(
        OutboundEmail(
            to=row.email,
            subject=subject,
            html=app_email_shell(
                subject,
                f"<p>{escape(message).replace(chr(10), '<br>')}</p>",
                eyebrow="RealMindX",
                preheader=subject,
            ),
            text=message,
        ),
        purpose="transactional",
        recipient_user_id=None,
        template_name="admin_contact_email",
        contact_id=row.id,
        initiated_by=current_user.id,
        attempt_id=attempt.id,
    )
    log_action("send_contact_email", "contact", row.id, {"attempt_id": attempt.id, "status": result.status})
    db.session.commit()
    return jsonify(
        message="Email accepted for delivery." if result.status in ("queued", "accepted", "sent", "delivered", "mocked") else "Email delivery failed.",
        attempt=_communication_attempt_json(attempt),
    ), 200 if result.status in ("queued", "accepted", "sent", "delivered", "mocked") else 502


@admin_bp.get("/newsletters")
@login_required
@permission_required("newsletters.view")
def list_newsletters():
    query = NewsletterSubscriber.query
    q = (request.args.get("q") or "").strip().lower()
    source = (request.args.get("source") or "").strip()
    status = (request.args.get("status") or "").strip()
    tag = (request.args.get("tag") or "").strip()
    if q:
        query = query.filter(NewsletterSubscriber.email.ilike(f"%{q}%"))
    if source:
        query = query.filter(or_(
            NewsletterSubscriber.source == source,
            NewsletterSubscriber.sources.contains([source]),
        ))
    if status:
        if status == "active":
            query = query.filter(NewsletterSubscriber.is_active.is_(True))
        elif status == "unsubscribed":
            query = query.filter(or_(
                NewsletterSubscriber.is_active.is_(False),
                NewsletterSubscriber.communication_status == UNSUBSCRIBED,
            ))
        else:
            query = query.filter(NewsletterSubscriber.communication_status == status)
    if tag:
        query = query.filter(NewsletterSubscriber.tags.contains([tag]))
    rows = query.order_by(NewsletterSubscriber.created_at.desc()).limit(500).all()
    return jsonify(items=[newsletter_subscriber_json(r) for r in rows])


@admin_bp.get("/newsletters/audience")
@login_required
@permission_required("newsletters.view")
def list_newsletter_audience():
    query = Contact.query
    q = (request.args.get("q") or "").strip()
    source = (request.args.get("source") or "").strip().lower()
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Contact.full_name.ilike(like), Contact.email.ilike(like), Contact.phone.ilike(like)))
    if source:
        query = query.filter(Contact.sources.any(ContactSource.source == source))
    rows = query.order_by(Contact.last_activity_at.desc(), Contact.id.desc()).limit(500).all()
    items = []
    for row in rows:
        item = contact_json(row)
        subscriber = row.newsletter_subscription
        item["newsletter_status"] = (
            UNSUBSCRIBED
            if subscriber and (not subscriber.is_active or subscriber.communication_status == UNSUBSCRIBED)
            else subscriber.communication_status if subscriber else None
        )
        items.append(item)
    return jsonify(items=items)


@admin_bp.post("/newsletters")
@login_required
@permission_required("newsletters.create")
def create_newsletter_contact():
    payload = request.get_json(silent=True) or {}
    try:
        email = normalize_contact_email(payload.get("email"))
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    source = (payload.get("source") or "admin_added").strip()
    status = payload.get("communication_status") or MARKETING_ACTIVE
    tags = payload.get("tags") or []
    if isinstance(tags, str):
        tags = [item.strip() for item in re.split(r"[,;]+", tags) if item.strip()]
    contact = upsert_contact(
        email,
        full_name=(payload.get("full_name") or "").strip() or None,
        phone=(payload.get("phone") or "").strip() or None,
        source="admin_added",
        metadata={"admin_source": source, "tags": tags},
    )
    row = NewsletterSubscriber.query.filter_by(email=email).first()
    if not row:
        row = NewsletterSubscriber(
            email=email,
            source=source,
            sources=[source],
            tags=tags,
            communication_status=status,
            is_active=status != UNSUBSCRIBED,
            confirmed_at=datetime.now(timezone.utc),
            unsubscribe_token=secrets.token_urlsafe(32),
            notes=(payload.get("notes") or "").strip() or None,
            contact=contact,
        )
        db.session.add(row)
    else:
        row.contact = contact
        row.tags = tags
        row.communication_status = status
        row.is_active = status != UNSUBSCRIBED
    db.session.commit()
    log_action("create_newsletter_contact", "newsletter_subscriber", row.id, {"email": row.email, "source": source})
    return jsonify(item=newsletter_subscriber_json(row)), 201


NEWSLETTER_LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)\s]+)\)")
NEWSLETTER_IMAGE_RE = re.compile(r"^!\[(?:(left|right|full):)?([^\]]*)\]\(([^)\s]+)\)$", re.IGNORECASE)


def _absolute_newsletter_url(raw_url, base_url=None):
    value = (raw_url or "").strip()
    if not value:
        return ""
    if value.startswith(("http://", "https://", "mailto:", "tel:")):
        return value
    origin = (base_url or current_app.config.get("BASE_URL", "https://realmindxgh.com")).rstrip("/")
    return f"{origin}/{value.lstrip('/')}"


def _render_newsletter_inline(text):
    rendered = []
    position = 0
    for match in NEWSLETTER_LINK_RE.finditer(text or ""):
        rendered.append(escape((text or "")[position:match.start()]).replace("\n", "<br>"))
        label = escape(match.group(1).strip())
        href = escape(_absolute_newsletter_url(match.group(2)), quote=True)
        rendered.append(
            f'<a class="newsletter-rich-link" href="{href}" '
            'style="color:#143670!important;font-weight:800;text-decoration:underline;text-underline-offset:3px;">'
            f"{label}</a>"
        )
        position = match.end()
    rendered.append(escape((text or "")[position:]).replace("\n", "<br>"))
    return "".join(rendered)


def _render_newsletter_image(match):
    align = (match.group(1) or "full").lower()
    alt = escape((match.group(2) or "Newsletter image").strip(), quote=True)
    src = escape(_absolute_newsletter_url(match.group(3)), quote=True)
    if align in {"left", "right"}:
        margin = "4px 18px 12px 0" if align == "left" else "4px 0 12px 18px"
        return (
            f'<img class="newsletter-rich-image newsletter-rich-image-{align}" src="{src}" alt="{alt}" '
            f'align="{align}" width="240" '
            f'style="display:block;max-width:46%;width:240px;height:auto;border-radius:12px;'
            f'border:1px solid #dce5f0;float:{align};margin:{margin};" />'
        )
    return (
        f'<img class="newsletter-rich-image" src="{src}" alt="{alt}" '
        'style="display:block;width:100%;max-width:100%;height:auto;border-radius:14px;'
        'border:1px solid #dce5f0;margin:18px 0;" />'
    )


def _render_newsletter_body(body):
    if contains_rich_html(body):
        return (
            '<div class="newsletter-rich" style="background:#ffffff;color:#1a2a40;">'
            + sanitize_rich_html(body)
            + '</div>'
        )
    blocks = []
    for block in re.split(r"\n\s*\n", body or ""):
        clean = block.strip()
        if not clean:
            continue
        image_match = NEWSLETTER_IMAGE_RE.match(clean)
        if image_match:
            blocks.append(_render_newsletter_image(image_match))
        else:
            blocks.append(f'<p style="margin:0 0 18px;">{_render_newsletter_inline(clean)}</p>')
    return (
        '<style>'
        '@media (prefers-color-scheme:dark){.newsletter-rich-link{color:#ffcc01!important;}}'
        '@media only screen and (max-width:600px){.newsletter-rich-image-left,.newsletter-rich-image-right{'
        'float:none!important;margin:16px 0!important;max-width:100%!important;width:100%!important;}}'
        '</style>'
        '<div class="newsletter-rich" style="background:#ffffff;color:#1a2a40;">'
        + "".join(blocks)
        + '<div style="clear:both;height:1px;line-height:1px;">&nbsp;</div></div>'
    )


def _render_newsletter_sections(sections, *, asset_base_url=None):
    if not isinstance(sections, list) or not sections:
        return ""
    blocks = []
    for index, section in enumerate(sections):
        heading = escape((section.get("heading") or "").strip())
        body = (section.get("body") or "").strip()
        caption = escape((section.get("caption") or "").strip())
        image_file_id = section.get("image_file_id")
        image_url = ""
        if image_file_id:
            image_file = db.session.get(UploadedFile, image_file_id)
            image_url = _upload_public_url(image_file) if image_file else ""
        position = (section.get("image_position") or "auto").strip().lower()
        if position == "full":
            position = "top"
        if position == "auto":
            position = "right" if index % 2 == 0 else "left"
        if position not in {"left", "right", "top", "bottom"}:
            position = "right"
        image_size = (section.get("image_size") or "medium").strip().lower()
        if image_size not in {"small", "medium", "large"}:
            image_size = "medium"

        text_html = ""
        if heading:
            text_html += f'<h2 style="margin:0 0 10px;color:#143670;font-size:20px;line-height:1.25;">{heading}</h2>'
        if body:
            text_html += _render_newsletter_body(body)

        image_html = ""
        if image_url:
            safe_url = escape(_absolute_newsletter_url(image_url, asset_base_url), quote=True)
            image_width = {"small": 180, "medium": 260, "large": 340}[image_size]
            if position in {"top", "bottom"}:
                image_width = 576
                image_style = "display:block;width:100%;max-width:100%;height:auto;border-radius:12px;border:1px solid #dce5f0;"
            else:
                image_style = (
                    f"display:block;width:100%;max-width:{image_width}px;height:auto;border-radius:12px;"
                    "border:1px solid #dce5f0;"
                )
            image_html = (
                f'<img src="{safe_url}" alt="{caption or heading or "Campaign image"}" width="{image_width}" '
                f'style="{image_style}" />'
            )
            if caption:
                image_html += f'<p style="margin:8px 0 0;color:#53657d;font-size:12px;line-height:1.4;">{caption}</p>'

        if image_html and position in {"left", "right"}:
            image_cell = f'<td class="newsletter-section-image" width="42%" style="width:42%;vertical-align:top;padding:0 0 16px;">{image_html}</td>'
            text_cell = f'<td class="newsletter-section-text" style="vertical-align:top;padding:0 0 16px;">{text_html}</td>'
            cells = image_cell + '<td width="18" style="width:18px;">&nbsp;</td>' + text_cell if position == "left" else text_cell + '<td width="18" style="width:18px;">&nbsp;</td>' + image_cell
            blocks.append(f'<table class="newsletter-section-row" role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin:0 0 22px;"><tr>{cells}</tr></table>')
        else:
            full_image = f'<div style="margin:0 0 14px;">{image_html}</div>' if image_html else ""
            blocks.append(
                '<div style="margin:0 0 24px;">'
                + (text_html + full_image if position == "bottom" else full_image + text_html)
                + '</div>'
            )
    return (
        '<style>@media only screen and (max-width:600px){'
        '.newsletter-section-row,.newsletter-section-row tbody,.newsletter-section-row tr,.newsletter-section-row td{display:block!important;width:100%!important;}'
        '.newsletter-section-image{padding:0 0 14px!important}.newsletter-section-text{padding:0!important}'
        '}</style>'
        + "".join(blocks)
    )


def _campaign_from_email(sender):
    sender = (sender or "news").strip().lower()
    if sender == "sales":
        return current_app.config.get("SALES_FROM_EMAIL") or "RealMindX Sales <sales@send.realmindxgh.com>"
    if sender == "bookshop":
        return current_app.config.get("BOOKSHOP_FROM_EMAIL")
    if sender == "default":
        return current_app.config.get("DEFAULT_FROM_EMAIL")
    return current_app.config.get("NEWSLETTER_FROM_EMAIL")


def _newsletter_footer_note(contact, subscriber):
    notes = []
    if any(link.source == "teacher" for link in contact.sources):
        notes.append(
            "You are receiving this email because you signed up for teacher recruitment "
            "with RealMindX Education."
        )
    if subscriber:
        if not subscriber.unsubscribe_token:
            subscriber.unsubscribe_token = secrets.token_urlsafe(32)
        base_url = current_app.config.get("SITE_BASE_URL", "https://realmindxgh.com").rstrip("/")
        unsubscribe_url = escape(
            f"{base_url}/unsubscribe?token={subscriber.unsubscribe_token}",
            quote=True,
        )
        notes.append(
            f'<a href="{unsubscribe_url}" style="color:#aaa;">Unsubscribe from newsletters</a>.'
        )
    return " ".join(notes) or None


def _newsletter_campaign_json(row, *, include_content=False):
    payload = {
        "id": row.id,
        "subject": row.subject,
        "title": row.title,
        "brand": row.brand,
        "sender": row.sender,
        "recipient_count": row.recipient_count,
        "sent_count": row.sent_count,
        "mocked_count": row.mocked_count,
        "failed_count": row.failed_count,
        "status": row.status,
        "sent_at": row.sent_at.isoformat() if row.sent_at else None,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }
    if include_content:
        payload["content"] = row.content or {}
        payload["audience"] = row.audience or {}
    return payload


NEWSLETTER_SUCCESS_STATUSES = {"mocked", "queued", "accepted", "sent", "delivered"}
NEWSLETTER_FAILED_STATUSES = {"disabled", "failed", "rejected", "expired", "skipped"}


def _newsletter_recipient_json(row):
    attempts = row.attempts or []
    return {
        "id": row.id,
        "campaign_id": row.campaign_id,
        "contact_id": row.contact_id,
        "email": row.email,
        "status": row.status,
        "successful": row.status in NEWSLETTER_SUCCESS_STATUSES,
        "provider": row.provider,
        "error_code": row.error_code,
        "error_message": row.error_message,
        "attempt_count": row.attempt_count,
        "last_attempt_at": row.last_attempt_at.isoformat() if row.last_attempt_at else None,
        "attempts": attempts,
    }


def _newsletter_attempt_json(result, attempted_at):
    return {
        "status": result.status,
        "provider": result.provider,
        "provider_message_id": result.provider_message_id,
        "error_code": result.error_code,
        "error_message": result.error_message,
        "attempted_at": attempted_at.isoformat(),
    }


def _apply_newsletter_result(recipient, result, attempted_at):
    recipient.status = result.status
    recipient.provider = result.provider
    recipient.provider_message_id = result.provider_message_id
    recipient.error_code = result.error_code
    recipient.error_message = (result.error_message or "")[:500] or None
    recipient.attempt_count = int(recipient.attempt_count or 0) + 1
    recipient.attempts = list(recipient.attempts or []) + [_newsletter_attempt_json(result, attempted_at)]
    recipient.last_attempt_at = attempted_at


def _refresh_newsletter_campaign_counts(campaign):
    rows = list(campaign.recipients)
    campaign.recipient_count = len(rows)
    campaign.mocked_count = sum(row.status == "mocked" for row in rows)
    campaign.sent_count = sum(row.status in (NEWSLETTER_SUCCESS_STATUSES - {"mocked"}) for row in rows)
    campaign.failed_count = sum(row.status in NEWSLETTER_FAILED_STATUSES for row in rows)
    successful = campaign.sent_count + campaign.mocked_count
    campaign.status = (
        "failed" if campaign.failed_count and not successful
        else "partial" if campaign.failed_count
        else "completed"
    )


def _send_saved_newsletter_recipient(campaign, recipient):
    contact = db.session.get(Contact, recipient.contact_id) if recipient.contact_id else None
    if contact is None:
        contact = Contact.query.filter(func.lower(Contact.email) == recipient.email.lower()).first()
    if contact is None:
        raise ValueError("This recipient is no longer in the contacts directory.")
    subscriber = contact.newsletter_subscription
    if subscriber and (subscriber.communication_status == UNSUBSCRIBED or not subscriber.is_active):
        raise ValueError("This recipient has unsubscribed and cannot be resent this newsletter.")

    content = campaign.content or {}
    sections = content.get("sections") or []
    body = (content.get("body") or "").strip()
    body_html = _render_newsletter_sections(sections) if sections else _render_newsletter_body(body)
    brand = (campaign.brand or content.get("brand") or "realmindx").strip().lower()
    sender = (campaign.sender or content.get("sender") or "news").strip().lower()
    shell = bookshop_email_shell if brand == "bookshop" else app_email_shell
    hero_image_url = None
    if content.get("image_file_id"):
        image_file = db.session.get(UploadedFile, content["image_file_id"])
        hero_image_url = _upload_public_url(image_file) if image_file else None
    footer_note = _newsletter_footer_note(contact, subscriber)

    return send_email(
        OutboundEmail(
            to=contact.email,
            subject=campaign.subject,
            from_email=_campaign_from_email(sender),
            html=shell(
                campaign.title,
                body_html,
                content.get("cta_label") or None,
                content.get("cta_url") or None,
                eyebrow="",
                preheader=content.get("preheader") or campaign.title,
                hero_image_url=hero_image_url,
                footer_note=footer_note,
            ),
        ),
        purpose="transactional",
        recipient_user_id=None,
        template_name="newsletter_campaign_retry",
        contact_id=contact.id,
        initiated_by=current_user.id,
        batch_id=f"newsletter-{campaign.id}",
        idempotency_key=f"newsletter-{campaign.id}-recipient-{recipient.id}-{uuid4().hex}",
    )


@admin_bp.get("/newsletters/campaigns")
@login_required
@permission_required("newsletters.view")
def list_newsletter_campaigns():
    rows = NewsletterCampaign.query.order_by(NewsletterCampaign.sent_at.desc(), NewsletterCampaign.id.desc()).limit(100).all()
    return jsonify(items=[_newsletter_campaign_json(row, include_content=True) for row in rows])


@admin_bp.get("/newsletters/campaigns/<int:campaign_id>/recipients")
@login_required
@permission_required("newsletters.view")
def list_newsletter_campaign_recipients(campaign_id):
    campaign = db.get_or_404(NewsletterCampaign, campaign_id)
    rows = (
        NewsletterCampaignRecipient.query
        .filter_by(campaign_id=campaign.id)
        .order_by(NewsletterCampaignRecipient.status.asc(), NewsletterCampaignRecipient.email.asc())
        .all()
    )
    return jsonify(
        campaign=_newsletter_campaign_json(campaign),
        recipients=[_newsletter_recipient_json(row) for row in rows],
        details_available=bool(rows),
    )


def _retry_newsletter_recipients(campaign, recipients):
    results = []
    for recipient in recipients:
        attempted_at = datetime.now(timezone.utc)
        try:
            result = _send_saved_newsletter_recipient(campaign, recipient)
            _apply_newsletter_result(recipient, result, attempted_at)
            results.append(_newsletter_recipient_json(recipient))
        except ValueError as exc:
            results.append({**_newsletter_recipient_json(recipient), "retry_error": str(exc)})
    _refresh_newsletter_campaign_counts(campaign)
    return results


@admin_bp.post("/newsletters/campaigns/<int:campaign_id>/recipients/<int:recipient_id>/resend")
@login_required
@permission_required("newsletters.create")
def resend_newsletter_recipient(campaign_id, recipient_id):
    campaign = db.get_or_404(NewsletterCampaign, campaign_id)
    recipient = NewsletterCampaignRecipient.query.filter_by(
        id=recipient_id,
        campaign_id=campaign.id,
    ).first_or_404()
    if recipient.status not in NEWSLETTER_FAILED_STATUSES:
        return jsonify(error="Only failed newsletter deliveries can be resent."), 409
    results = _retry_newsletter_recipients(campaign, [recipient])
    log_action(
        "resend_newsletter_recipient",
        "newsletter",
        campaign.id,
        {"recipient_id": recipient.id, "status": recipient.status},
    )
    db.session.commit()
    retry_error = results[0].get("retry_error")
    if retry_error:
        return jsonify(error=retry_error, recipient=results[0], campaign=_newsletter_campaign_json(campaign)), 409
    return jsonify(recipient=results[0], campaign=_newsletter_campaign_json(campaign))


@admin_bp.post("/newsletters/campaigns/<int:campaign_id>/recipients/resend-failed")
@login_required
@permission_required("newsletters.create")
def resend_failed_newsletter_recipients(campaign_id):
    campaign = db.get_or_404(NewsletterCampaign, campaign_id)
    failed_rows = NewsletterCampaignRecipient.query.filter(
        NewsletterCampaignRecipient.campaign_id == campaign.id,
        NewsletterCampaignRecipient.status.in_(NEWSLETTER_FAILED_STATUSES),
    ).all()
    if not failed_rows:
        return jsonify(error="This campaign has no failed recipients to resend."), 409
    results = _retry_newsletter_recipients(campaign, failed_rows)
    log_action(
        "resend_failed_newsletter_recipients",
        "newsletter",
        campaign.id,
        {
            "requested": len(failed_rows),
            "successful": sum(row.get("status") in NEWSLETTER_SUCCESS_STATUSES for row in results),
            "failed": sum(row.get("status") in NEWSLETTER_FAILED_STATUSES or bool(row.get("retry_error")) for row in results),
        },
    )
    db.session.commit()
    return jsonify(results=results, campaign=_newsletter_campaign_json(campaign))


@admin_bp.delete("/newsletters/campaigns/<int:campaign_id>")
@login_required
@permission_required("newsletters.delete")
def delete_newsletter_campaign(campaign_id):
    campaign = db.get_or_404(NewsletterCampaign, campaign_id)
    audit_details = {
        "subject": campaign.subject,
        "sender": campaign.sender,
        "sent_at": campaign.sent_at.isoformat() if campaign.sent_at else None,
    }
    db.session.delete(campaign)
    log_action("delete_newsletter_campaign", "newsletter", campaign_id, audit_details)
    db.session.commit()
    return jsonify(success=True, deleted_id=campaign_id)


@admin_bp.post("/newsletters/preview")
@login_required
@permission_required("newsletters.create")
def preview_newsletter_campaign():
    payload = request.get_json(silent=True) or {}
    subject = (payload.get("subject") or "Newsletter preview").strip()
    title = (payload.get("title") or subject).strip()
    body = (payload.get("body") or "").strip()
    sections = payload.get("sections") or []
    preview_asset_origin = request.host_url.rstrip("/")
    body_html = _render_newsletter_sections(sections, asset_base_url=preview_asset_origin) if sections else _render_newsletter_body(body)
    if not body_html.strip():
        body_html = '<p style="margin:0;color:#53657d;">Your newsletter content will appear here.</p>'

    image_url = None
    image_file_id = payload.get("image_file_id")
    if image_file_id:
        image_file = db.session.get(UploadedFile, image_file_id)
        image_url = _upload_public_url(image_file) if image_file else None

    brand = (payload.get("brand") or "realmindx").strip().lower()
    is_bookshop = brand == "bookshop"
    shell = bookshop_email_shell if is_bookshop else app_email_shell
    sender = (payload.get("sender") or ("bookshop" if is_bookshop else "news")).strip().lower()
    html = shell(
        title,
        body_html,
        payload.get("cta_label") or None,
        payload.get("cta_url") or None,
        eyebrow="",
        preheader=payload.get("preheader") or payload.get("summary") or title,
        hero_image_url=image_url,
        footer_note=None,
    )
    return jsonify(html=html, subject=subject, brand=brand, sender=sender)


@admin_bp.post("/newsletters/send")
@login_required
@permission_required("newsletters.create")
def send_newsletter_campaign():
    payload = request.get_json(silent=True) or {}
    subject = (payload.get("subject") or "").strip()
    title = (payload.get("title") or subject).strip()
    body = (payload.get("body") or "").strip()
    sections = payload.get("sections") or []
    if not subject or not title or (not body and not sections):
        return jsonify(error="Subject, title, and message content are required."), 400

    body_html = _render_newsletter_sections(sections) if sections else _render_newsletter_body(body)
    image_url = None
    image_file_id = payload.get("image_file_id")
    if image_file_id:
        image_file = db.session.get(UploadedFile, image_file_id)
        image_url = _upload_public_url(image_file) if image_file else None

    brand = (payload.get("brand") or "realmindx").strip().lower()
    is_bookshop = brand == "bookshop"
    shell = bookshop_email_shell if is_bookshop else app_email_shell
    sender = (payload.get("sender") or payload.get("purpose") or ("bookshop" if is_bookshop else "news")).strip().lower()
    from_email = _campaign_from_email(sender)
    contact_ids = payload.get("contact_ids") or []
    recipient_ids = payload.get("recipient_ids") or []  # legacy subscriber IDs
    recipient_emails = payload.get("recipient_emails") or payload.get("recipients") or []
    if isinstance(recipient_emails, str):
        recipient_emails = [item.strip() for item in re.split(r"[\s,;]+", recipient_emails) if item.strip()]
    recipients = []
    seen = set()
    if contact_ids:
        recipients.extend((contact, contact.newsletter_subscription) for contact in Contact.query.filter(Contact.id.in_(contact_ids)).all())
    if recipient_ids:
        recipients.extend((subscriber.contact, subscriber) for subscriber in NewsletterSubscriber.query.filter(NewsletterSubscriber.id.in_(recipient_ids)).all() if subscriber.contact)
    for raw_email in recipient_emails:
        try:
            email = normalize_contact_email(raw_email)
        except ValueError:
            continue
        contact = Contact.query.filter_by(email=email).first()
        if contact:
            recipients.append((contact, contact.newsletter_subscription))
    if not recipients:
        return jsonify(error="Select at least one gathered contact before sending."), 400

    sent = 0
    mocked = 0
    failed = 0
    delivery_results = []
    for contact, subscriber in recipients:
        if contact.email in seen:
            continue
        if subscriber and (subscriber.communication_status == UNSUBSCRIBED or not subscriber.is_active):
            continue
        seen.add(contact.email)
        footer_note = _newsletter_footer_note(contact, subscriber)
        attempted_at = datetime.now(timezone.utc)
        result = send_email(
            OutboundEmail(
                to=contact.email,
                subject=subject,
                from_email=from_email,
                html=shell(
                    title,
                    body_html,
                    payload.get("cta_label") or None,
                    payload.get("cta_url") or None,
                    eyebrow="",
                    preheader=payload.get("preheader") or payload.get("summary") or title,
                    hero_image_url=image_url,
                    footer_note=footer_note,
                ),
            ),
            purpose="transactional",
            recipient_user_id=None,
            template_name="newsletter_campaign",
            contact_id=contact.id,
            initiated_by=current_user.id,
        )
        delivery_results.append((contact, result, attempted_at))
        if result.status == "mocked":
            mocked += 1
        elif result.status in ("queued", "accepted", "sent", "delivered"):
            sent += 1
        else:
            failed += 1

    campaign = NewsletterCampaign(
        subject=subject,
        title=title,
        brand=brand,
        sender=sender,
        content={
            "subject": subject,
            "title": title,
            "brand": brand,
            "sender": sender,
            "preheader": payload.get("preheader") or "",
            "body": body,
            "sections": sections,
            "cta_label": payload.get("cta_label") or "",
            "cta_url": payload.get("cta_url") or "",
            "image_file_id": image_file_id,
        },
        audience={"contact_ids": contact_ids, "recipient_emails": recipient_emails},
        recipient_count=len(seen),
        sent_count=sent,
        mocked_count=mocked,
        failed_count=failed,
        status="failed" if failed and not (sent or mocked) else "partial" if failed else "completed",
        initiated_by=current_user.id,
        sent_at=datetime.now(timezone.utc),
    )
    db.session.add(campaign)
    db.session.flush()
    for contact, result, attempted_at in delivery_results:
        recipient = NewsletterCampaignRecipient(
            campaign_id=campaign.id,
            contact_id=contact.id,
            email=contact.email,
            status="pending",
            attempt_count=0,
            attempts=[],
        )
        _apply_newsletter_result(recipient, result, attempted_at)
        db.session.add(recipient)
    log_action(
        "send_newsletter_campaign",
        "newsletter",
        campaign.id,
        {
            "subject": subject,
            "brand": brand,
            "sender": sender,
            "sent": sent,
            "mocked": mocked,
            "failed": failed,
        },
    )
    db.session.commit()
    return jsonify(
        message=f"Newsletter accepted for {sent} subscriber(s).",
        sent=sent,
        mocked=mocked,
        failed=failed,
        campaign=_newsletter_campaign_json(campaign, include_content=True),
    )


@admin_bp.put("/newsletters/<int:subscriber_id>")
@login_required
@permission_required("newsletters.edit")
def update_newsletter_subscriber(subscriber_id):
    row = db.get_or_404(NewsletterSubscriber, subscriber_id)
    payload = request.get_json(silent=True) or {}
    requested_status = payload.get("communication_status") or payload.get("status")
    if "is_active" in payload:
        row.is_active = bool(payload["is_active"])
        if not row.is_active:
            row.communication_status = UNSUBSCRIBED
    if "status" in payload:
        row.communication_status = payload["status"]
        row.is_active = payload["status"] != UNSUBSCRIBED
    if "communication_status" in payload:
        row.communication_status = payload["communication_status"]
        row.is_active = payload["communication_status"] != UNSUBSCRIBED
    if "source" in payload:
        row.source = (payload.get("source") or row.source or "site").strip()
    if "sources" in payload and isinstance(payload["sources"], list):
        row.sources = [str(item).strip() for item in payload["sources"] if str(item).strip()]
    if "tags" in payload:
        tags = payload["tags"]
        if isinstance(tags, str):
            tags = [item.strip() for item in re.split(r"[,;]+", tags) if item.strip()]
        row.tags = tags if isinstance(tags, list) else []
    if "notes" in payload:
        row.notes = (payload.get("notes") or "").strip() or None
    log_action("update_newsletter_subscriber", "newsletter_subscriber", row.id)
    db.session.commit()
    return jsonify(item=newsletter_subscriber_json(row))


@admin_bp.delete("/newsletters/<int:subscriber_id>")
@login_required
@permission_required("newsletters.delete")
def delete_newsletter_subscriber(subscriber_id):
    row = db.get_or_404(NewsletterSubscriber, subscriber_id)
    contact = row.contact
    log_action("delete_newsletter_subscriber", "newsletter_subscriber", row.id, {"email": row.email})
    db.session.delete(row)
    db.session.flush()
    if contact:
        remove_contact_source(contact, "newsletter")
    db.session.commit()
    return jsonify(message="Subscriber deleted.")


@admin_bp.get("/services")
@login_required
@permission_required("services.view")
def list_services_content():
    rows = _admin_collection_items("services", DEFAULT_SERVICES)
    rows = _enrich_service_media(rows)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("sort_order") or 0, item.get("label") or "")))


@admin_bp.post("/services")
@login_required
@permission_required("services.create")
def create_service_content():
    payload = request.get_json(silent=True) or {}
    if not (payload.get("label") or payload.get("title")):
        return jsonify(error="Service label or title is required."), 400
    row, error = _create_admin_collection_item("services", DEFAULT_SERVICES, payload, "service")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0]), 201


@admin_bp.put("/services/<string:service_id>")
@login_required
@permission_required("services.edit")
def update_service_content(service_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("services", DEFAULT_SERVICES, service_id, payload, "service")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0])


@admin_bp.delete("/services/<string:service_id>")
@login_required
@permission_required("services.delete")
def delete_service_content(service_id):
    return _delete_admin_collection_item("services", DEFAULT_SERVICES, service_id, "service")


@admin_bp.get("/partners")
@login_required
@permission_required("partners.view")
def list_partners_content():
    rows = _admin_collection_items("partners", DEFAULT_PARTNERS)
    rows = _enrich_service_media(rows)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("sort_order") or 0, item.get("name") or "")))


@admin_bp.post("/partners")
@login_required
@permission_required("partners.create")
def create_partner_content():
    payload = request.get_json(silent=True) or {}
    if not payload.get("name"):
        return jsonify(error="Partner name is required."), 400
    row, error = _create_admin_collection_item("partners", DEFAULT_PARTNERS, payload, "partner")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0]), 201


@admin_bp.put("/partners/<string:partner_id>")
@login_required
@permission_required("partners.edit")
def update_partner_content(partner_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("partners", DEFAULT_PARTNERS, partner_id, payload, "partner")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0])


@admin_bp.delete("/partners/<string:partner_id>")
@login_required
@permission_required("partners.delete")
def delete_partner_content(partner_id):
    return _delete_admin_collection_item("partners", DEFAULT_PARTNERS, partner_id, "partner")


@admin_bp.get("/people")
@login_required
@permission_required("people.view")
def list_people_content():
    rows = _admin_collection_items("people", DEFAULT_PEOPLE)
    rows = _enrich_service_media(rows)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("sort_order") or 0, item.get("name") or "")))


@admin_bp.post("/people")
@login_required
@permission_required("people.create")
def create_people_content():
    payload = request.get_json(silent=True) or {}
    if not payload.get("name"):
        return jsonify(error="Person name is required."), 400
    row, error = _create_admin_collection_item("people", DEFAULT_PEOPLE, payload, "person")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0]), 201


@admin_bp.put("/people/<string:person_id>")
@login_required
@permission_required("people.edit")
def update_people_content(person_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("people", DEFAULT_PEOPLE, person_id, payload, "person")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0])


@admin_bp.delete("/people/<string:person_id>")
@login_required
@permission_required("people.delete")
def delete_people_content(person_id):
    return _delete_admin_collection_item("people", DEFAULT_PEOPLE, person_id, "person")


@admin_bp.get("/testimonials")
@login_required
@permission_required("testimonials.view")
def list_testimonials_content():
    rows = _admin_collection_items("testimonials", DEFAULT_TESTIMONIALS)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("sort_order") or 0, item.get("name") or "")))


@admin_bp.post("/testimonials")
@login_required
@permission_required("testimonials.create")
def create_testimonial_content():
    payload = request.get_json(silent=True) or {}
    if not payload.get("quote"):
        return jsonify(error="Testimonial quote is required."), 400
    if not payload.get("name"):
        return jsonify(error="Client name is required."), 400
    row, error = _create_admin_collection_item("testimonials", DEFAULT_TESTIMONIALS, payload, "testimonial")
    if error:
        return error
    return jsonify(row), 201


@admin_bp.put("/testimonials/<string:testimonial_id>")
@login_required
@permission_required("testimonials.edit")
def update_testimonial_content(testimonial_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("testimonials", DEFAULT_TESTIMONIALS, testimonial_id, payload, "testimonial")
    if error:
        return error
    return jsonify(row)


@admin_bp.delete("/testimonials/<string:testimonial_id>")
@login_required
@permission_required("testimonials.delete")
def delete_testimonial_content(testimonial_id):
    return _delete_admin_collection_item("testimonials", DEFAULT_TESTIMONIALS, testimonial_id, "testimonial")


@admin_bp.get("/home-hero-slides")
@login_required
@permission_required("homeHeroSlides.view")
def list_home_hero_slides():
    rows = _admin_collection_items("home_hero_slides", DEFAULT_HOME_HERO_SLIDES)
    rows = _enrich_service_media(rows)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("sort_order") or 0, item.get("label") or "")))


@admin_bp.post("/home-hero-slides")
@login_required
@permission_required("homeHeroSlides.create")
def create_home_hero_slide():
    payload = request.get_json(silent=True) or {}
    if not payload.get("label"):
        return jsonify(error="Slide label is required."), 400
    row, error = _create_admin_collection_item("home_hero_slides", DEFAULT_HOME_HERO_SLIDES, payload, "home_hero_slide")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0]), 201


@admin_bp.put("/home-hero-slides/<string:slide_id>")
@login_required
@permission_required("homeHeroSlides.edit")
def update_home_hero_slide(slide_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("home_hero_slides", DEFAULT_HOME_HERO_SLIDES, slide_id, payload, "home_hero_slide")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0])


@admin_bp.delete("/home-hero-slides/<string:slide_id>")
@login_required
@permission_required("homeHeroSlides.delete")
def delete_home_hero_slide(slide_id):
    return _delete_admin_collection_item("home_hero_slides", DEFAULT_HOME_HERO_SLIDES, slide_id, "home_hero_slide")


@admin_bp.get("/donation-slides")
@login_required
@permission_required("donationSlides.view")
def list_donation_slides():
    rows = _admin_collection_items("donation_slides", DEFAULT_DONATION_SLIDES)
    rows = _enrich_service_media(rows)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("sort_order") or 0, item.get("label") or "")))


@admin_bp.post("/donation-slides")
@login_required
@permission_required("donationSlides.create")
def create_donation_slide():
    payload = request.get_json(silent=True) or {}
    if not payload.get("label"):
        return jsonify(error="Slide label is required."), 400
    row, error = _create_admin_collection_item("donation_slides", DEFAULT_DONATION_SLIDES, payload, "donation_slide")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0]), 201


@admin_bp.put("/donation-slides/<string:slide_id>")
@login_required
@permission_required("donationSlides.edit")
def update_donation_slide(slide_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("donation_slides", DEFAULT_DONATION_SLIDES, slide_id, payload, "donation_slide")
    if error:
        return error
    return jsonify(_enrich_service_media([row])[0])


@admin_bp.delete("/donation-slides/<string:slide_id>")
@login_required
@permission_required("donationSlides.delete")
def delete_donation_slide(slide_id):
    return _delete_admin_collection_item("donation_slides", DEFAULT_DONATION_SLIDES, slide_id, "donation_slide")


@admin_bp.get("/site-copy")
@login_required
@permission_required("siteCopy.view")
def list_site_copy_content():
    rows = _admin_collection_items("site_copy", DEFAULT_SITE_COPY)
    return jsonify(items=sorted(rows, key=lambda item: (item.get("area") or "", item.get("label") or "")))


@admin_bp.post("/site-copy")
@login_required
@permission_required("siteCopy.create")
def create_site_copy_content():
    payload = request.get_json(silent=True) or {}
    if not payload.get("key"):
        return jsonify(error="Copy key is required."), 400
    payload["id"] = payload.get("id") or payload["key"]
    row, error = _create_admin_collection_item("site_copy", DEFAULT_SITE_COPY, payload, "site_copy")
    if error:
        return error
    return jsonify(row), 201


@admin_bp.put("/site-copy/<string:copy_id>")
@login_required
@permission_required("siteCopy.edit")
def update_site_copy_content(copy_id):
    payload = request.get_json(silent=True) or {}
    row, error = _update_admin_collection_item("site_copy", DEFAULT_SITE_COPY, copy_id, payload, "site_copy")
    if error:
        return error
    return jsonify(row)


@admin_bp.delete("/site-copy/<string:copy_id>")
@login_required
@permission_required("siteCopy.delete")
def delete_site_copy_content(copy_id):
    return _delete_admin_collection_item("site_copy", DEFAULT_SITE_COPY, copy_id, "site_copy")


SETTING_SCOPES = {"all", "main", "bookshop"}


def _split_setting_key(storage_key):
    for scope in ("main", "bookshop"):
        prefix = f"{scope}__"
        if storage_key.startswith(prefix):
            return scope, storage_key[len(prefix):]
    return "all", storage_key


def _storage_setting_key(key, scope):
    clean_key = (key or "").strip()
    clean_scope = (scope or "all").strip().lower()
    if not clean_key:
        raise ValueError("Detail Name is required.")
    if "__" in clean_key:
        raise ValueError("Detail Name cannot contain double underscores.")
    if clean_scope not in SETTING_SCOPES:
        raise ValueError("Choose Both sites, Main website only, or Bookshop only.")
    return clean_key if clean_scope == "all" else f"{clean_scope}__{clean_key}"


@admin_bp.get("/settings")
@login_required
@permission_required("settings.view")
def list_settings():
    rows = SiteSetting.query.order_by(SiteSetting.key.asc()).all()
    items = []
    for row in rows:
        scope, display_key = _split_setting_key(row.key)
        items.append({
            "id": row.key,
            "database_id": row.id,
            "key": display_key,
            "site_scope": scope,
            "value": row.value,
            "public": row.public,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        })
    return jsonify(items=items)


@admin_bp.get("/flyers")
@login_required
@permission_required("flyers.view")
def list_flyers():
    rows = Flyer.query.order_by(Flyer.sort_order.asc(), Flyer.created_at.asc()).all()
    def _img(r):
        return f"/uploads/{r.image_file.visibility}/{r.image_file.category}/{r.image_file.stored_filename}" if r.image_file else None
    return jsonify(items=[{
        "id": r.id,
        "headline": r.headline,
        "accent": r.accent,
        "subline": r.subline,
        "badge": r.badge,
        "sort_order": r.sort_order,
        "show_overlay": r.show_overlay,
        "image_fit": r.image_fit,
        "image_position": r.image_position,
        "image_url": _img(r),
        "image_file_id": r.image_file_id,
        "status": r.status,
        "is_focus": r.is_focus,
    } for r in rows])


@admin_bp.post("/flyers")
@login_required
@permission_required("flyers.create")
def create_flyer():
    payload = request.get_json(silent=True) or {}
    row = Flyer(
        headline=(payload.get("headline") or "").strip() or None,
        accent=(payload.get("accent") or "").strip() or None,
        subline=(payload.get("subline") or "").strip() or None,
        badge=(payload.get("badge") or "").strip() or None,
        sort_order=int(payload.get("sort_order") or 0),
        image_file_id=payload.get("image_file_id") or None,
        show_overlay=bool(payload.get("show_overlay", False)),
        image_fit=payload.get("image_fit") or "cover",
        image_position=payload.get("image_position") or "center",
        status=payload.get("status") or "published",
        is_focus=bool(payload.get("is_focus", False)),
    )
    if not row.headline and not row.image_file_id:
        return jsonify(error="Add headline text, an image, or both."), 400
    if row.is_focus:
        Flyer.query.update({Flyer.is_focus: False}, synchronize_session=False)
    db.session.add(row)
    db.session.flush()
    log_action("create_flyer", "flyer", row.id)
    db.session.commit()
    return jsonify(id=row.id, headline=row.headline, status=row.status), 201


@admin_bp.put("/flyers/<int:flyer_id>")
@login_required
@permission_required("flyers.edit")
def update_flyer(flyer_id):
    row = db.get_or_404(Flyer, flyer_id)
    payload = request.get_json(silent=True) or {}
    for field in ["headline", "accent", "subline", "badge", "sort_order", "image_file_id", "show_overlay", "image_fit", "image_position", "status", "is_focus"]:
        if field in payload:
            value = payload[field]
            if field in {"headline", "accent", "subline", "badge", "image_file_id"}:
                value = (str(value).strip() if value is not None else "") or None
            setattr(row, field, value)
    if row.is_focus:
        Flyer.query.filter(Flyer.id != row.id).update({Flyer.is_focus: False}, synchronize_session=False)
    if not row.headline and not row.image_file_id:
        return jsonify(error="Add headline text, an image, or both."), 400
    log_action("update_flyer", "flyer", row.id)
    db.session.commit()
    return jsonify(id=row.id, headline=row.headline, status=row.status)


@admin_bp.delete("/flyers/<int:flyer_id>")
@login_required
@permission_required("flyers.delete")
def delete_flyer(flyer_id):
    row = db.get_or_404(Flyer, flyer_id)
    log_action("delete_flyer", "flyer", row.id, {"headline": row.headline})
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="Flyer deleted.")


# ============================================================
# PROMO CODES
# ============================================================

@admin_bp.get("/promo-codes")
@login_required
@permission_required("priceAdjustment.view")
def list_promo_codes():
    rows = PromoCode.query.order_by(PromoCode.created_at.desc()).all()
    return jsonify(items=[_promo_json(r) for r in rows])


def _apply_promo_payload(row, payload):
    discount_type = (payload.get("discount_type", row.discount_type or "percentage") or "percentage").strip().lower()
    applies_to = (payload.get("applies_to", row.applies_to or "products") or "products").strip().lower()
    if discount_type not in {"percentage", "fixed"}:
        raise ValueError("Discount type must be percentage or fixed.")
    if applies_to not in {"products", "delivery", "all"}:
        raise ValueError("Promo codes can apply to products, delivery, or all.")

    discount_value = _decimalish(payload.get("discount_value"), float(row.discount_value or 0))
    min_order_amount = _decimalish(payload.get("min_order_amount"), float(row.min_order_amount or 0))
    commission_percent = _decimalish(
        payload.get("affiliate_commission_percent"),
        float(getattr(row, "affiliate_commission_percent", 0) or 0),
    )
    if discount_value < 0:
        raise ValueError("Discount value cannot be negative.")
    if discount_type == "percentage" and discount_value > 100:
        raise ValueError("Percentage discount cannot exceed 100%.")
    if min_order_amount < 0:
        raise ValueError("Minimum order amount cannot be negative.")
    if commission_percent < 0 or commission_percent > 100:
        raise ValueError("Affiliate commission percent must be between 0 and 100.")

    max_uses_value = payload.get("max_uses", row.max_uses)
    max_uses = None if max_uses_value in {None, ""} else max(1, _intish(max_uses_value, 0))
    row.description = (payload.get("description", row.description) or "").strip() or None
    row.discount_type = discount_type
    row.discount_value = discount_value
    row.applies_to = applies_to
    row.min_order_amount = min_order_amount
    row.max_uses = max_uses
    if "valid_from" in payload:
        row.valid_from = _dateish(payload.get("valid_from"))
    if "valid_until" in payload:
        row.valid_until = _dateish(payload.get("valid_until"))
    if row.valid_from and row.valid_until and row.valid_from > row.valid_until:
        raise ValueError("Valid From cannot be later than Valid Until.")
    if "is_active" in payload:
        row.is_active = _boolish(payload.get("is_active"))

    row.affiliate_name = (payload.get("affiliate_name", row.affiliate_name) or "").strip() or None
    row.affiliate_email = (payload.get("affiliate_email", row.affiliate_email) or "").strip().lower() or None
    row.affiliate_phone = (payload.get("affiliate_phone", row.affiliate_phone) or "").strip() or None
    row.affiliate_commission_percent = commission_percent
    if "affiliate_notify_on_use" in payload:
        row.affiliate_notify_on_use = _boolish(payload.get("affiliate_notify_on_use"))


@admin_bp.post("/promo-codes")
@login_required
@permission_required("priceAdjustment.edit")
def create_promo_code():
    payload = request.get_json(silent=True) or {}
    code = (payload.get("code") or "").strip().upper()
    if not code:
        return jsonify(error="Code is required."), 400
    if PromoCode.query.filter_by(code=code).first():
        return jsonify(error="A promo code with that name already exists."), 409
    row = PromoCode(code=code, is_active=True)
    try:
        _apply_promo_payload(row, payload)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    db.session.add(row)
    db.session.flush()
    log_action("create_promo_code", "promo_code", row.id, {"code": row.code})
    db.session.commit()
    return jsonify(promo_code=_promo_json(row)), 201


@admin_bp.put("/promo-codes/<int:promo_id>")
@login_required
@permission_required("priceAdjustment.edit")
def update_promo_code(promo_id):
    row = db.get_or_404(PromoCode, promo_id)
    payload = request.get_json(silent=True) or {}
    try:
        _apply_promo_payload(row, payload)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    log_action("update_promo_code", "promo_code", row.id, {"code": row.code})
    db.session.commit()
    return jsonify(promo_code=_promo_json(row))


@admin_bp.delete("/promo-codes/<int:promo_id>")
@login_required
@permission_required("priceAdjustment.edit")
def delete_promo_code(promo_id):
    row = db.get_or_404(PromoCode, promo_id)
    log_action("delete_promo_code", "promo_code", row.id, {"code": row.code})
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="Promo code deleted.")


def _promo_json(r):
    return {
        "id": r.id, "code": r.code, "description": r.description,
        "discount_type": r.discount_type, "discount_value": float(r.discount_value or 0),
        "applies_to": r.applies_to, "min_order_amount": float(r.min_order_amount or 0),
        "max_uses": r.max_uses, "uses_count": r.uses_count,
        "valid_from": str(r.valid_from) if r.valid_from else None,
        "valid_until": str(r.valid_until) if r.valid_until else None,
        "is_active": r.is_active,
        "affiliate_name": getattr(r, "affiliate_name", None),
        "affiliate_email": getattr(r, "affiliate_email", None),
        "affiliate_phone": getattr(r, "affiliate_phone", None),
        "affiliate_commission_percent": float(getattr(r, "affiliate_commission_percent", 0) or 0),
        "affiliate_notify_on_use": bool(getattr(r, "affiliate_notify_on_use", True)),
        "completed_affiliate_uses": len(getattr(r, "usages", []) or []),
        "created_at": r.created_at.isoformat(),
    }


# ============================================================
# BULK PRICE ADJUSTMENT
# ============================================================

@admin_bp.post("/products/bulk-price-adjust")
@login_required
@permission_required("priceAdjustment.edit")
def bulk_price_adjust():
    """Apply a percentage or fixed amount increase/decrease to ALL active product prices."""
    payload = request.get_json(silent=True) or {}
    adjustment_type = payload.get("type") or "percentage"   # percentage | fixed
    value = float(payload.get("value") or 0)
    direction = payload.get("direction") or "decrease"       # increase | decrease
    if value <= 0:
        return jsonify(error="Value must be greater than zero."), 400

    products = Product.query.filter_by(is_active=True).all()
    count = 0
    for p in products:
        old_price = float(p.price or 0)
        if adjustment_type == "percentage":
            delta = old_price * (value / 100)
        else:
            delta = value
        if direction == "increase":
            new_price = old_price + delta
        else:
            new_price = max(0.01, old_price - delta)
        p.price = round(new_price, 2)
        count += 1
    log_action("bulk_price_adjust", "product", None, {
        "type": adjustment_type, "value": value, "direction": direction, "count": count,
    })
    db.session.commit()
    return jsonify(message=f"Updated {count} products.", count=count)


# ============================================================
# DELIVERY FEE BULK ADJUST
# ============================================================

@admin_bp.post("/delivery-zones/bulk-adjust")
@login_required
@permission_required("priceAdjustment.edit")
def bulk_delivery_adjust():
    """Apply a percentage or fixed amount increase/decrease to ALL active delivery zone fees."""
    from ..models import DeliveryZone
    payload = request.get_json(silent=True) or {}
    adjustment_type = payload.get("type") or "percentage"
    value = float(payload.get("value") or 0)
    direction = payload.get("direction") or "increase"
    if value <= 0:
        return jsonify(error="Value must be greater than zero."), 400

    zones = DeliveryZone.query.filter_by(is_active=True).all()
    count = 0
    for z in zones:
        if float(z.fee or 0) == 0:
            continue  # skip free zones
        old_fee = float(z.fee)
        if adjustment_type == "percentage":
            delta = old_fee * (value / 100)
        else:
            delta = value
        if direction == "increase":
            new_fee = old_fee + delta
        else:
            new_fee = max(0, old_fee - delta)
        z.fee = round(new_fee, 2)
        count += 1
    log_action("bulk_delivery_adjust", "delivery_zone", None, {
        "type": adjustment_type, "value": value, "direction": direction, "count": count,
    })
    db.session.commit()
    return jsonify(message=f"Updated {count} delivery zones.", count=count)


@admin_bp.put("/settings/<string:key>")
@login_required
@permission_required("settings.edit")
def upsert_setting(key):
    payload = request.get_json(silent=True) or {}
    row = SiteSetting.query.filter_by(key=key).first()
    display_key = payload.get("key")
    if display_key is None:
        _, display_key = _split_setting_key(key)
    scope = payload.get("site_scope")
    if scope is None:
        scope, _ = _split_setting_key(key)
    try:
        storage_key = _storage_setting_key(display_key, scope)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    conflict = SiteSetting.query.filter_by(key=storage_key).first()
    if conflict is not None and conflict is not row:
        return jsonify(error="That detail already exists for the selected website."), 409
    if not row:
        row = SiteSetting(key=storage_key)
        db.session.add(row)
    else:
        row.key = storage_key
    row.value = payload.get("value")
    row.public = bool(payload.get("public", row.public))
    log_action("upsert_setting", "site_setting", storage_key, {"scope": scope, "detail": display_key})
    db.session.commit()
    return jsonify(id=row.key, key=display_key, site_scope=scope, value=row.value, public=row.public)


@admin_bp.delete("/settings/<string:key>")
@login_required
@permission_required("settings.delete")
def delete_setting(key):
    row = SiteSetting.query.filter_by(key=key).first_or_404()
    log_action("delete_setting", "site_setting", key)
    db.session.delete(row)
    db.session.commit()
    return jsonify(message="Setting deleted.")
